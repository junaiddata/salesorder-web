"""
Brandwise Sales Analysis
========================
A pivot table of net sales with:
  - Rows    = brands (Items.item_firm)
  - Columns = months (Jan..Dec) of the selected year, plus a row/column total
  - Filters = Year, Store (HO / Others / Total), Salesman (multi-select),
              Brand (multi-select; optional — empty means all brands)

Net sales = AR Invoice line values + AR Credit Memo line values (credit memos net
negative, so they reduce sales) — the same "sales" definition used on the home page.
Sales value per line uses ``line_total_after_discount`` (LineTotal after the header
discount is applied). Brand comes from the linked Items.item_firm.
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, DecimalField, Case, When, Q, F, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render

from .models import (
    Items,
    SAPARInvoice,
    SAPARInvoiceItem,
    SAPARCreditMemo,
    SAPARCreditMemoItem,
)
from .sap_salesorder_views import salesman_scope_q_salesorder

MONTH_NAMES_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _net_value_expr():
    """Per-line net sales value, matching the trusted item-sold analysis convention.

    Use ``line_total_after_discount`` only when it is populated and non-zero;
    otherwise fall back to the raw ``line_total``. This avoids under-counting when
    ``line_total_after_discount`` is NULL/0 on some synced rows (which otherwise
    makes invoice rows vanish and leaves only negative credit-memo rows).
    Credit-memo line values are stored negative, so summing invoices + credit
    memos yields net sales — the same definition used on the home page.
    """
    return Sum(
        Case(
            When(
                Q(line_total_after_discount__isnull=False) & ~Q(line_total_after_discount=0),
                then=F('line_total_after_discount'),
            ),
            default=F('line_total'),
            output_field=DecimalField(),
        )
    )


def _gp_value_expr():
    """Per-line gross profit. Credit-memo GP is stored negative, so summing
    invoices + credit memos yields net GP — mirrors the net-sales convention."""
    return Sum(Coalesce(F('gross_profit'), Value(0, output_field=DecimalField())))


def _pct(part, whole):
    """Safe percentage: part / whole * 100, rounded to 1 dp."""
    w = float(whole or 0)
    return round(float(part or 0) / w * 100, 1) if w else 0


def _user_is_admin(user):
    """Admin = superuser / staff / 'manager' / Role == 'Admin'. GP is admin-only,
    matching the home page. Guarded so a missing Role row can't raise."""
    try:
        if user.is_superuser or user.is_staff:
            return True
        if (user.username or '').strip().lower() == 'manager':
            return True
        role = getattr(user, 'role', None)
        if role is not None and getattr(role, 'role', None) == 'Admin':
            return True
    except Exception:
        pass
    return False


def _compute_brandwise_sales(request):
    """Shared computation for the HTML page and the PDF/Excel exports.

    Returns the full context dict (filters, brand rows, month/grand totals,
    and the option lists for the filter widgets).
    """
    today = date.today()
    current_year = today.year

    # ── Year filter ──────────────────────────────────────────────
    calendar_years = list(range(2024, current_year + 2))
    try:
        selected_year = int(request.GET.get('year', '').strip() or current_year)
    except (ValueError, TypeError):
        selected_year = current_year
    if selected_year not in calendar_years:
        selected_year = current_year

    # ── Store filter (HO / Others / Total) — default HO, same as home page ──
    store_filter = (request.GET.get('store', 'HO') or 'HO').strip()
    if store_filter not in ('HO', 'Others', 'Total'):
        store_filter = 'HO'

    # ── Month range filter (From / To) — default Jan..Dec (whole year) ──
    def _parse_month(val, default):
        try:
            m = int(str(val).strip())
        except (ValueError, TypeError):
            return default
        return m if 1 <= m <= 12 else default

    from_month = _parse_month(request.GET.get('from_month'), 1)
    to_month = _parse_month(request.GET.get('to_month'), 12)
    if from_month > to_month:                       # swap if entered backwards
        from_month, to_month = to_month, from_month
    month_range = list(range(from_month - 1, to_month))   # 0-based indices in range

    # ── Salesman filter (multi-select) ───────────────────────────
    selected_salesmen = [s.strip() for s in request.GET.getlist('salesman') if s.strip()]

    # ── Brand filter (multi-select) — optional; empty = all brands ──
    selected_brands = list(dict.fromkeys(
        b.strip() for b in request.GET.getlist('brand') if b and b.strip()
    ))

    # ── Scoped header querysets (respect per-user salesman scope) ─
    scope_q = salesman_scope_q_salesorder(request.user)
    inv_headers = (
        SAPARInvoice.objects.filter(scope_q)
        .exclude(salesman_name__iexact='Z.DUTY')
        .filter(posting_date__year=selected_year)
    )
    cm_headers = (
        SAPARCreditMemo.objects.filter(scope_q)
        .exclude(salesman_name__iexact='Z.DUTY')
        .filter(posting_date__year=selected_year)
    )
    if store_filter in ('HO', 'Others'):
        inv_headers = inv_headers.filter(store=store_filter)
        cm_headers = cm_headers.filter(store=store_filter)
    if selected_salesmen:
        inv_headers = inv_headers.filter(salesman_name__in=selected_salesmen)
        cm_headers = cm_headers.filter(salesman_name__in=selected_salesmen)

    is_admin = _user_is_admin(request.user)

    # ── Aggregate sales + GP by brand + month ────────────────────
    # brand -> {'sales': [12], 'gp': [12]}
    grid = defaultdict(lambda: {'sales': [Decimal('0')] * 12, 'gp': [Decimal('0')] * 12})

    def _accumulate(rows, brand_key, month_key):
        for r in rows:
            brand = (r[brand_key] or '').strip() or 'Unknown'
            month = r[month_key]
            if not month:
                continue
            grid[brand]['sales'][month - 1] += r['sales'] or Decimal('0')
            grid[brand]['gp'][month - 1] += r['gp'] or Decimal('0')

    inv_items = SAPARInvoiceItem.objects.filter(invoice__in=inv_headers)
    cm_items = SAPARCreditMemoItem.objects.filter(credit_memo__in=cm_headers)
    if selected_brands:
        inv_items = inv_items.filter(item__item_firm__in=selected_brands)
        cm_items = cm_items.filter(item__item_firm__in=selected_brands)

    inv_rows = (
        inv_items
        .values('item__item_firm', 'invoice__posting_date__month')
        .annotate(sales=_net_value_expr(), gp=_gp_value_expr())
    )
    _accumulate(inv_rows, 'item__item_firm', 'invoice__posting_date__month')

    cm_rows = (
        cm_items
        .values('item__item_firm', 'credit_memo__posting_date__month')
        .annotate(sales=_net_value_expr(), gp=_gp_value_expr())
    )
    _accumulate(cm_rows, 'item__item_firm', 'credit_memo__posting_date__month')

    # ── Build table rows (highest sales first) ───────────────────
    # Row/column/grand totals are summed over the selected month range only,
    # and only the in-range month cells are shown.
    brand_rows = []
    for brand, d in grid.items():
        sales_months = d['sales']
        gp_months = d['gp']
        row_total = sum((sales_months[i] for i in month_range), Decimal('0'))
        gp_total = sum((gp_months[i] for i in month_range), Decimal('0'))
        cells = [
            {
                'sales': sales_months[i],
                'gp': gp_months[i],
                'gp_pct': _pct(gp_months[i], sales_months[i]),
            }
            for i in month_range
        ]
        brand_rows.append({
            'brand': brand,
            'months': sales_months,      # full 12 (kept for totals/exports)
            'gp_months': gp_months,
            'cells': cells,
            'total': row_total,
            'gp_total': gp_total,
            'gp_pct': _pct(gp_total, row_total),
        })
    brand_rows.sort(key=lambda r: r['total'], reverse=True)

    # ── Column (month) totals + grand totals (range-aware) ───────
    month_totals = [sum((r['months'][i] for r in brand_rows), Decimal('0')) for i in range(12)]
    gp_month_totals = [sum((r['gp_months'][i] for r in brand_rows), Decimal('0')) for i in range(12)]
    grand_total = sum((month_totals[i] for i in month_range), Decimal('0'))
    gp_grand_total = sum((gp_month_totals[i] for i in month_range), Decimal('0'))
    grand_gp_pct = _pct(gp_grand_total, grand_total)

    # Zip month name + sales total + gp total + gp% for the header/footer row
    # (only the in-range months)
    month_headers = [
        {'name': MONTH_NAMES_SHORT[i], 'sales': month_totals[i],
         'gp': gp_month_totals[i], 'gp_pct': _pct(gp_month_totals[i], month_totals[i])}
        for i in month_range
    ]

    # ── Salesman list for the filter dropdown (scoped) ───────────
    salesmen = list(
        SAPARInvoice.objects.filter(scope_q)
        .exclude(salesman_name__isnull=True)
        .exclude(salesman_name='')
        .exclude(salesman_name__iexact='Z.DUTY')
        .values_list('salesman_name', flat=True)
        .distinct()
        .order_by('salesman_name')
    )

    # ── Brand list for the filter dropdown (all firms in Items) ──
    brands = list(
        Items.objects.exclude(item_firm__isnull=True)
        .exclude(item_firm='')
        .values_list('item_firm', flat=True)
        .distinct()
        .order_by('item_firm')
    )

    all_months = [{'num': i + 1, 'name': MONTH_NAMES_SHORT[i]} for i in range(12)]
    period_label = MONTH_NAMES_SHORT[from_month - 1] if from_month == to_month else \
        f"{MONTH_NAMES_SHORT[from_month - 1]} – {MONTH_NAMES_SHORT[to_month - 1]}"

    context = {
        'month_names': MONTH_NAMES_SHORT,
        'month_headers': month_headers,
        'all_months': all_months,
        'from_month': from_month,
        'to_month': to_month,
        'month_range': month_range,
        'period_label': period_label,
        'brand_rows': brand_rows,
        'month_totals': month_totals,
        'gp_month_totals': gp_month_totals,
        'grand_total': grand_total,
        'gp_grand_total': gp_grand_total,
        'grand_gp_pct': grand_gp_pct,
        'brand_count': len(brand_rows),
        'is_admin': is_admin,
        'calendar_years': calendar_years,
        'selected_year': selected_year,
        'store_filter': store_filter,
        'salesmen': salesmen,
        'selected_salesmen': selected_salesmen,
        'brands': brands,
        'selected_brands': selected_brands,
    }
    return context


@login_required
def brandwise_sales_analysis(request):
    context = _compute_brandwise_sales(request)
    return render(request, 'salesorders/brandwise_sales_analysis.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _filter_label(ctx):
    """Human-readable one-line summary of the active filters (for file names / subtitles)."""
    parts = [str(ctx['selected_year']), ctx['store_filter']]
    fm, tm = ctx['from_month'], ctx['to_month']
    if not (fm == 1 and tm == 12):                  # only note a partial range
        rng = MONTH_NAMES_SHORT[fm - 1] if fm == tm else \
            f"{MONTH_NAMES_SHORT[fm - 1]}–{MONTH_NAMES_SHORT[tm - 1]}"
        parts.append(rng)
    if ctx['selected_salesmen']:
        parts.append(f"{len(ctx['selected_salesmen'])} salesman" if len(ctx['selected_salesmen']) == 1
                     else f"{len(ctx['selected_salesmen'])} salesmen")
    if ctx['selected_brands']:
        parts.append(f"{len(ctx['selected_brands'])} brand" if len(ctx['selected_brands']) == 1
                     else f"{len(ctx['selected_brands'])} brands")
    return ' · '.join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_brandwise_sales_analysis_excel(request):
    from io import BytesIO
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ctx = _compute_brandwise_sales(request)
    month_names = ctx['month_names']
    month_range = ctx['month_range']
    brand_rows = ctx['brand_rows']
    month_totals = ctx['month_totals']
    gp_month_totals = ctx['gp_month_totals']
    is_admin = ctx['is_admin']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Brandwise Sales'

    navy = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    total_fill = PatternFill(start_color='EBF0F7', end_color='EBF0F7', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8F9FB', end_color='F8F9FB', fill_type='solid')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    bold = Font(bold=True, size=10)
    gp_font = Font(size=9, color='059669')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = '#,##0'

    # Column layout: Brand [| Metric] <range months> Total  (Metric column only for admins)
    n_months = len(month_range)
    metric_col = 1 if is_admin else 0
    month_start = 2 + metric_col          # first month column index
    total_col = month_start + n_months    # Total column index
    n_cols = total_col

    # Title + filter summary
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tcell = ws.cell(row=1, column=1, value='Brandwise Sales Analysis')
    tcell.font = Font(bold=True, size=14, color='1B2A4A')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    scell = ws.cell(row=2, column=1, value=f"{_filter_label(ctx)}   ·   Generated {datetime.now().strftime('%d %b %Y, %H:%M')}")
    scell.font = Font(italic=True, size=9, color='6B7280')

    # Header row
    header_row = 4
    month_names_range = [month_names[i] for i in month_range]
    headers = ['Brand'] + (['Metric'] if is_admin else []) + month_names_range + ['Total']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = navy
        cell.font = white_bold
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left', vertical='center')
        cell.border = border

    def _write_line(r, brand_label, metric_label, values, total_value, is_gp=False, fill=None):
        """Write one data line across the row. `values` = full list of 12 monthly
        numbers; only the in-range months are written."""
        ws.cell(row=r, column=1, value=brand_label).border = border
        if is_admin:
            mcell = ws.cell(row=r, column=2, value=metric_label)
            mcell.border = border
            mcell.font = gp_font if is_gp else Font(size=9, color='6B7280')
        for pos, m in enumerate(month_range):
            v = float(values[m])
            cell = ws.cell(row=r, column=month_start + pos, value=v if v else None)
            cell.number_format = num_fmt
            cell.border = border
            if is_gp:
                cell.font = gp_font
        tcell = ws.cell(row=r, column=total_col, value=float(total_value))
        tcell.number_format = num_fmt
        tcell.border = border
        tcell.font = gp_font if is_gp else bold
        if fill is not None:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill

    # Data rows
    r = header_row + 1
    for idx, row in enumerate(brand_rows):
        fill = zebra_fill if idx % 2 == 1 else None
        _write_line(r, row['brand'], 'Sales', row['months'], row['total'], is_gp=False, fill=fill)
        r += 1
        if is_admin:
            _write_line(r, '', 'GP', row['gp_months'], row['gp_total'], is_gp=True, fill=fill)
            r += 1

    # Totals row(s)
    _write_line(r, 'TOTAL', 'Sales', month_totals, ctx['grand_total'], is_gp=False, fill=total_fill)
    for c in range(1, n_cols + 1):
        ws.cell(row=r, column=c).font = bold
    r += 1
    if is_admin:
        _write_line(r, '', 'GP', gp_month_totals, ctx['gp_grand_total'], is_gp=True, fill=total_fill)
        r += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    if is_admin:
        ws.column_dimensions[get_column_letter(2)].width = 8
    for c in range(month_start, total_col):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 14

    # Freeze header + brand column
    ws.freeze_panes = ws.cell(row=header_row + 1, column=month_start)
    ws.row_dimensions[header_row].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f"Brandwise_Sales_{ctx['selected_year']}_{ctx['store_filter']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_brandwise_sales_analysis_pdf(request):
    from io import BytesIO
    from datetime import datetime

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    from .finance_statement_pdf_export import (
        _build_document_header, _build_kpi_bar, _build_styles, _fmt,
        CLR_WHITE, CLR_BG_TOTAL, CLR_BG_ZEBRA, CLR_BORDER, CLR_BORDER_HEAVY, CLR_PRIMARY, CLR_TEXT_FAINT,
    )
    from reportlab.lib.colors import HexColor

    ctx = _compute_brandwise_sales(request)
    month_names = ctx['month_names']
    month_range = ctx['month_range']
    brand_rows = ctx['brand_rows']
    month_totals = ctx['month_totals']
    gp_month_totals = ctx['gp_month_totals']
    is_admin = ctx['is_admin']

    styles = _build_styles()
    base = getSampleStyleSheet()['Normal']
    th = ParagraphStyle('bwTH', parent=base, fontName='Helvetica-Bold', fontSize=6.5,
                        textColor=CLR_WHITE, alignment=TA_CENTER, leading=8)
    th_l = ParagraphStyle('bwTHL', parent=base, fontName='Helvetica-Bold', fontSize=6.5,
                          textColor=CLR_WHITE, alignment=TA_LEFT, leading=8)
    td = ParagraphStyle('bwTD', parent=base, fontName='Helvetica', fontSize=6.5,
                        textColor=HexColor('#1F2937'), alignment=TA_RIGHT, leading=8)
    td_l = ParagraphStyle('bwTDL', parent=base, fontName='Helvetica', fontSize=6.5,
                          textColor=HexColor('#1F2937'), alignment=TA_LEFT, leading=8)
    td_faint = ParagraphStyle('bwTDF', parent=base, fontName='Helvetica', fontSize=6.5,
                              textColor=CLR_TEXT_FAINT, alignment=TA_RIGHT, leading=8)
    td_tot = ParagraphStyle('bwTDT', parent=base, fontName='Helvetica-Bold', fontSize=6.5,
                            textColor=CLR_PRIMARY, alignment=TA_RIGHT, leading=8)
    td_tot_l = ParagraphStyle('bwTDTL', parent=base, fontName='Helvetica-Bold', fontSize=6.5,
                              textColor=CLR_PRIMARY, alignment=TA_LEFT, leading=8)

    def _gp_line(gp, pct):
        """Second (green) line inside a cell showing GP and GP%."""
        return f"<br/><font size=5 color='#059669'>GP {float(gp or 0):,.0f} · {pct:.1f}%</font>"

    def cell(sales, gp=None, gp_pct=None, total=False):
        fv = float(sales or 0)
        style = td_tot if total else td
        if not fv and not total:
            body = "–"
            style = td_faint
        else:
            body = f"{fv:,.0f}"
        if is_admin and (fv or total):
            body += _gp_line(gp, gp_pct or 0)
        return Paragraph(body, style)

    # Build table header
    def _hdr(label):
        return f"{label}{'<br/><font size=5>Sales / GP</font>' if is_admin else ''}"
    header = [Paragraph('Brand', th_l)] + [Paragraph(_hdr(month_names[i]), th) for i in month_range] + [Paragraph(_hdr('Total'), th)]
    data = [header]
    for row in brand_rows:
        line = [Paragraph(str(row['brand']), td_l)]
        line += [cell(c['sales'], c['gp'], c['gp_pct']) for c in row['cells']]
        line.append(cell(row['total'], row['gp_total'], row['gp_pct'], total=True))
        data.append(line)
    # Totals row
    totals_line = [Paragraph('TOTAL', td_tot_l)]
    for m in month_range:
        pct = _pct(gp_month_totals[m], month_totals[m])
        totals_line.append(cell(month_totals[m], gp_month_totals[m], pct, total=True))
    totals_line.append(cell(ctx['grand_total'], ctx['gp_grand_total'], ctx['grand_gp_pct'], total=True))
    data.append(totals_line)

    # Column widths (landscape A4 usable ≈ 10.7in)
    page_size = landscape(A4)
    usable_w = page_size[0] - 0.8 * inch  # left+right margins 0.4 each
    n_months = len(month_range)
    brand_w = 1.5 * inch
    total_w = 0.95 * inch
    month_w = (usable_w - brand_w - total_w) / max(n_months, 1)
    col_widths = [brand_w] + [month_w] * n_months + [total_w]

    n_rows = len(data)
    tstyle = [
        ('BACKGROUND', (0, 0), (-1, 0), CLR_PRIMARY),
        ('BOX', (0, 0), (-1, -1), 0.75, CLR_BORDER_HEAVY),
        ('LINEBELOW', (0, 0), (-1, 0), 1, CLR_BORDER_HEAVY),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, CLR_BORDER),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('LINEAFTER', (0, 0), (0, -1), 0.6, CLR_BORDER_HEAVY),
        ('BACKGROUND', (0, -1), (-1, -1), CLR_BG_TOTAL),
        ('LINEABOVE', (0, -1), (-1, -1), 1.2, CLR_PRIMARY),
    ]
    for i in range(1, n_rows - 1):
        if i % 2 == 0:
            tstyle.append(('BACKGROUND', (0, i), (-1, i), CLR_BG_ZEBRA))

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(tstyle))

    # KPI bar
    kpi_items = [('Total Net Sales', _fmt(ctx['grand_total']))]
    if is_admin:
        kpi_items.append(('Gross Profit', _fmt(ctx['gp_grand_total'])))
        kpi_items.append(('GP %', f"{ctx['grand_gp_pct']:.1f}%"))
    fm, tm = ctx['from_month'], ctx['to_month']
    period = MONTH_NAMES_SHORT[fm - 1] if fm == tm else \
        f"{MONTH_NAMES_SHORT[fm - 1]}–{MONTH_NAMES_SHORT[tm - 1]}"
    kpi_items += [
        ('Brands', str(ctx['brand_count'])),
        ('Year', str(ctx['selected_year'])),
        ('Period', period),
        ('Store', ctx['store_filter']),
    ]

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(CLR_TEXT_FAINT)
        canvas.drawString(0.4 * inch, 0.28 * inch,
                          f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')}")
        canvas.drawRightString(page_size[0] - 0.4 * inch, 0.28 * inch, f"Page {doc_.page}")
        canvas.restoreState()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.4 * inch, bottomMargin=0.5 * inch,
    )
    elements = []
    elements += _build_document_header(styles, 'Brandwise Sales Analysis', _filter_label(ctx), usable_w)
    elements += [_build_kpi_bar(kpi_items, styles, usable_w), Spacer(1, 10)]
    elements.append(table)

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    fname = f"Brandwise_Sales_{ctx['selected_year']}_{ctx['store_filter']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
