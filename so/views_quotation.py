from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Quotation, QuotationItem, Customer, Salesman, Items, CustomerPrice
from .models import Quotation, QuotationItem, CustomerPrice, Customer, Items, Salesman, QuotationLog
from .models import SalesOrder, OrderItem
from .utils import get_client_ip, label_network
from django.db.models import Q
from .utils import parse_device_info
from urllib.parse import quote
from .salesman_mapping import (
    expand_quotation_salesman_picks,
    get_quotation_salesman_canonical_choices_sorted,
    normalize_quotation_salesman_picks_to_canonicals,
)


def _is_manager_account(user):
    return bool(
        user
        and user.is_authenticated
        and (getattr(user, "username", None) or "").strip().lower() == "manager"
    )


@csrf_exempt
@transaction.atomic
def create_quotation(request):
    if request.method == 'POST':
        try:
            # -------------------------
            # 0. Division & Remarks Setup
            # -------------------------
            division = 'JUNAID' # Default
            if request.user.is_authenticated and 'alabama' in request.user.username.lower():
                division = 'ALABAMA'
            
            remarks = request.POST.get('remarks', '').strip()

            # -------------------------
            # 1. Customer handling
            # -------------------------
            customer_id = request.POST.get('customer')
            customer_display_name = request.POST.get('new_customer_name', '').strip()  # Used as display name for CASH customers

            if not customer_id:
                messages.error(request, 'Please select a customer.')
                return redirect('create_quotation')

            # Get salesman FIRST (required)
            salesman_id = request.POST.get('salesman')
            if not salesman_id:
                messages.error(request, 'Salesman is required.')
                return redirect('create_quotation')
                
            salesman = get_object_or_404(Salesman, id=salesman_id)
            customer = get_object_or_404(Customer, id=customer_id)

            # -------------------------
            # 2. Items validation
            # -------------------------
            item_ids = request.POST.getlist('item')
            quantities = request.POST.getlist('quantity')
            prices = request.POST.getlist('price')
            units = request.POST.getlist('unit')

            # Check if we have items
            if not item_ids:
                messages.error(request, 'Please add at least one item.')
                return redirect('create_quotation')

            if len(item_ids) != len(quantities) or len(item_ids) != len(prices) or len(item_ids) != len(units):
                messages.error(request, 'Invalid form data. Please try again.')
                return redirect('create_quotation')

            # -------------------------
            # 3. Create Quotation Object
            # -------------------------
            quotation = Quotation.objects.create(
                customer=customer,
                salesman=salesman,
                division=division,
                remarks=remarks,
                customer_display_name=customer_display_name if customer_display_name else None
            )

            # -------------------------
            # 4. Logging Logic
            # -------------------------
            ip = get_client_ip(request)
            network_label = label_network(ip)
            ua_string = request.META.get('HTTP_USER_AGENT', '')[:500]
            device_type, device_os, device_browser = parse_device_info(ua_string)

            try:
                lat = request.POST.get("location_lat")
                lng = request.POST.get("location_lng")
                lat_val = float(lat) if lat not in (None, "",) else None
                lng_val = float(lng) if lng not in (None, "",) else None
            except ValueError:
                lat_val = None
                lng_val = None

            QuotationLog.objects.create(
                quotation=quotation,
                user=request.user if request.user.is_authenticated else None,
                ip_address=ip,
                user_agent=ua_string,
                device_type=device_type,
                device_os=device_os,
                device_browser=device_browser,
                location_lat=lat_val,
                location_lng=lng_val,
                network_label=network_label,
                device=getattr(request, 'device_obj', None), 
                action="created",
            )

            # -------------------------
            # 5. Process Items
            # -------------------------
            # Batch-fetch all items and customer prices to avoid N+1 queries
            items_map = {
                str(obj.id): obj
                for obj in Items.objects.filter(id__in=item_ids)
            }
            customer_prices_map = {
                str(cp.item_id): cp
                for cp in CustomerPrice.objects.filter(customer=customer, item_id__in=item_ids)
            }

            quotation_items = []
            customer_price_updates = []
            total_amount = 0

            for i, (item_id, qty, price_input, unit) in enumerate(zip(item_ids, quantities, prices, units)):
                try:
                    item = items_map.get(str(item_id))
                    if item is None:
                        raise Items.DoesNotExist(f"Item {item_id} not found")
                    quantity_val = int(qty)
                    unit_val = unit if unit in ['pcs', 'ctn','roll'] else 'pcs'

                    # Automatic price from CustomerPrice or default item price
                    customer_price = customer_prices_map.get(str(item_id))
                    price_val = float(price_input) if price_input else (customer_price.custom_price if customer_price else float(item.item_price))

                    if quantity_val <= 0:
                        messages.error(request, f'Quantity must be positive for item {i+1}.')
                        return redirect('create_quotation')
                    if price_val < 0:
                        messages.error(request, f'Price cannot be negative for item {i+1}.')
                        return redirect('create_quotation')

                    line_total = quantity_val * price_val
                    total_amount += line_total

                    quotation_items.append(QuotationItem(
                        quotation=quotation,
                        item=item,
                        quantity=quantity_val,
                        price=price_val,
                        unit=unit_val,
                        line_total=line_total
                    ))

                    # CustomerPrice update if new price entered
                    if price_input:
                        customer_price_updates.append((customer, item, price_val))

                except (ValueError, Items.DoesNotExist) as e:
                    messages.error(request, f'Invalid data for item {i+1}: {str(e)}')
                    return redirect('create_quotation')

            # Bulk insert quotation items
            QuotationItem.objects.bulk_create(quotation_items)

            # Update CustomerPrice
            for c, it, pr in customer_price_updates:
                CustomerPrice.objects.update_or_create(
                    customer=c,
                    item=it,
                    defaults={'custom_price': pr}
                )

            # Save totals
            quotation.total_amount = total_amount
            quotation.grand_total = total_amount # Add Tax logic here if needed
            quotation.save()

            messages.success(request, f'Quotation {quotation.quotation_number} created successfully!')
            return redirect('view_quotations')

        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('create_quotation')

    else:
       # GET request → render empty form
        customers = Customer.objects.all().order_by('customer_name')
        
        # ---------------------------------------------------------------------
        # START: Salesman Filtering Logic
        # ---------------------------------------------------------------------
        salesmen = Salesman.objects.all()

        # Filter Logic: Only apply if user is logged in AND is NOT a superuser/admin
        if request.user.is_authenticated and not request.user.is_superuser:
            current_username = request.user.username.lower()

            user_salesman_map = {
                'alabamakadhar': ['KADER'],
                'alabamamusharaf': ['MUSHARAF'],   # Multiple allowed
                'alabamaadmin': ['KADER','MUSHARAF','AIJAZ','CASH'],               # Single allowed
                
            }


            if current_username in user_salesman_map:
                target_names = user_salesman_map[current_username]
                
                # Build a complex query: (name contains A) OR (name contains B)
                query = Q()
                for name in target_names:
                    query |= Q(salesman_name__icontains=name)
                
                salesmen = salesmen.filter(query)
        # ---------------------------------------------------------------------
        # END: Salesman Filtering Logic
        # ---------------------------------------------------------------------
        firms = Items.objects.values_list('item_firm', flat=True).distinct().order_by('item_firm')
        return render(request, 'so/quotations/create_quotation.html', {
            'customers': customers,
            'salesmen': salesmen,
            'firms': firms,
        })

from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Quotation, Salesman
from django.db.models import Q
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required


def _inapp_quotation_scope_for_calendar(request):
    """Same visibility as the quotations list, without search/date/status/division filters."""
    qs = Quotation.objects.all()
    if hasattr(request.user, 'role') and request.user.role.role == 'Admin':
        if request.user.username.lower() in ['so', 'manager']:
            pass
        elif getattr(request.user.role, 'company', 'Junaid') == 'Alabama':
            qs = qs.filter(division='ALABAMA')
        else:
            qs = qs.filter(division='JUNAID')
    elif request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
        from .views import SALES_USER_MAP
        current_username = (request.user.username or "").strip().lower()
        allowed_names = SALES_USER_MAP.get(current_username)
        if allowed_names:
            qs = qs.filter(salesman__salesman_name__in=allowed_names)
        else:
            qs = qs.none()
    return qs


def _combined_quotation_sap_effective_division(request):
    """
    Same rules as in-app admin company + division dropdown: returns
    ('ALABAMA'|'JUNAID'|None, impossible: bool). None = no extra SAP division filter.
    """
    admin_div = None
    if hasattr(request.user, 'role') and request.user.role.role == 'Admin':
        uname = (request.user.username or '').strip().lower()
        if uname in ('so', 'manager'):
            admin_div = None
        elif getattr(request.user.role, 'company', 'Junaid') == 'Alabama':
            admin_div = 'ALABAMA'
        else:
            admin_div = 'JUNAID'

    sel = (request.GET.get('division') or 'All').strip()
    if sel in ('', 'All'):
        chosen = None
    else:
        chosen = sel.upper()
        if chosen not in ('ALABAMA', 'JUNAID'):
            chosen = None

    if admin_div:
        if chosen:
            if chosen != admin_div:
                return None, True
            return chosen, False
        return admin_div, False
    if chosen:
        return chosen, False
    return None, False


def _sap_quotation_alabama_hints_q():
    """Heuristic match for SAP rows tied to Alabama (no division field on SAP header)."""
    return (
        Q(salesman_name__icontains='ALABAMA')
        | Q(customer_name__icontains='Alabama')
        | Q(bill_to__icontains='Alabama')
    )


def _quotation_salesman_pick_list(request):
    """Names from multiselect GET (repeated salesman_filter). Empty list = no salesman filter."""
    out = []
    for s in request.GET.getlist('salesman_filter'):
        t = (s or '').strip()
        if t and t != 'All':
            out.append(t)
    return out


def inapp_quotations_filtered_qs(request):
    """
    Quotations queryset matching view_quotations / view_quotations_ajax filters (GET params).
    Ordered by -created_at. Uses select_related for list/export efficiency.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    salesmen_pick = _quotation_salesman_pick_list(request)
    status = request.GET.get('status', 'All')
    division = request.GET.get('division', 'All')
    q = (request.GET.get('q') or '').strip()

    quotations = Quotation.objects.all().select_related('customer', 'salesman')

    if status and status != 'All':
        quotations = quotations.filter(status=status)

    if hasattr(request.user, 'role') and request.user.role.role == 'Admin':
        if request.user.username.lower() in ['so', 'manager']:
            pass
        elif getattr(request.user.role, 'company', 'Junaid') == 'Alabama':
            quotations = quotations.filter(division='ALABAMA')
        else:
            quotations = quotations.filter(division='JUNAID')

    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
        from .views import SALES_USER_MAP
        current_username = (request.user.username or "").strip().lower()
        allowed_names = SALES_USER_MAP.get(current_username)
        if allowed_names:
            quotations = quotations.filter(salesman__salesman_name__in=allowed_names)
        else:
            quotations = quotations.none()
    elif salesmen_pick:
        sm_q = Q()
        for name in expand_quotation_salesman_picks(salesmen_pick):
            sm_q |= Q(salesman__salesman_name__iexact=name)
        quotations = quotations.filter(sm_q)

    if start_date:
        quotations = quotations.filter(quotation_date__gte=start_date)
    if end_date:
        quotations = quotations.filter(quotation_date__lte=end_date)

    if division and division != 'All':
        div = (division or '').strip().upper()
        if div in ('ALABAMA', 'JUNAID'):
            quotations = quotations.filter(division=div)

    if q:
        quotations = quotations.filter(
            Q(quotation_number__icontains=q)
            | Q(customer__customer_name__icontains=q)
            | Q(customer__customer_code__icontains=q)
            | Q(salesman__salesman_name__icontains=q)
            | Q(remarks__icontains=q)
            | Q(customer_display_name__icontains=q)
        )

    return quotations.order_by('-created_at')


def sap_quotations_filtered_qs_combined(request):
    """
    SAP quotations with salesman_scope_q and GET filters aligned with the combined list:
    dates, search, salesman (iexact), division/admin company (heuristic), app-style status.
    """
    from .models import SAPQuotation
    from .views import salesman_scope_q

    eff_div, div_impossible = _combined_quotation_sap_effective_division(request)
    if div_impossible:
        return SAPQuotation.objects.none()

    qs = SAPQuotation.objects.filter(salesman_scope_q(request.user))
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    salesmen_pick = _quotation_salesman_pick_list(request)
    status = (request.GET.get('status') or 'All').strip()
    q = (request.GET.get('q') or '').strip()

    if eff_div == 'ALABAMA':
        qs = qs.filter(_sap_quotation_alabama_hints_q())
    elif eff_div == 'JUNAID':
        qs = qs.exclude(_sap_quotation_alabama_hints_q())

    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
        pass
    elif salesmen_pick:
        sm_q = Q()
        for name in expand_quotation_salesman_picks(salesmen_pick):
            sm_q |= Q(salesman_name__iexact=name)
        qs = qs.filter(sm_q)

    if status and status != 'All':
        if status in ('Pending', 'On Hold'):
            qs = qs.filter(status__in=['O', 'OPEN', 'Open', 'open'])
        # App "Approved" has no SAP equivalent; leave SAP unfiltered by document status.

    if start_date:
        qs = qs.filter(posting_date__gte=start_date)
    if end_date:
        qs = qs.filter(posting_date__lte=end_date)

    if q:
        if q.isdigit():
            qs = qs.filter(q_number__istartswith=q)
        elif len(q) < 3:
            qs = qs.filter(
                Q(customer_name__istartswith=q) |
                Q(salesman_name__istartswith=q)
            )
        else:
            qs = qs.filter(
                Q(q_number__icontains=q) |
                Q(customer_code__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(salesman_name__icontains=q)
            )
    return qs


def _style_inapp_quotation_excel_worksheet(worksheet):
    """Header styling, autosized columns (with sensible mins/maxes), freeze header row."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_align_wrap = Alignment(vertical="center", wrap_text=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column in worksheet.columns:
        col_letter = column[0].column_letter
        header_title = str(column[0].value or "")
        max_length = len(header_title)
        for cell in column:
            if cell.row == 1:
                continue
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
            if header_title in ('Price', 'Total Value', 'Quantity'):
                cell.alignment = Alignment(vertical="center", horizontal="right")
            else:
                cell.alignment = body_align_wrap
        if header_title == 'Source':
            adjusted_width = max(max_length + 2, 11)
        elif header_title in ('Customer Name', 'Item Name'):
            adjusted_width = min(max(max_length + 4, 28), 55)
        elif header_title in ('Price', 'Total Value'):
            adjusted_width = min(max(max_length + 4, 16), 22)
        elif header_title == 'Quantity':
            adjusted_width = min(max(max_length + 3, 14), 18)
        elif header_title == 'Date':
            adjusted_width = max(max_length + 2, 14)
        else:
            adjusted_width = min(max(max_length + 3, 16), 48)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    worksheet.row_dimensions[1].height = 24
    worksheet.freeze_panes = "A2"


def view_quotations(request):
    import calendar
    from datetime import date as date_cls
    from django.db.models import Sum, Count, FloatField, Value
    from django.db.models.functions import Coalesce
    from django.http import HttpResponse, HttpResponseForbidden

    show_inapp_calendar = (
        request.user.is_authenticated
        and hasattr(request.user, 'role')
        and getattr(request.user.role, 'role', None) == 'Admin'
    )

    if request.GET.get('ajax') == 'inapp_quotation_calendar' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if not show_inapp_calendar:
            return HttpResponseForbidden('Forbidden')

    if show_inapp_calendar:
        today = date_cls.today()
        current_year = today.year
        current_month = today.month
        qcal_year_raw = request.GET.get('qcal_year', '').strip()
        qcal_month_raw = request.GET.get('qcal_month', '').strip()
        calendar_year_min = 2020
        calendar_year_max = current_year + 1
        calendar_years = list(range(calendar_year_min, calendar_year_max + 1))
        try:
            qcal_year = int(qcal_year_raw) if qcal_year_raw else current_year
            qcal_month = int(qcal_month_raw) if qcal_month_raw else current_month
        except (ValueError, TypeError):
            qcal_year = current_year
            qcal_month = current_month
        if qcal_year not in calendar_years:
            qcal_year = current_year
        if qcal_month < 1 or qcal_month > 12:
            qcal_month = current_month

        month_names_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        calendar_months = [(i, month_names_short[i - 1]) for i in range(1, 13)]
        calendar_display = date_cls(qcal_year, qcal_month, 1)

        scope_qs = _inapp_quotation_scope_for_calendar(request)

        _, last_day_in_month = calendar.monthrange(qcal_year, qcal_month)
        is_calendar_current_month = qcal_year == current_year and qcal_month == current_month
        last_calendar_day = min(today.day, last_day_in_month) if is_calendar_current_month else last_day_in_month

        def _agg_division(qs, div):
            r = qs.filter(division=div).aggregate(
                total=Coalesce(Sum('grand_total'), Value(0.0, output_field=FloatField())),
                cnt=Count('id'),
            )
            return float(r['total'] or 0), r['cnt'] or 0

        inapp_month_days = []
        for day_num in range(1, last_calendar_day + 1):
            day_date = date_cls(qcal_year, qcal_month, day_num)
            day_qs = scope_qs.filter(quotation_date=day_date)
            j_total, j_cnt = _agg_division(day_qs, 'JUNAID')
            a_total, a_cnt = _agg_division(day_qs, 'ALABAMA')
            inapp_month_days.append({
                'date': day_date,
                'formatted_date': f"{month_names_short[qcal_month - 1]} {day_num}",
                'junaid_total': j_total,
                'junaid_count': j_cnt,
                'alabama_total': a_total,
                'alabama_count': a_cnt,
                'has_quotations': (j_cnt + a_cnt) > 0,
            })

        month_qs = scope_qs.filter(quotation_date__year=qcal_year, quotation_date__month=qcal_month)
        junaid_month_total, junaid_month_count = _agg_division(month_qs, 'JUNAID')
        alabama_month_total, alabama_month_count = _agg_division(month_qs, 'ALABAMA')

        if request.GET.get('ajax') == 'inapp_quotation_calendar' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string(
                'so/quotations/_inapp_quotation_calendar_body.html',
                {
                    'month_days': inapp_month_days,
                    'today': today,
                    'junaid_month_total': junaid_month_total,
                    'junaid_month_count': junaid_month_count,
                    'alabama_month_total': alabama_month_total,
                    'alabama_month_count': alabama_month_count,
                },
                request=request,
            )
            return HttpResponse(html, content_type='text/html')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page = request.GET.get('page', 1)
    status = request.GET.get('status', 'All')  # Default to 'All'
    division = request.GET.get('division', 'All')
    q = (request.GET.get('q') or '').strip()

    quotations = inapp_quotations_filtered_qs(request)

    # Get all unique salesmen for the filter dropdown
    all_salesmen = Salesman.objects.all().order_by('salesman_name')

    # Pagination - 12 items per page (3x4 grid)
    paginator = Paginator(quotations, 12)

    try:
        quotations_page = paginator.page(page)
    except PageNotAnInteger:
        quotations_page = paginator.page(1)
    except EmptyPage:
        quotations_page = paginator.page(paginator.num_pages)

    # Build query string for pagination links
    query_params = []
    if status and status != 'All':
        query_params.append(f"status={status}")
    if start_date:
        query_params.append(f"start_date={start_date}")
    if end_date:
        query_params.append(f"end_date={end_date}")
    for sm in _quotation_salesman_pick_list(request):
        query_params.append('salesman_filter=' + quote(sm))
    if division and division != 'All':
        query_params.append(f"division={division}")
    if q:
        query_params.append(f"q={q}")
    query_string = "&".join(query_params)

    context = {
        'quotations': quotations_page,
        'all_salesmen': all_salesmen,
        'selected_salesman': request.GET.get('salesman_filter'),
        'current_status': status,
        'start_date': start_date,
        'end_date': end_date,
        'selected_division': division or 'All',
        'search_query': q,
        'query_string': query_string,
        'show_inapp_calendar': show_inapp_calendar,
    }
    if show_inapp_calendar:
        context.update({
            'qcal_year': qcal_year,
            'qcal_month': qcal_month,
            'calendar_years': calendar_years,
            'calendar_months': calendar_months,
            'calendar_display': calendar_display,
            'inapp_month_days': inapp_month_days,
            'junaid_month_total': junaid_month_total,
            'junaid_month_count': junaid_month_count,
            'alabama_month_total': alabama_month_total,
            'alabama_month_count': alabama_month_count,
            'today': today,
        })
    return render(request, 'so/quotations/view_quotations.html', context)


@login_required
@require_GET
def view_quotations_ajax(request):
    """
    AJAX endpoint for backend filtering/search/pagination on quotations dashboard.
    Supports:
    - q (search)
    - status
    - start_date / end_date
    - salesman_filter
    - division: All | ALABAMA | JUNAID
    - page
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page = request.GET.get('page', 1)
    status = request.GET.get('status', 'All')
    division = request.GET.get('division', 'All')
    q = (request.GET.get('q') or '').strip()

    quotations = inapp_quotations_filtered_qs(request)

    paginator = Paginator(quotations, 12)
    try:
        quotations_page = paginator.page(page)
    except PageNotAnInteger:
        quotations_page = paginator.page(1)
    except EmptyPage:
        quotations_page = paginator.page(paginator.num_pages)

    query_params = []
    if status and status != 'All':
        query_params.append(f"status={status}")
    if start_date:
        query_params.append(f"start_date={start_date}")
    if end_date:
        query_params.append(f"end_date={end_date}")
    for sm in _quotation_salesman_pick_list(request):
        query_params.append('salesman_filter=' + quote(sm))
    if division and division != 'All':
        query_params.append(f"division={division}")
    if q:
        query_params.append(f"q={q}")
    query_string = "&".join(query_params)

    html = render_to_string(
        'so/quotations/_quotations_results.html',
        {
            'quotations': quotations_page,
            'current_status': status or 'All',
            'selected_salesman': request.GET.get('salesman_filter'),
            'selected_division': division or 'All',
            'start_date': start_date,
            'end_date': end_date,
            'query_string': query_string,
        },
        request=request
    )

    return JsonResponse({'html': html, 'count': paginator.count})


@login_required
@require_GET
def export_inapp_quotations_consolidated_excel(request):
    """Excel: one row per quotation — number, customer code, name, grand total. Respects list filters."""
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone

    qs = inapp_quotations_filtered_qs(request)
    rows = []
    for quot in qs.iterator(chunk_size=500):
        cust = quot.customer
        name = quot.customer_display_name or (cust.customer_name if cust else '')
        d = quot.quotation_date
        date_str = d.strftime('%Y-%m-%d') if d else ''
        rows.append({
            'Quotation Number': quot.quotation_number or '',
            'Date': date_str,
            'Customer Code': cust.customer_code if cust else '',
            'Customer Name': name,
            'Total Value': float(quot.grand_total or 0),
        })
    cols = ['Quotation Number', 'Date', 'Customer Code', 'Customer Name', 'Total Value']
    df = pd.DataFrame(rows, columns=cols)
    if df.empty:
        df = pd.DataFrame(columns=cols)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Quotations', index=False)
        _style_inapp_quotation_excel_worksheet(writer.sheets['Quotations'])
    output.seek(0)
    filename = 'inapp_quotations_%s.xlsx' % timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response


@login_required
@require_GET
def export_inapp_quotations_items_excel(request):
    """Excel: one row per line item — quotation, customer, item no/name, unit price, line total. Respects list filters."""
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone

    base_qs = inapp_quotations_filtered_qs(request)
    item_qs = (
        QuotationItem.objects.filter(quotation__in=base_qs)
        .select_related('quotation', 'quotation__customer', 'item')
        .order_by('quotation_id', 'id')
    )
    rows = []
    for li in item_qs.iterator(chunk_size=1000):
        q = li.quotation
        cust = q.customer
        name = q.customer_display_name or (cust.customer_name if cust else '')
        item_no = li.item.item_code if li.item_id else ''
        item_name = li.item.item_description if li.item_id else ''
        price = float(li.price or 0)
        line_total = float(li.line_total or ((li.quantity or 0) * price))
        d = q.quotation_date
        date_str = d.strftime('%Y-%m-%d') if d else ''
        rows.append({
            'Quotation Number': q.quotation_number or '',
            'Date': date_str,
            'Customer Code': cust.customer_code if cust else '',
            'Customer Name': name,
            'Item No': item_no,
            'Item Name': item_name,
            'Price': price,
            'Total Value': line_total,
        })
    cols = [
        'Quotation Number', 'Date', 'Customer Code', 'Customer Name',
        'Item No', 'Item Name', 'Price', 'Total Value',
    ]
    df = pd.DataFrame(rows, columns=cols)
    if df.empty:
        df = pd.DataFrame(columns=cols)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Quotation Lines', index=False)
        _style_inapp_quotation_excel_worksheet(writer.sheets['Quotation Lines'])
    output.seek(0)
    filename = 'inapp_quotation_lines_%s.xlsx' % timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response


def _combined_list_export_source(request):
    source = (request.GET.get('source') or 'all').strip().lower()
    if source not in ('all', 'sap', 'app'):
        source = 'all'
    return source


def _combined_sort_date(s):
    from datetime import date as date_cls
    from datetime import datetime as dt_cls
    if not s:
        return date_cls(1900, 1, 1)
    try:
        return dt_cls.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return date_cls(1900, 1, 1)


@login_required
@require_GET
def export_combined_quotations_consolidated_excel(request):
    """Combined SAP + in-app quotations, one row per quote; same filters as combined list."""
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone

    source = _combined_list_export_source(request)
    rows = []
    if source in ('all', 'sap'):
        sap_qs = sap_quotations_filtered_qs_combined(request).order_by('-posting_date', '-id')
        for s in sap_qs.iterator(chunk_size=500):
            d = s.posting_date
            date_str = d.strftime('%Y-%m-%d') if d else ''
            rows.append({
                'Source': 'SAP',
                'Quotation Number': s.q_number,
                'Date': date_str,
                'Customer Code': s.customer_code or '',
                'Customer Name': s.customer_name or '',
                'Total Value': float(s.document_total or 0),
                'Status': (s.status or ''),
                'Salesman': s.salesman_name or '',
            })
    if source in ('all', 'app'):
        app_qs = inapp_quotations_filtered_qs(request).order_by('-quotation_date', '-id')
        for a in app_qs.iterator(chunk_size=500):
            cust = a.customer
            d = a.quotation_date
            date_str = d.strftime('%Y-%m-%d') if d else ''
            name = a.customer_display_name or (cust.customer_name if cust else '')
            rows.append({
                'Source': 'App',
                'Quotation Number': a.quotation_number or '',
                'Date': date_str,
                'Customer Code': cust.customer_code if cust else '',
                'Customer Name': name,
                'Total Value': float(a.grand_total or 0),
                'Status': (a.status or ''),
                'Salesman': a.salesman.salesman_name if a.salesman_id else '',
            })
    rows.sort(
        key=lambda r: (
            _combined_sort_date(r['Date']),
            r.get('Quotation Number') or '',
            r.get('Source') or '',
        ),
        reverse=True,
    )
    cols = [
        'Source', 'Quotation Number', 'Date', 'Customer Code', 'Customer Name',
        'Total Value', 'Status', 'Salesman',
    ]
    df = pd.DataFrame(rows, columns=cols)
    if df.empty:
        df = pd.DataFrame(columns=cols)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Combined Quotations', index=False)
        _style_inapp_quotation_excel_worksheet(writer.sheets['Combined Quotations'])
    output.seek(0)
    filename = 'combined_quotations_%s.xlsx' % timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response


@login_required
@require_GET
def export_combined_quotations_items_excel(request):
    """Combined SAP + in-app line items; respects combined filters."""
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    from .models import SAPQuotationItem

    source = _combined_list_export_source(request)
    rows = []
    if source in ('all', 'sap'):
        sap_base = sap_quotations_filtered_qs_combined(request)
        sap_items = (
            SAPQuotationItem.objects.filter(quotation__in=sap_base)
            .select_related('quotation')
            .order_by('quotation_id', 'id')
        )
        for li in sap_items.iterator(chunk_size=1000):
            q = li.quotation
            d = q.posting_date
            date_str = d.strftime('%Y-%m-%d') if d else ''
            price = float(li.price or 0)
            qty = float(li.quantity or 0)
            line_total = float(li.row_total) if li.row_total is not None else qty * price
            rows.append({
                'Source': 'SAP',
                'Quotation Number': q.q_number,
                'Date': date_str,
                'Customer Code': q.customer_code or '',
                'Customer Name': q.customer_name or '',
                'Item No': li.item_no or '',
                'Item Name': li.description or '',
                'Quantity': qty,
                'Price': price,
                'Total Value': line_total,
            })
    if source in ('all', 'app'):
        app_base = inapp_quotations_filtered_qs(request)
        app_items = (
            QuotationItem.objects.filter(quotation__in=app_base)
            .select_related('quotation', 'quotation__customer', 'item')
            .order_by('quotation_id', 'id')
        )
        for li in app_items.iterator(chunk_size=1000):
            qo = li.quotation
            cust = qo.customer
            d = qo.quotation_date
            date_str = d.strftime('%Y-%m-%d') if d else ''
            name = qo.customer_display_name or (cust.customer_name if cust else '')
            item_no = li.item.item_code if li.item_id else ''
            item_name = li.item.item_description if li.item_id else ''
            price = float(li.price or 0)
            qty = float(li.quantity or 0)
            line_total = float(li.line_total or (qty * price))
            rows.append({
                'Source': 'App',
                'Quotation Number': qo.quotation_number or '',
                'Date': date_str,
                'Customer Code': cust.customer_code if cust else '',
                'Customer Name': name,
                'Item No': item_no,
                'Item Name': item_name,
                'Quantity': qty,
                'Price': price,
                'Total Value': line_total,
            })
    rows.sort(
        key=lambda r: (
            _combined_sort_date(r['Date']),
            r.get('Quotation Number') or '',
            r.get('Source') or '',
            r.get('Item No') or '',
        ),
        reverse=True,
    )
    cols = [
        'Source', 'Quotation Number', 'Date', 'Customer Code', 'Customer Name',
        'Item No', 'Item Name', 'Quantity', 'Price', 'Total Value',
    ]
    df = pd.DataFrame(rows, columns=cols)
    if df.empty:
        df = pd.DataFrame(columns=cols)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Combined Lines', index=False)
        _style_inapp_quotation_excel_worksheet(writer.sheets['Combined Lines'])
    output.seek(0)
    filename = 'combined_quotation_lines_%s.xlsx' % timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response


@login_required
def combined_quotations_list(request):
    """
    Single list merging SAP quotations (posting_date) and in-app quotations (quotation_date),
    sorted by date descending. Reuses in-app filter GET params; source=all|sap|app.
    """
    from datetime import date as date_cls
    from django.urls import reverse

    source = (request.GET.get('source') or 'all').strip().lower()
    if source not in ('all', 'sap', 'app'):
        source = 'all'

    rows = []
    if source in ('all', 'sap'):
        sap_qs = sap_quotations_filtered_qs_combined(request).order_by('-posting_date', '-id')
        for s in sap_qs.iterator(chunk_size=500):
            d = s.posting_date
            rows.append({
                'display_date': d,
                'source': 'SAP',
                'source_label': 'SAP',
                'number': s.q_number,
                'customer_code': s.customer_code or '',
                'customer_name': s.customer_name or '',
                'total': float(s.document_total or 0),
                'status': (s.status or '')[:80],
                'salesman': s.salesman_name or '',
                'detail_url': reverse('quotation_detail', args=[s.q_number]),
            })
    if source in ('all', 'app'):
        app_qs = inapp_quotations_filtered_qs(request).order_by('-quotation_date', '-id')
        for a in app_qs.iterator(chunk_size=500):
            cust = a.customer
            d = a.quotation_date
            name = a.customer_display_name or (cust.customer_name if cust else '')
            rows.append({
                'display_date': d,
                'source': 'APP',
                'source_label': 'App',
                'number': a.quotation_number or '',
                'customer_code': cust.customer_code if cust else '',
                'customer_name': name,
                'total': float(a.grand_total or 0),
                'status': (a.status or '')[:80],
                'salesman': a.salesman.salesman_name if a.salesman_id else '',
                'detail_url': reverse('view_quotation_details', args=[a.id]),
            })

    _min_date = date_cls(1900, 1, 1)
    rows.sort(
        key=lambda r: ((r['display_date'] or _min_date), r.get('number') or ''),
        reverse=True,
    )

    page = request.GET.get('page', 1)
    try:
        page_size = int(request.GET.get('page_size', 50))
    except ValueError:
        page_size = 50
    page_size = max(10, min(page_size, 200))

    paginator = Paginator(rows, page_size)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    qd = request.GET.copy()
    qd.pop('page', None)
    query_string = qd.urlencode()

    salesman_canonical_choices = get_quotation_salesman_canonical_choices_sorted()

    return render(request, 'so/quotations/combined_quotations.html', {
        'page_obj': page_obj,
        'total_count': paginator.count,
        'query_string': query_string,
        'salesman_canonical_choices': salesman_canonical_choices,
        'selected_salesmen': normalize_quotation_salesman_picks_to_canonicals(
            _quotation_salesman_pick_list(request)
        ),
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'search_query': request.GET.get('q', ''),
        'selected_division': request.GET.get('division', 'All'),
        'current_status': request.GET.get('status', 'All'),
        'source': source,
        'page_size': page_size,
    })


from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages
from .models import Quotation, QuotationItem

def view_quotation_details(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    quotation_items = quotation.items.all()

    # ✅ Compute totals & undercost logic
    grand_total = 0
    has_undercost_items = False
    is_admin = hasattr(request.user, 'role') and request.user.role.role == 'Admin'
    total_cost = 0.0
    total_margin = 0.0
    margin_percent = 0.0
    has_zero_cost_items = False

    for item in quotation_items:
        item.line_total = item.quantity * item.price
        grand_total += item.line_total

        # Check undercost
        if hasattr(item, "item") and item.item:
            undercost_limit = item.item.item_cost   
            item.is_undercost = item.price < undercost_limit
            if item.is_undercost:
                has_undercost_items = True

            # Margin per item (Admin only) - skip if cost is zero
            cost_val = float(item.item.item_cost or 0)
            if cost_val == 0:
                has_zero_cost_items = True
                item.line_cost = None
                item.line_margin = None
                item.margin_pct = None
            else:
                item.line_cost = cost_val * item.quantity
                item.line_margin = item.line_total - item.line_cost
                item.margin_pct = (item.line_margin / item.line_total * 100) if item.line_total else None
        else:
            item.is_undercost = False
            item.line_cost = None
            item.line_margin = None
            item.margin_pct = None

    # Overall margin (Admin only, and only if no zero-cost items)
    if is_admin and not has_zero_cost_items:
        for item in quotation_items:
            if getattr(item, 'line_cost') is not None:
                total_cost += item.line_cost
        total_margin = grand_total - total_cost
        margin_percent = (total_margin / grand_total * 100) if grand_total else 0.0

    # Grand total with 5% VAT
    vat_rate = 0.05
    vat_amount = round(grand_total * vat_rate, 2)
    grand_total_with_vat = round(grand_total + vat_amount, 2)

    # 🔹 Automatic approval if no undercost items
    if not has_undercost_items and quotation.status != 'Approved':
        quotation.status = 'Approved'
        quotation.save()
        messages.success(request, "Quotation auto-approved as all prices are above minimum selling price.")

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_remarks':
            quotation.remarks = request.POST.get('remarks', '')
            quotation.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'remarks': quotation.remarks
                })
            else:
                messages.success(request, 'Remarks updated successfully!')
                return redirect('view_quotation_details', quotation_id=quotation_id)

        elif action == 'approve':
            quotation.status = 'Approved'
            quotation.save()
            messages.success(request, 'Quotation approved successfully!')
            return redirect('view_quotation_details', quotation_id=quotation_id)

        elif action == 'hold':
            quotation.status = 'On Hold'
            quotation.save()
            messages.warning(request, 'Quotation put on hold.')
            return redirect('view_quotation_details', quotation_id=quotation_id)

        elif action == 'manager_change_division' and _is_manager_account(request.user):
            new_div = (request.POST.get('division') or '').strip().upper()
            if new_div not in ('JUNAID', 'ALABAMA'):
                messages.error(request, 'Invalid division.')
            elif quotation.division == new_div:
                messages.info(request, 'Division is already set to that value.')
            else:
                quotation.division = new_div
                quotation.save(update_fields=['division'])
                messages.success(
                    request,
                    f'Division updated to {quotation.get_division_display()}. '
                    f'PDF exports will use that letterhead. Quotation number {quotation.quotation_number} is unchanged.',
                )
            return redirect('view_quotation_details', quotation_id=quotation_id)

    return render(request, "so/quotations/view_quotation_details.html", {
        "quotation": quotation,
        "quotation_items": quotation_items,
        "grand_total": grand_total,
        "vat_amount": vat_amount,
        "grand_total_with_vat": grand_total_with_vat,
        "has_undercost_items": has_undercost_items,
        "is_admin": is_admin,
        "has_zero_cost_items": has_zero_cost_items,
        "total_cost": total_cost,
        "total_margin": total_margin,
        "margin_percent": margin_percent,
        "is_manager": _is_manager_account(request.user),
    })


@login_required
@transaction.atomic
def convert_quotation_to_sales_order(request, quotation_id):
    """
    Convert a quotation to a sales order.
    Copies all quotation data (customer, salesman, items, prices) to create a new sales order.
    """
    quotation = get_object_or_404(Quotation, id=quotation_id)
    
    # Check if already converted
    if quotation.converted_to_sales_order:
        messages.warning(
            request, 
            f'This quotation has already been converted to Sales Order {quotation.converted_to_sales_order.order_number}.'
        )
        return redirect('view_sales_order_details', order_id=quotation.converted_to_sales_order.id)
    
    quotation_items = quotation.items.all()
    
    # Validate that quotation has items
    if not quotation_items.exists():
        messages.error(request, 'Cannot convert quotation with no items.')
        return redirect('view_quotation_details', quotation_id=quotation_id)
    
    try:
        # Determine division (same logic as create_sales_order)
        division = quotation.division  # Use quotation's division
        if not division:
            # Fallback to user-based division
            division = 'JUNAID'  # Default
            if request.user.is_authenticated and 'alabama' in request.user.username.lower():
                division = 'ALABAMA'
        
        # Create the sales order
        sales_order = SalesOrder.objects.create(
            customer=quotation.customer,
            division=division,
            salesman=quotation.salesman,
            remarks=quotation.remarks or '',
        )
        
        # Process order items from quotation items
        order_items = []
        customer_price_updates = []
        total_amount = 0.0
        
        for quotation_item in quotation_items:
            if not quotation_item.item:
                continue  # Skip items without valid item reference
            
            item = quotation_item.item
            quantity = quotation_item.quantity
            price = quotation_item.price
            unit = quotation_item.unit if quotation_item.unit in ['pcs', 'ctn', 'roll'] else 'pcs'
            
            # Check if price is custom (differs from item default price)
            is_custom_price = abs(float(price) - float(item.item_price)) > 0.01
            
            # Calculate line total
            line_total = quantity * price
            total_amount += line_total
            
            order_items.append(OrderItem(
                order=sales_order,
                item=item,
                quantity=quantity,
                price=price,
                unit=unit,
                is_custom_price=is_custom_price
            ))
            
            # Track customer price updates if custom price
            if is_custom_price:
                customer_price_updates.append((quotation.customer, item, price))
        
        # Bulk create order items
        if order_items:
            OrderItem.objects.bulk_create(order_items)
            
            # Update customer prices for custom prices
            for customer, item, price in customer_price_updates:
                CustomerPrice.objects.update_or_create(
                    customer=customer,
                    item=item,
                    defaults={'custom_price': price}
                )
            
            # Calculate totals
            tax = round(0.05 * total_amount, 2)
            sales_order.tax = tax
            sales_order.total_amount = total_amount
            sales_order.save()
            
            # Mark quotation as converted
            quotation.converted_to_sales_order = sales_order
            quotation.save()
            
            messages.success(
                request, 
                f'Quotation {quotation.quotation_number} successfully converted to Sales Order {sales_order.order_number}!'
            )
            return redirect('view_sales_order_details', order_id=sales_order.id)
        else:
            # No valid items, delete the sales order
            sales_order.delete()
            messages.error(request, 'No valid items found in quotation to convert.')
            return redirect('view_quotation_details', quotation_id=quotation_id)
            
    except Exception as e:
        messages.error(request, f'An error occurred while converting quotation: {str(e)}')
        return redirect('view_quotation_details', quotation_id=quotation_id)


from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from .models import Quotation, QuotationItem, Customer, Salesman, Items

@csrf_exempt
@transaction.atomic
def edit_quotation(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)

    if request.method == 'POST':
        customer = get_object_or_404(Customer, id=request.POST.get('customer'))
        salesman = get_object_or_404(Salesman, id=request.POST.get('salesman')) if request.POST.get('salesman') else None
        customer_display_name = request.POST.get('customer_display_name', '').strip()

        # Update main quotation fields
        quotation.customer = customer
        quotation.salesman = salesman
        quotation.customer_display_name = customer_display_name if customer_display_name else None
        quotation.remarks = request.POST.get('remarks', '').strip() or None

        # Get new items from POST
        item_ids = request.POST.getlist('item')  # dropdown selection
        quantities = request.POST.getlist('quantity')
        prices = request.POST.getlist('price')
        units = request.POST.getlist('unit')

        # Validate we have the same number of items, quantities, prices, and units
        if len(item_ids) != len(quantities) or len(item_ids) != len(prices) or len(item_ids) != len(units):
            messages.error(request, 'Invalid form data: mismatched item fields')
            return redirect('edit_quotation', quotation_id=quotation.id)

        # Batch-fetch all items to avoid N+1 queries inside the loop
        valid_item_ids = [iid for iid in item_ids if iid]
        items_map = {
            str(obj.id): obj
            for obj in Items.objects.filter(id__in=valid_item_ids)
        }

        quotation_items = []
        has_undercost_items = False
        
        for i, (item_id, qty, price, unit) in enumerate(zip(item_ids, quantities, prices, units)):
            if not item_id:  # Skip empty items
                continue
                
            try:
                quantity = int(qty) if qty else 0
                price_val = float(price) if price else 0.0
                unit_val = unit if unit in ['pcs', 'ctn', 'roll'] else 'pcs'
                
                if quantity <= 0 or price_val < 0:
                    continue  # Skip invalid items
                    
                item = items_map.get(str(item_id))
                if not item:
                    continue
                
                # Check if item is undercost
                undercost_limit = item.item_cost  # 10% above cost
                if price_val < undercost_limit:
                    has_undercost_items = True
                
                quotation_items.append(QuotationItem(
                    quotation=quotation,
                    item=item,  # Link to Items model
                    quantity=quantity,
                    price=price_val,
                    unit=unit_val,
                    line_total=quantity * price_val
                ))
            except (ValueError, Items.DoesNotExist):
                # Skip invalid items but continue processing others
                continue

        # Only proceed if we have valid items
        if quotation_items:
            # Remove old items
            quotation.items.all().delete()
            
            # Bulk create new items
            QuotationItem.objects.bulk_create(quotation_items)

            # Recalculate totals
            total = sum(qi.line_total for qi in quotation_items)
            quotation.total_amount = total
            quotation.grand_total = total
            
            # 🔥 Update status based on undercost items
            if has_undercost_items:
                quotation.status = 'Pending'
                status_message = 'Quotation updated! Status changed to Pending due to undercost items.'
            else:
                quotation.status = 'Approved'
                status_message = 'Quotation updated and auto-approved! All prices are above minimum selling price.'
            
            quotation.save()
            
            messages.success(request, status_message)
        else:
            messages.error(request, 'No valid items found in the quotation')

        return redirect('view_quotation_details', quotation_id=quotation.id)

    # GET request → render form
    salesmen = Salesman.objects.all()

    # Filter Logic: Only apply if user is logged in AND is NOT a superuser/admin
    if request.user.is_authenticated and not request.user.is_superuser:
        current_username = request.user.username.lower()

        # Define Mapping: 'username' -> List of ['Name1', 'Name2']
        user_salesman_map = {
            'alabamakadhar': ['KADER'],
            'alabamamusharaf': ['MUSHARAF'],   # Multiple allowed
            'alabamaadmin': ['KADER','MUSHARAF','AIJAZ','CASH'],               # Single allowed
            
        }


        if current_username in user_salesman_map:
            target_names = user_salesman_map[current_username]
            
            # Build a complex query: (name contains A) OR (name contains B)
            query = Q()
            for name in target_names:
                query |= Q(salesman_name__icontains=name)
            
            salesmen = salesmen.filter(query)
    # -------------------------------------------------------------------------
    # END: Salesman Filtering Logic
    firms = Items.objects.values_list('item_firm', flat=True).distinct()

    return render(request, 'so/quotations/edit_quotation.html', {
        'quotation': quotation,
        'salesmen': salesmen,
        'firms': firms,
        'quotation_items': quotation.items.all(),
    })

import os
import requests
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    PageBreak, Frame, BaseDocTemplate, PageTemplate, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
# Adjust import based on your actual app name
from .models import Quotation 

# --- PDF Styles (Shared) ---
styles = getSampleStyleSheet()

if 'MainTitle' not in styles:
    styles.add(ParagraphStyle(
        name='MainTitle', 
        fontSize=18, 
        leading=22, 
        alignment=TA_CENTER, 
        spaceAfter=12, 
        fontName='Helvetica-Bold'
    ))

if 'SectionHeader' not in styles:
    styles.add(ParagraphStyle(
        name='SectionHeader', 
        fontSize=12, 
        leading=14, 
        alignment=TA_LEFT, 
        spaceAfter=6, 
        fontName='Helvetica-Bold', 
        textColor=HexColor("#FFFFFF")
    ))

if 'ItemDescription' not in styles:
    styles.add(ParagraphStyle(
        name='ItemDescription', 
        fontSize=10, 
        leading=12, 
        alignment=TA_LEFT
    ))

if 'ItemCode' not in styles:
    styles.add(ParagraphStyle(
        name='ItemCode',
        parent=styles['Normal'],
        fontSize=9,
        leading=10,
        alignment=TA_CENTER,
        wordWrap='CJK'
    ))

if 'h3' not in styles:
    styles.add(ParagraphStyle(
        name='h3',
        parent=styles['Heading3'],
        fontSize=14,
        leading=16,
        spaceAfter=12,
        fontName='Helvetica-Bold'
    ))

# --- Dynamic Template Class ---
class QuotationPDFTemplate(BaseDocTemplate):
    """
    Dynamic template that accepts company details and theme colors.
    """
    def __init__(self, filename, company_config, theme_config, **kwargs):
        self.company_name = company_config.get('name', "JUNAID SANITARY & ELECTRICAL REQUISITES TRADING LLC")
        self.company_address = company_config.get('address', "Dubai Investment Parks 2, Dubai, UAE")
        self.company_contact = company_config.get('contact', "Email: sales@junaid.ae | Phone: +97142367723")
        self.logo_url = company_config.get('logo_url')
        self.local_logo_path = company_config.get('local_logo_path')
        
        self.theme_color = theme_config.get('primary', HexColor('#2C5530'))
        
        self.page_count = 1 
        self.logo_image = None
        
        kwargs.setdefault('bottomMargin', 1.0 * inch)
        super().__init__(filename, **kwargs)
        
        self.logo_image = self._get_logo()

        top_margin = 1.75 * inch
        bottom_margin = 1.0 * inch
        
        frame = Frame(
            self.leftMargin,
            self.bottomMargin,
            self.width,
            self.height - top_margin - bottom_margin,
            leftPadding=0,
            rightPadding=0,
            bottomPadding=0,
            topPadding=0,
            id='normal'
        )
        
        template = PageTemplate(id='QuotationPage', frames=[frame], onPage=self.on_page)
        self.addPageTemplates([template])

    def _get_logo(self):
        # Try URL first
        if self.logo_url:
            try:
                response_img = requests.get(self.logo_url, timeout=5)
                if response_img.status_code == 200:
                    return Image(BytesIO(response_img.content), width=1.5*inch, height=0.5*inch)
            except Exception:
                pass
        
        # Try Local fallback
        if self.local_logo_path:
            try:
                if os.path.exists(self.local_logo_path):
                    return Image(self.local_logo_path, width=1.5*inch, height=0.5*inch)
            except Exception:
                pass
        return None

    def on_page(self, canvas, doc):
        self._header(canvas, doc)
        self._footer(canvas, doc)

    def _header(self, canvas, doc):
        canvas.saveState()
        
        header_content = []
        company_text = f'{self.company_name}\n{self.company_address}\n{self.company_contact}'
        
        if self.logo_image:
            header_content.append([self.logo_image, company_text])
        else:
            header_content.append(['', company_text])

        header_table = Table(header_content, colWidths=[1.7*inch, 5.8*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        w, h = header_table.wrap(doc.width, doc.topMargin)
        header_table.drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h)

        # Line color based on theme
        canvas.setStrokeColor(self.theme_color)
        canvas.setLineWidth(2)
        canvas.line(doc.leftMargin, doc.height + doc.topMargin - h - 5, 
                   doc.leftMargin + doc.width, doc.height + doc.topMargin - h - 5)
        
        canvas.restoreState()

    def _footer(self, canvas, doc):
        canvas.saveState()
        
        footer_text = Paragraph(f"Thank you for your business! | {self.company_name}", styles['Normal'])
        page_num_text = Paragraph(f"Page {canvas.getPageNumber()} of {self.page_count}", styles['Normal'])
        
        footer_table = Table([[footer_text, page_num_text]], colWidths=[doc.width/2, doc.width/2])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        
        w, h = footer_table.wrap(doc.width, doc.bottomMargin)
        footer_table.drawOn(canvas, doc.leftMargin, h + 0.2*inch)
        
        canvas.restoreState()
    
    def afterFlowable(self, flowable):
        self.page_count = self.page

# ==========================================
# 1. MAIN DISPATCHER VIEW
# ==========================================
def export_quotation_to_pdf(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    
    # Setup Response
    response = HttpResponse(content_type='application/pdf')
    filename = f"Quotation_{quotation.quotation_number}_{quotation.quotation_date.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = BytesIO()
    
    # Dispatch based on Division
    if quotation.division == 'ALABAMA':
        generate_alabama_quotation(buffer, quotation)
    else:
        generate_junaid_quotation(buffer, quotation)
        
    pdf_content = buffer.getvalue()
    buffer.close()
    response.write(pdf_content)
    return response


def export_quotation_to_excel(request, quotation_id):
    """Export a single quotation to Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    quotation = get_object_or_404(Quotation, id=quotation_id)
    quotation_items = quotation.items.select_related('item').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Quotation_{quotation.quotation_number[:30]}"

    # Styles
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    table_header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    table_header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thick'), bottom=Side(style='thin')
    )
    currency_format = '#,##0.00'

    row_num = 1

    # Header
    ws.merge_cells(f'A{row_num}:G{row_num}')
    company_cell = ws[f'A{row_num}']
    company_cell.value = "QUOTATION"
    company_cell.font = header_font
    company_cell.fill = header_fill
    company_cell.alignment = header_alignment
    ws.row_dimensions[row_num].height = 30
    row_num += 2

    # Quotation info
    display_name = quotation.customer_display_name or quotation.customer.customer_name
    order_info = [
        ("Quotation Number:", quotation.quotation_number),
        ("Date:", quotation.quotation_date.strftime("%d-%m-%Y")),
        ("Customer:", display_name),
        ("Customer Code:", quotation.customer.customer_code),
        ("Salesman:", quotation.salesman.salesman_name if quotation.salesman else "N/A"),
        ("Division:", quotation.division or "N/A"),
    ]
    for label, value in order_info:
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = Font(bold=True, size=10)
        ws[f'B{row_num}'] = value
        ws[f'B{row_num}'].font = Font(size=10)
        ws.merge_cells(f'B{row_num}:D{row_num}')
        row_num += 1

    if quotation.remarks:
        row_num += 1
        ws[f'A{row_num}'] = "Remarks:"
        ws[f'A{row_num}'].font = Font(bold=True, size=10)
        row_num += 1
        ws[f'A{row_num}'] = quotation.remarks
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical="top")
        row_num += 1

    row_num += 1

    # Items table header
    headers = ['S.No', 'Item Code', 'Item Description', 'Qty', 'Unit', 'Unit Price', 'Total']
    col_widths = [8, 15, 40, 10, 8, 15, 15]
    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width
    row_num += 1

    # Item rows
    subtotal = 0.0
    for idx, qi in enumerate(quotation_items, 1):
        line_total = qi.quantity * qi.price
        subtotal += line_total
        desc = qi.item.item_description if qi.item else "N/A"
        code = qi.item.item_code if qi.item else "N/A"
        row_data = [idx, code, desc, qi.quantity, qi.unit or 'pcs', float(qi.price), float(line_total)]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col_num in [1, 4, 5] else "left",
                vertical="center"
            )
            if col_num in [6, 7]:
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal="right", vertical="center")
        row_num += 1

    # Summary
    row_num += 1
    vat_rate = 0.05
    vat_amount = round(subtotal * vat_rate, 2)
    grand_total = round(subtotal + vat_amount, 2)

    ws[f'E{row_num}'] = "Subtotal:"
    ws[f'E{row_num}'].font = Font(bold=True, size=10)
    ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
    ws[f'G{row_num}'] = float(subtotal)
    ws[f'G{row_num}'].number_format = currency_format
    ws[f'G{row_num}'].font = Font(bold=True, size=10)
    ws[f'G{row_num}'].border = thick_border
    row_num += 1

    ws[f'E{row_num}'] = f"VAT ({vat_rate:.0%}):"
    ws[f'E{row_num}'].font = Font(size=10)
    ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
    ws[f'G{row_num}'] = float(vat_amount)
    ws[f'G{row_num}'].number_format = currency_format
    ws[f'G{row_num}'].border = thin_border
    row_num += 1

    ws[f'E{row_num}'] = "Grand Total:"
    ws[f'E{row_num}'].font = Font(bold=True, size=12)
    ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
    ws[f'G{row_num}'] = float(grand_total)
    ws[f'G{row_num}'].number_format = currency_format
    ws[f'G{row_num}'].font = Font(bold=True, size=12, color="2C3E50")
    ws[f'G{row_num}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws[f'G{row_num}'].border = thin_border

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Quotation_{quotation.quotation_number}_{quotation.quotation_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ==========================================
# 2. JUNAID GENERATOR (Green/White Theme)
# ==========================================
def generate_junaid_quotation(buffer, quotation):
    quotation_items = quotation.items.all()
    
    # --- CONFIGURATION ---
    company_config = {
        'name': "JUNAID SANITARY & ELECTRICAL REQUISITES TRADING LLC",
        'address': "Dubai Investment Parks 2, Dubai, UAE",
        'contact': "Email: sales@junaid.ae | Phone: +97142367723",
        'logo_url': "https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp",
        'local_logo_path': os.path.join(settings.BASE_DIR, 'static', 'images', 'footer-logo.png.webp')
    }
    
    # Colors
    PRIMARY_COLOR = HexColor('#2C5530')  # Dark Green
    SECONDARY_COLOR = HexColor('#4A7C59') # Light Green
    ROW_BG_COLOR = HexColor('#F0F7F4')    # Very light green
    
    theme_config = {'primary': PRIMARY_COLOR}

    # --- DOCUMENT SETUP ---
    main_table_width = 7.2 * inch 
    doc = QuotationPDFTemplate(
        buffer,
        company_config=company_config,
        theme_config=theme_config,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=1.0*inch
    )
    
    elements = []
    
    # --- Title ---
    elements.append(Spacer(1, -1.3*inch))
    title_table_data = [[Paragraph('QUOTATION', styles['MainTitle'])]]
    title_table = Table(title_table_data, colWidths=[5.2*inch, 2*inch])
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 12),
        ('TEXTCOLOR', (1, 0), (1, 0), PRIMARY_COLOR),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 0.1*inch))

    # --- Info Tables ---
    quotation_data = [
        [Paragraph('Quotation Details', styles['SectionHeader'])],
        [Paragraph(f'<b>Number:</b> {quotation.quotation_number}', styles['Normal'])],
        [Paragraph(f'<b>Date:</b> {quotation.quotation_date.strftime("%d-%b-%Y")}', styles['Normal'])],
    ]

    quotation_info_table = Table(quotation_data, colWidths=[main_table_width / 2])
    quotation_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), SECONDARY_COLOR),
    ]))

    # Use customer_display_name if available, otherwise use customer.customer_name
    display_name = quotation.customer_display_name or quotation.customer.customer_name
    customer_data = [
        [Paragraph('Customer Information', styles['SectionHeader'])],
        [Paragraph(f'<b>Name:</b> {display_name}', styles['Normal'])],
    ]

    if quotation.salesman:
        customer_data.append([Paragraph(f'<b>Salesman:</b> {quotation.salesman.salesman_name}', styles['Normal'])])
    else:
        customer_data.append([Paragraph('', styles['Normal'])])

    customer_info_table = Table(customer_data, colWidths=[main_table_width / 2])
    customer_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), SECONDARY_COLOR),
    ]))

    info_table = Table([[quotation_info_table, customer_info_table]], colWidths=[main_table_width / 2, main_table_width / 2])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.2 * inch))

    # --- Items Table ---
    items_header = ['#', 'UPC Code', 'Description', 'Qty', 'Unit Price', 'Total']
    items_data = [items_header]
    subtotal = 0.0

    for idx, item in enumerate(quotation_items, 1):
        line_total = item.quantity * item.price
        subtotal += line_total

        desc_style = styles['ItemDescription']
        description_text = getattr(item.item, 'item_description', 'No description available')
        description_para = Paragraph(description_text, desc_style)

        upc_raw = getattr(item.item, 'item_upvc', '')
        if upc_raw is None: upc_raw = ""
        upc_para = Paragraph(str(upc_raw), styles['ItemCode'])

        items_data.append([
            str(idx),
            upc_para,
            description_para,
            f"{item.quantity} {item.unit}",
            f"AED {item.price:,.2f}",
            f"AED {line_total:,.2f}"
        ])

    items_table = Table(
        items_data,
        colWidths=[
            main_table_width * 0.05,
            main_table_width * 0.15,
            main_table_width * 0.43,
            main_table_width * 0.07,
            main_table_width * 0.15,
            main_table_width * 0.15
        ],
        repeatRows=1
    )
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [ROW_BG_COLOR, white]),
        ('ALIGN', (0, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 0.1 * inch))

    # --- Summary Table ---
    tax_rate = 0.05
    tax_amount = subtotal * tax_rate
    grand_total = subtotal + tax_amount

    summary_data = [
        ['Subtotal:', f"AED {subtotal:,.2f}"],
        [f'VAT ({tax_rate:.0%}):', f"AED {tax_amount:,.2f}"],
        ['Grand Total:', f"AED {grand_total:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[main_table_width * 0.5, main_table_width * 0.5])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('BACKGROUND', (0, 2), (-1, 2), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 2), (-1, 2), white),
    ]))
    
    summary_wrapper = Table([[summary_table]], colWidths=[main_table_width])
    summary_wrapper.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    elements.append(KeepTogether(summary_wrapper))
    elements.append(Spacer(1, 0.3 * inch))
    
    # --- Remarks & Terms ---
    if hasattr(quotation, 'remarks') and quotation.remarks:
        remarks_section = [
            Paragraph("Remarks:", styles['h3']),
            Paragraph(quotation.remarks, styles['Normal']),
            Spacer(1, 0.2 * inch)
        ]
        elements.extend(remarks_section)

    terms_section = [
        Paragraph("Terms & Conditions:", styles['h3'])
    ]
    terms = [
        "1. This quotation is valid as Agreed.",
        "2. Prices are subject to change without prior notice after the validity period.",
        "3. Delivery timelines will be confirmed upon order confirmation.",
        "4. This is a system-generated document and does not require a signature.",
    ]
    for term in terms:
        terms_section.append(Paragraph(term, styles['Normal']))
    
    elements.extend(terms_section)
    
    doc.multiBuild(elements)

# ==========================================
# 3. ALABAMA GENERATOR (Red/Black Theme)
# ==========================================
def generate_alabama_quotation(buffer, quotation):
    quotation_items = quotation.items.all()
    
    # --- CONFIGURATION ---
    company_config = {
        'name': "ALABAMA BUILDING MATERIALS TRADING", # Or full legal name
        'address': "Dubai Investment Parks 2, Dubai, UAE",
        'contact': "Email: sales@alabamauae.com", # Update if needed
        'logo_url': "https://alabamauae.com/alabama4.png",
        'local_logo_path': os.path.join(settings.BASE_DIR, 'static', 'images', 'alabama-logo.png')
    }
    
    # Colors (Red & Black Theme)
    PRIMARY_COLOR = HexColor("#211F1F")  # Dark Red
    SECONDARY_COLOR = HexColor('#211F1F') # Black/Dark Grey
    ROW_BG_COLOR = HexColor('#FFF5F5')    # Very light red for rows
    
    theme_config = {'primary': PRIMARY_COLOR}

    # --- DOCUMENT SETUP ---
    main_table_width = 7.2 * inch 
    doc = QuotationPDFTemplate(
        buffer,
        company_config=company_config,
        theme_config=theme_config,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=1.0*inch
    )
    
    elements = []
    
    # --- Title ---
    elements.append(Spacer(1, -1.3*inch))
    title_table_data = [[Paragraph('QUOTATION', styles['MainTitle'])]]
    title_table = Table(title_table_data, colWidths=[5.2*inch, 2*inch])
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 12),
        ('TEXTCOLOR', (1, 0), (1, 0), PRIMARY_COLOR),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 0.1*inch))

    # --- Info Tables ---
    quotation_data = [
        [Paragraph('Quotation Details', styles['SectionHeader'])],
        [Paragraph(f'<b>Number:</b> {quotation.quotation_number}', styles['Normal'])],
        [Paragraph(f'<b>Date:</b> {quotation.quotation_date.strftime("%d-%b-%Y")}', styles['Normal'])],
    ]

    quotation_info_table = Table(quotation_data, colWidths=[main_table_width / 2])
    quotation_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), PRIMARY_COLOR), # Use Red for headers
    ]))

    # Use customer_display_name if available, otherwise use customer.customer_name
    display_name = quotation.customer_display_name or quotation.customer.customer_name
    customer_data = [
        [Paragraph('Customer Information', styles['SectionHeader'])],
        [Paragraph(f'<b>Name:</b> {display_name}', styles['Normal'])],
    ]

    if quotation.salesman:
        customer_data.append([Paragraph(f'<b>Salesman:</b> {quotation.salesman.salesman_name}', styles['Normal'])])
    else:
        customer_data.append([Paragraph('', styles['Normal'])])

    customer_info_table = Table(customer_data, colWidths=[main_table_width / 2])
    customer_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), PRIMARY_COLOR), # Use Red for headers
    ]))

    info_table = Table([[quotation_info_table, customer_info_table]], colWidths=[main_table_width / 2, main_table_width / 2])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.2 * inch))

    # --- Items Table ---
    items_header = ['#', 'UPC Code', 'Description', 'Qty', 'Unit Price', 'Total']
    items_data = [items_header]
    subtotal = 0.0

    for idx, item in enumerate(quotation_items, 1):
        line_total = item.quantity * item.price
        subtotal += line_total

        desc_style = styles['ItemDescription']
        # Set text color to black/grey for readability
        desc_style.textColor = black 
        
        description_text = getattr(item.item, 'item_description', 'No description available')
        description_para = Paragraph(description_text, desc_style)

        upc_raw = getattr(item.item, 'item_upvc', '')
        if upc_raw is None: upc_raw = ""
        upc_para = Paragraph(str(upc_raw), styles['ItemCode'])

        items_data.append([
            str(idx),
            upc_para,
            description_para,
            f"{item.quantity} {item.unit}",
            f"AED {item.price:,.2f}",
            f"AED {line_total:,.2f}"
        ])

    items_table = Table(
        items_data,
        colWidths=[
            main_table_width * 0.05,
            main_table_width * 0.15,
            main_table_width * 0.43,
            main_table_width * 0.07,
            main_table_width * 0.15,
            main_table_width * 0.15
        ],
        repeatRows=1
    )
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), SECONDARY_COLOR), # Black Header
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [ROW_BG_COLOR, white]),
        ('ALIGN', (0, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),  # Set row content fontsize to 9
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 0.1 * inch))

    # --- Summary Table ---
    tax_rate = 0.05
    tax_amount = subtotal * tax_rate
    grand_total = subtotal + tax_amount

    summary_data = [
        ['Subtotal:', f"AED {subtotal:,.2f}"],
        [f'VAT ({tax_rate:.0%}):', f"AED {tax_amount:,.2f}"],
        ['Grand Total:', f"AED {grand_total:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[main_table_width * 0.5, main_table_width * 0.5])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('BACKGROUND', (0, 2), (-1, 2), PRIMARY_COLOR), # Red Total Background
        ('TEXTCOLOR', (0, 2), (-1, 2), white),
    ]))
    
    summary_wrapper = Table([[summary_table]], colWidths=[main_table_width])
    summary_wrapper.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    elements.append(KeepTogether(summary_wrapper))
    elements.append(Spacer(1, 0.3 * inch))
    
    # --- Remarks & Terms ---
    if hasattr(quotation, 'remarks') and quotation.remarks:
        remarks_section = [
            Paragraph("Remarks:", styles['h3']),
            Paragraph(quotation.remarks, styles['Normal']),
            Spacer(1, 0.2 * inch)
        ]
        elements.extend(remarks_section)

    terms_section = [
        Paragraph("Terms & Conditions:", styles['h3'])
    ]
    terms = [
        "1. This quotation is valid as Agreed.",
        "2. Prices are subject to change without prior notice after the validity period.",
        "3. Delivery timelines will be confirmed upon order confirmation.",
        "4. This is a system-generated document and does not require a signature.",
    ]
    for term in terms:
        terms_section.append(Paragraph(term, styles['Normal']))
    
    elements.extend(terms_section)
    
    doc.multiBuild(elements)