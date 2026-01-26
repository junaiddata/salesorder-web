from django.shortcuts import render
import openpyxl
from openpyxl.styles import Font
from django.contrib.auth import logout, authenticate, login
# Create your views here .
# Upload Items file from .xlsx file which contains item_code, item_description, item_firm
from django.http import HttpResponse
import pandas as pd
from .models import Items,Customer
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, redirect
from .models import SalesOrder, OrderItem
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from .models import Customer, Items, SalesOrder, OrderItem,Salesman,CustomerPrice,Role
import datetime
from django.contrib.auth.decorators import login_required
from functools import wraps
from django.db import IntegrityError
from django.http import HttpResponseForbidden
from django.views.decorators.http import require_POST
from .forms import CustomerForm, ItemForm
#import reverse
from django.db import transaction
from django.urls import reverse
from django.http import JsonResponse
from .utils import send_telegram_message 
from datetime import date, datetime, timedelta
from django.db.models.functions import Coalesce
from django.db.models import Sum, Value, DecimalField
from django.contrib.auth.decorators import login_required

def role_required(*required_roles):  # Accept multiple roles
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return HttpResponseForbidden("You are not authorized to view this page.")
            
            try:
                user_role = request.user.role.role  # Assuming `role` is a related object
                if user_role in required_roles:  # Check if user role is in allowed roles
                    return view_func(request, *args, **kwargs)
            except AttributeError:  # If role doesn't exist
                pass

            return HttpResponseForbidden("You are not authorized to view this page.")
        return _wrapped_view
    return decorator
    



def upload_items(request):
    if request.method == 'POST':
        file = request.FILES['file']
        df = pd.read_excel(file)
        # df['item_code'] = df['item_code'].astype(str).str.rstrip('.0')
        # df['item_upvc'] = df['item_upvc'].astype(str).str.rstrip('.0')
        df['item_code'] = df['item_code'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
        df['item_upvc'] = df['item_upvc'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))

        for index, row in df.iterrows():
            item_code = row['item_code']
            item_description = row['item_description']
            item_upvc = row.get('item_upvc', '')  # Default to empty string if not provided
            item_cost = row.get('item_cost', 0.00)  # Default to 0.00 if not provided
            item_firm = row['item_firm']
            item_price = row.get('item_price', 0.00)  # Default to 0.00 if not provided  # Default to 0 if not provided
            item_stock = row.get('item_stock', 0)  # Default to 0 if not provided
            
            Items.objects.update_or_create(
                item_code=item_code,
                defaults={
                    'item_description': item_description,
                    'item_firm': item_firm,
                    'item_price': item_price,
                    'item_upvc': item_upvc,  # Save UPVC if provided
                    'item_cost': item_cost,  # Save cost if provided
                    'item_stock': item_stock  # Save stock if provided
                }
            )
        
        return HttpResponse("Items uploaded successfully.")
    
    return render(request, 'so/upload_items.html')


# In this code, we define a view `upload_items` that handles the file upload and processes the items from an Excel file.

# View for uploading customer data
def upload_customers(request):
    if request.method == 'POST':
        file = request.FILES['file']
        df = pd.read_excel(file)
        
        # Normalize column names (handle variations)
        df.columns = df.columns.str.strip()
        
        # Map possible column name variations
        code_col = None
        name_col = None
        vat_col = None
        salesman_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'bp code' in col_lower or 'customer_code' in col_lower or 'customer code' in col_lower:
                code_col = col
            elif 'bp name' in col_lower or 'customer_name' in col_lower or 'customer name' in col_lower:
                name_col = col
            elif 'vat' in col_lower and 'number' in col_lower:
                vat_col = col
            elif 'salesman' in col_lower:
                salesman_col = col
        
        if not code_col:
            return HttpResponse("Error: Could not find 'BP Code' or 'Customer Code' column in Excel file.", status=400)
        if not name_col:
            return HttpResponse("Error: Could not find 'BP Name' or 'Customer Name' column in Excel file.", status=400)
        
        updated_count = 0
        created_count = 0
        
        for index, row in df.iterrows():
            customer_code = str(row[code_col]).strip() if pd.notna(row[code_col]) else None
            customer_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else None
            vat_number = str(row[vat_col]).strip() if vat_col and pd.notna(row.get(vat_col)) else None
            salesman_name = str(row[salesman_col]).strip() if salesman_col and pd.notna(row.get(salesman_col)) else None

            if not customer_code or not customer_name:
                continue

            # Get or create the salesman
            salesman = None
            if salesman_name:
                salesman, _ = Salesman.objects.get_or_create(salesman_name=salesman_name)

            # Update or create the customer
            customer, created = Customer.objects.update_or_create(
                customer_code=customer_code,
                defaults={
                    'customer_name': customer_name,
                    'salesman': salesman,
                    'vat_number': vat_number if vat_number else None,
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1

        return HttpResponse(f"Upload successful! Created: {created_count}, Updated: {updated_count} customers.")
    
    return render(request, 'so/upload_customers.html')

from django.http import JsonResponse
from django.db.models import Q
import json

def get_unique_firms():
    return Items.objects.values_list('item_firm', flat=True).distinct()



@csrf_exempt
@transaction.atomic
def create_sales_order(request):
    if request.method == 'POST':

        division = 'JUNAID' # Default
    
    # OPTION A: Your suggested Username check (Simplest)
        if 'alabama' in request.user.username.lower():
            division = 'ALABAMA'
        try:
            customer_id = request.POST.get('customer')
            new_customer_name = request.POST.get('new_customer_name', '').strip()

            # Validate that we have either an existing customer or a new customer name
            if not customer_id and not new_customer_name:
                messages.error(request, 'Please select a customer or enter a new customer name.')
                return redirect('create_sales_order')

            # Handle customer creation/selection
            if new_customer_name:
                # Find the last NEWCUSTOMER code
                last_customer = Customer.objects.filter(customer_code__startswith='NEWCUSTOMER') \
                                                .order_by('-id').first()
                if last_customer and last_customer.customer_code[11:].isdigit():
                    last_number = int(last_customer.customer_code[11:])
                else:
                    last_number = 0

                new_code = f'NEWCUSTOMER{last_number + 1}'

                # Assign the selected salesman to the new customer
                salesman_id = request.POST.get('salesman')
                if not salesman_id:
                    messages.error(request, 'Salesman is required when creating a new customer.')
                    return redirect('create_sales_order')
                    
                selected_salesman = get_object_or_404(Salesman, id=salesman_id)

                customer, created = Customer.objects.get_or_create(
                    customer_name=new_customer_name,
                    defaults={
                        'customer_code': new_code,
                        'salesman': selected_salesman
                    }
                )
            else:
                customer = get_object_or_404(Customer, id=customer_id)

            salesman_id = request.POST.get('salesman')
            salesman = get_object_or_404(Salesman, id=salesman_id) if salesman_id else None
            lpo_image = request.FILES.get('lpo_image')
            location = request.POST.get('location', '').strip()
            salesman_remarks = request.POST.get('salesman_remarks', '').strip()

            # Validate items
            item_ids = request.POST.getlist('item')
            quantities = request.POST.getlist('quantity')
            prices = request.POST.getlist('price')
            units = request.POST.getlist('unit')  # Get units list
            
            if not item_ids:
                messages.error(request, 'Please add at least one item to the order.')
                return redirect('create_sales_order')
                
            if len(item_ids) != len(quantities) or len(item_ids) != len(prices) or len(item_ids) != len(units):
                messages.error(request, 'Invalid form data. Please try again.')
                return redirect('create_sales_order')

            # Create the sales order
            sales_order = SalesOrder.objects.create(
                customer=customer,
                division=division,
                salesman=salesman,
                lpo_image=lpo_image,
                location=location,
                salesman_remarks=salesman_remarks
            )

            # Process order items
            order_items = []
            customer_price_updates = []

            for i, (item_id, qty, price, unit) in enumerate(zip(item_ids, quantities, prices, units)):
                try:
                    item = Items.objects.get(id=item_id)
                    price_val = float(price) if price else float(item.item_price)
                    quantity_val = int(qty)
                    unit_val = unit if unit in ['pcs', 'ctn','roll'] else 'pcs'  # Validate unit
                    
                    if quantity_val <= 0:
                        messages.error(request, f'Quantity must be positive for item {i+1}.')
                        return redirect('create_sales_order')
                        
                    if price_val < 0:
                        messages.error(request, f'Price cannot be negative for item {i+1}.')
                        return redirect('create_sales_order')

                    order_items.append(OrderItem(
                        order=sales_order,
                        item=item,
                        quantity=quantity_val,
                        price=price_val,
                        unit=unit_val,  # Add unit field
                        is_custom_price=bool(price)
                    ))

                    if price_val:
                        customer_price_updates.append((customer, item, price_val))
                        
                except (ValueError, Items.DoesNotExist) as e:
                    messages.error(request, f'Invalid data for item {i+1}.')
                    return redirect('create_sales_order')

            # Bulk create order items
            OrderItem.objects.bulk_create(order_items)

            # Update customer prices
            for customer, item, price in customer_price_updates:
                CustomerPrice.objects.update_or_create(
                    customer=customer,
                    item=item,
                    defaults={'custom_price': price}
                )

            # Calculate totals
            total = sum(oi.quantity * oi.price for oi in order_items)
            tax = round(0.05 * total, 2)
            grand_total = round(total + tax, 2)

            sales_order.tax = tax
            sales_order.total_amount = total
            sales_order.save()

            # Send notification
            msg = (
                f"🆕 New Sales Order Created\n"
                f"Customer: {sales_order.customer.customer_name}\n"
                f"Location: {sales_order.location}\n"
                f"Remarks: {sales_order.salesman_remarks}\n"
                f"Amount: {sales_order.total_amount}\n\n"
                f"Please review and approve."
            )
            send_telegram_message(settings.TELEGRAM_CREATE_CHAT_ID, msg)

            messages.success(request, 'Sales order created successfully!')
            return redirect('view_sales_orders')

        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('create_sales_order')

    # GET request - show form
    # -------------------------------------------------------------------------
    # GET request - show form (With Multiple Salesman Filtering Logic)
    # -------------------------------------------------------------------------
    salesmen = Salesman.objects.all()

    # Filter Logic: Only apply if user is logged in AND is NOT a superuser/admin
    if request.user.is_authenticated and not request.user.is_superuser:
        current_username = request.user.username.lower()

        # Define Mapping: 'username' -> List of ['Name1', 'Name2']
        user_salesman_map = {
            'alabamakadhar': ['KADER'],
            'alabamamusharaf': ['MUSHARAF'],   # Multiple allowed
            'alabamaadmin': ['KADER','MUSHARAF','AIJAZ','CASH'],               # Single allowed
            
        }

        if current_username in user_salesman_map:
            target_names = user_salesman_map[current_username]
            
            # Build a complex query: (name contains A) OR (name contains B)
            query = Q()
            for name in target_names:
                query |= Q(salesman_name__icontains=name)
            
            salesmen = salesmen.filter(query)

    # -------------------------------------------------------------------------
    firms = Items.objects.values_list('item_firm', flat=True).distinct()
    customers = Customer.objects.all()

    return render(request, 'so/create_sales_order.html', {
        'salesmen': salesmen,
        'firms': firms,
        'customers': customers,
    })


from django.http import JsonResponse
from django.db.models import Q

def items_search(request):
    query = request.GET.get('q', '')
    
    if query:
        items = Items.objects.filter(
            Q(item_code__icontains=query) |
            Q(item_description__icontains=query) |
            Q(item_upvc__icontains=query)
        )[:20]  # Limit results for performance
    else:
        items = Items.objects.all()[:20]
    
    results = []
    for item in items:
        results.append({
            'id': item.id,
            'item_code': item.item_code,
            'item_description': item.item_description,
            'item_upvc': item.item_upvc,
            'item_stock': item.item_stock,
            'text': f"{item.item_description}"
        })
    
    return JsonResponse(results, safe=False)

def get_item_details(request):
    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Item ID required'}, status=400)
    
    try:
        item = Items.objects.get(id=item_id)
        return JsonResponse({
            'stock': item.item_stock,
            'cost': item.item_cost if hasattr(item, 'item_cost') else 'N/A'
        })
    except Items.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)

from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from so.models import SalesOrder, OrderItem, Customer, Salesman, Items, CustomerPrice

@csrf_exempt
@transaction.atomic
def edit_sales_order(request, order_id):
    sales_order = get_object_or_404(SalesOrder, id=order_id)

    if request.method == 'POST':
        customer = get_object_or_404(Customer, id=request.POST.get('customer'))
        salesman = get_object_or_404(Salesman, id=request.POST.get('salesman')) if request.POST.get('salesman') else None
        lpo_image = request.FILES.get('lpo_image') or sales_order.lpo_image

        # Update order fields
        sales_order.customer = customer
        sales_order.salesman = salesman
        sales_order.lpo_image = lpo_image
        sales_order.save()

        # Remove old items
        sales_order.items.all().delete()

        # Re-create items
        item_ids = request.POST.getlist('item')
        quantities = request.POST.getlist('quantity')
        prices = request.POST.getlist('price')
        units = request.POST.getlist('unit') 

        order_items = []
        for item_id, qty, price, unit in zip(item_ids, quantities, prices, units):
            item = get_object_or_404(Items, id=item_id)
            price = float(price) if price else float(item.item_price)
            quantity = int(qty)
            unit_val = unit if unit in ['pcs', 'ctn', 'roll'] else 'pcs'

            order_items.append(OrderItem(
                order=sales_order,
                item=item,
                quantity=quantity,
                price=price,
                unit=unit_val,
                is_custom_price=True
            ))

        OrderItem.objects.bulk_create(order_items)

        # Recalculate totals
        total = sum(oi.quantity * oi.price for oi in order_items)
        tax = round(0.05 * total, 2)
        sales_order.total_amount = total
        sales_order.tax = tax
        sales_order.save()

        return redirect('view_sales_order_details', order_id=sales_order.id)

    # -------------------------------------------------------------------------
    # GET request → render form with existing data (With Filtering Logic)
    # -------------------------------------------------------------------------
    salesmen = Salesman.objects.all()

    # Filter Logic: Only apply if user is logged in AND is NOT a superuser/admin
    if request.user.is_authenticated and not request.user.is_superuser:
        current_username = request.user.username.lower()

        # Define Mapping: 'username' -> List of ['Name1', 'Name2']
        user_salesman_map = {
            'alabamakadhar': ['KADER'],
            'alabamamusharaf': ['MUSHARAF'],   # Multiple allowed
            'alabamaadmin': ['KADER','MUSHARAF','AIJAZ','CASH'],               # Single allowed
            
        }

        if current_username in user_salesman_map:
            target_names = user_salesman_map[current_username]
            
            # Build a complex query: (name contains A) OR (name contains B)
            query = Q()
            for name in target_names:
                query |= Q(salesman_name__icontains=name)
            
            salesmen = salesmen.filter(query)

    # -------------------------------------------------------------------------
    
    firms = Items.objects.values_list('item_firm', flat=True).distinct()

    return render(request, 'so/edit_sales_order.html', {
        'sales_order': sales_order,
        'salesmen': salesmen,
        'firms': firms,
        'order_items': sales_order.items.all()
    })
# @csrf_exempt
# def create_sales_order(request):
#     if request.method == 'POST':
#         customer_id = request.POST.get('customer')
#         customer = get_object_or_404(Customer, id=customer_id)
#         salesman_id = request.POST.get('salesman')
#         salesman = get_object_or_404(Salesman, id=salesman_id) if salesman_id else None
#         lpo_image = request.FILES.get('lpo_image')

#         sales_order = SalesOrder.objects.create(
#             customer=customer,
#             salesman=salesman,
#             lpo_image=lpo_image,  # Save the uploaded LPO image
#         )

#         item_ids = request.POST.getlist('item')
#         quantities = request.POST.getlist('quantity')
#         prices = request.POST.getlist('price')  # New field for custom prices

#         for item_id, qty, price in zip(item_ids, quantities, prices):
#             item = get_object_or_404(Items, id=item_id)
#             custom_price = price if price else item.item_price

#             OrderItem.objects.create(
#                 order=sales_order,
#                 item=item,
#                 quantity=int(qty),
#                 price=custom_price,
#                 is_custom_price=bool(price)  # True if custom price was entered
#             )
#             if price:
#                 CustomerPrice.objects.update_or_create(
#                     customer=customer,
#                     item=item,
#                     defaults={'custom_price': custom_price}
#                 )

#             # IF YOU WANT THE HIGHEST CUSTOM PRICE TO BE SAVED
#             # if existing_cp:
#             #     if custom_price > existing_cp.custom_price:
#             #         existing_cp.custom_price = custom_price
#             #         existing_cp.save()
#             # else:
#             #     CustomerPrice.objects.create(
#             #         customer=customer,
#             #         item=item,
#             #         custom_price=custom_price
#             #     )

#         total = sum(oi.quantity * oi.price for oi in sales_order.items.all())
#         tax = round(0.05 * total, 2)
#         grand_total = round(total + tax,2)
#         sales_order.tax = tax
#         sales_order.total_amount = total
#         sales_order.save()

#         return redirect('view_sales_orders')

#     salesmen = Salesman.objects.all()
#     firms = Items.objects.values_list('item_firm', flat=True).distinct()

#     return render(request, 'so/create_sales_order.html', {
#         'salesmen': salesmen,
#         'firms': firms,
#     })


def get_item_price(request):
    item_id = request.GET.get('item_id')
    customer_id = request.GET.get('customer_id')
    
    item = get_object_or_404(Items, id=item_id)
    
    # Try to get custom price if exists
    custom_price_obj = CustomerPrice.objects.filter(
        customer_id=customer_id,
        item_id=item_id
    ).first()
    
    default_price = item.item_price
    custom_price = custom_price_obj.custom_price if custom_price_obj else None

    # Decide final price: higher of custom or default
    final_price = custom_price if custom_price is not None and custom_price > item.item_cost else default_price

    return JsonResponse({
        'default_price': float(default_price),
        'custom_price': float(custom_price) if custom_price else None,
        'final_price': float(final_price),
    })

from django.db.models import Case, When, Value, IntegerField

def get_customers_by_salesman(request):
    salesman_id = request.GET.get('salesman_id')
    priority_customer_name = "A NEW CUSTOMER"  # replace with your customer name

    customers = Customer.objects.filter(salesman_id=salesman_id).annotate(
        priority_order=Case(
            When(customer_name=priority_customer_name, then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        )
    ).order_by('priority_order', 'customer_name').values('id', 'customer_name')

    return JsonResponse({'customers': list(customers)})

# def get_items_by_firm(request):
#     firm = request.GET.get('firm')

#     if firm == 'All' or not firm:
#         items = Items.objects.all()
#     else:
#         items = Items.objects.filter(item_firm=firm)
    
#     items = items.values('id', 'item_description', 'item_code', 'item_firm','item_upvc')

#     # items_data = items.values('id', 'item_description')
#     return JsonResponse({'items': list(items)})
from django.views.decorators.http import require_GET
from django.core.cache import cache
from django.http import JsonResponse

@require_GET
def get_items_by_firm(request):
    firm = request.GET.get('firm')

    # Don't cache since stock changes frequently
    # Always fetch fresh data from database
    if firm == 'All' or not firm:
        qs = Items.objects.all()
    else:
        qs = Items.objects.filter(item_firm=firm)

    items = list(qs.values('id', 'item_description', 'item_code', 'item_firm', 'item_upvc','item_stock'))

    return JsonResponse({'items': items})

def get_item_stock(request):
    item_id = request.GET.get('item_id')
    item = get_object_or_404(Items, id=item_id)
    return JsonResponse({'stock': item.item_stock, 'cost': float(item.item_cost)})


def format_whatsapp_order(sales_order, order_items, request):
    """
    Format sales order for WhatsApp with professional styling
    """
    # Calculate totals
    grand_total = sales_order.total_amount + sales_order.tax
    
    # Currency formatter
    def format_currency(amount):
        return f"{amount:,.2f} AED"
    
    # Build message sections
    sections = []
    
    # Header
    sections.append("🏢 *JUNAID GROUP*")
    sections.append("━" * 25)
    sections.append("")
    
    # Order info
    sections.append(f"📋 *CUSTOMER ORDER {sales_order.order_number}*")
    sections.append(
    f"📅 {sales_order.order_date.strftime('%d-%b-%Y') if sales_order.order_date else 'N/A'}"
)
    sections.append("")
    
    # Customer
    sections.append("👤 *Customer Details*")
    sections.append(f"Name: {sales_order.customer.customer_name}")
    if hasattr(sales_order.customer, 'phone'):
        sections.append(f"Phone: {sales_order.customer.phone}")
    sections.append("")
    
    # Items
    sections.append("📦 *Order Items*")
    sections.append("─" * 25)
    
    for idx, item in enumerate(order_items, 1):
        sections.append(f"\n*{idx}. {item.item.item_description}*")
        sections.append(f"   Qty: {item.quantity} × {format_currency(item.price)}")
        sections.append(f"   Amount: *{format_currency(item.line_total)}*")
    
    sections.append("\n" + "─" * 25)
    
    # Summary
    sections.append("\n💰 *Payment Summary*")
    sections.append(f"Subtotal: {format_currency(sales_order.total_amount)}")
    sections.append(f"VAT (5%): {format_currency(sales_order.tax)}")
    sections.append("─" * 25)
    sections.append(f"*TOTAL: {format_currency(grand_total)}*")
    
    # PDF Link
    pdf_url = request.build_absolute_uri(
        reverse('export_sales_order_to_pdf', args=[sales_order.id])
    )
    sections.append(f"\n📄 *Download PDF*")
    sections.append(f"👉 {pdf_url}")
    
    # Footer
    sections.append("\n✅ Thank you for your order!")
    
    return "\n".join(sections)
# A view to view the sales order first it will list all the sales orders and then it will show the details of a specific sales order
from django.shortcuts import render
from django.db.models import Q # Import Q for complex lookups
import datetime

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import render_to_string
from django.views.decorators.http import require_GET

def view_sales_orders(request):
    status = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    salesman_filter = request.GET.get('salesman_filter')
    division = request.GET.get('division', 'All')
    q = (request.GET.get('q') or '').strip()
    page = request.GET.get('page', 1)

    # Initial queryset - all sales orders
    sales_orders = SalesOrder.objects.all()

    if request.user.role.role == 'Admin' and request.user.username.lower() == 'alabamaadmin':
        sales_orders = SalesOrder.objects.filter(division='ALABAMA')
    elif request.user.role.role == 'Admin' and request.user.username.lower() not in ['so','manager']:
        sales_orders = SalesOrder.objects.filter(division='JUNAID')
    else:
        sales_orders = SalesOrder.objects.all()

    # Salesman restriction
    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
        salesman_name = request.user.first_name
        sales_orders = sales_orders.filter(salesman__salesman_name=salesman_name)
    elif salesman_filter and salesman_filter != 'All':
        # Apply salesman filter if it's not "All" and the user is not a salesman
        sales_orders = sales_orders.filter(salesman__salesman_name=salesman_filter)

    # Apply status filter
    if status and status != "All":
        sales_orders = sales_orders.filter(order_status=status)

    # Division filter (based on order_number prefix)
    if division and division != "All":
        div = (division or "").strip().upper()
        if div == "ALABAMA":
            sales_orders = sales_orders.filter(Q(order_number__istartswith="AL") | Q(division="ALABAMA"))
        elif div == "JUNAID":
            sales_orders = sales_orders.filter(Q(order_number__istartswith="CO") | Q(division="JUNAID"))

    # Search query
    if q:
        sales_orders = sales_orders.filter(
            Q(order_number__icontains=q)
            | Q(customer__customer_name__icontains=q)
            | Q(customer__customer_code__icontains=q)
            | Q(salesman__salesman_name__icontains=q)
            | Q(salesman_remarks__icontains=q)
        )

    # Apply date filters
    if start_date:
        sales_orders = sales_orders.filter(order_date__gte=start_date)
    
    if end_date:
        sales_orders = sales_orders.filter(order_date__lte=end_date)
    
    # Get all unique salesmen for the filter dropdown
    all_salesmen = Salesman.objects.all().order_by('salesman_name')

    # Pagination - 12 items per page (3x4 grid layout)
    paginator = Paginator(sales_orders.order_by('-order_date_time'), 12)
    
    try:
        sales_orders_page = paginator.page(page)
    except PageNotAnInteger:
        sales_orders_page = paginator.page(1)
    except EmptyPage:
        sales_orders_page = paginator.page(paginator.num_pages)

    # Build query string for pagination links
    query_params = []
    if status and status != "All":
        query_params.append(f"status={status}")
    if start_date:
        query_params.append(f"start_date={start_date}")
    if end_date:
        query_params.append(f"end_date={end_date}")
    if salesman_filter:
        query_params.append(f"salesman_filter={salesman_filter}")
    if division and division != "All":
        query_params.append(f"division={division}")
    if q:
        query_params.append(f"q={q}")
    
    query_string = "&".join(query_params)

    return render(request, 'so/view_sales_orders.html', {
        'sales_orders': sales_orders_page,
        'current_status': status or "All",
        'all_salesmen': all_salesmen,
        'selected_salesman': salesman_filter,
        'selected_division': division or "All",
        'search_query': q,
        'start_date': start_date,
        'end_date': end_date,
        'query_string': query_string,
    })


@login_required
@require_GET
def view_sales_orders_ajax(request):
    """
    AJAX endpoint for backend filtering/search/pagination on Customer Order Forms dashboard.
    Supports:
    - q (search)
    - status
    - start_date / end_date
    - salesman_filter
    - division: All | ALABAMA | JUNAID (prefix filter: AL/CO)
    - page
    """
    status = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    salesman_filter = request.GET.get('salesman_filter')
    division = request.GET.get('division', 'All')
    q = (request.GET.get('q') or '').strip()
    page = request.GET.get('page', 1)

    sales_orders = SalesOrder.objects.all()

    if request.user.role.role == 'Admin' and request.user.username.lower() == 'alabamaadmin':
        sales_orders = SalesOrder.objects.filter(division='ALABAMA')
    elif request.user.role.role == 'Admin' and request.user.username.lower() not in ['so', 'manager']:
        sales_orders = SalesOrder.objects.filter(division='JUNAID')
    else:
        sales_orders = SalesOrder.objects.all()

    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
        salesman_name = request.user.first_name
        sales_orders = sales_orders.filter(salesman__salesman_name=salesman_name)
    elif salesman_filter and salesman_filter != 'All':
        sales_orders = sales_orders.filter(salesman__salesman_name=salesman_filter)

    if status and status != "All":
        sales_orders = sales_orders.filter(order_status=status)

    if division and division != "All":
        div = (division or "").strip().upper()
        if div == "ALABAMA":
            sales_orders = sales_orders.filter(Q(order_number__istartswith="AL") | Q(division="ALABAMA"))
        elif div == "JUNAID":
            sales_orders = sales_orders.filter(Q(order_number__istartswith="CO") | Q(division="JUNAID"))

    if q:
        sales_orders = sales_orders.filter(
            Q(order_number__icontains=q)
            | Q(customer__customer_name__icontains=q)
            | Q(customer__customer_code__icontains=q)
            | Q(salesman__salesman_name__icontains=q)
            | Q(salesman_remarks__icontains=q)
        )

    if start_date:
        sales_orders = sales_orders.filter(order_date__gte=start_date)
    if end_date:
        sales_orders = sales_orders.filter(order_date__lte=end_date)

    paginator = Paginator(sales_orders.order_by('-order_date_time'), 12)
    try:
        sales_orders_page = paginator.page(page)
    except PageNotAnInteger:
        sales_orders_page = paginator.page(1)
    except EmptyPage:
        sales_orders_page = paginator.page(paginator.num_pages)

    query_params = []
    if status and status != "All":
        query_params.append(f"status={status}")
    if start_date:
        query_params.append(f"start_date={start_date}")
    if end_date:
        query_params.append(f"end_date={end_date}")
    if salesman_filter:
        query_params.append(f"salesman_filter={salesman_filter}")
    if division and division != "All":
        query_params.append(f"division={division}")
    if q:
        query_params.append(f"q={q}")
    query_string = "&".join(query_params)

    html = render_to_string(
        'so/_sales_orders_results.html',
        {
            'sales_orders': sales_orders_page,
            'current_status': status or "All",
            'selected_salesman': salesman_filter,
            'selected_division': division or "All",
            'start_date': start_date,
            'end_date': end_date,
            'query_string': query_string,
        },
        request=request
    )

    return JsonResponse({
        'html': html,
        'count': paginator.count,
    })

from django.http import JsonResponse

def mark_so_created(request, order_id):
    if request.method == "POST":
        sales_order = get_object_or_404(SalesOrder, id=order_id)
        sales_order.order_status = "SO Created"
        sales_order.save()
        return JsonResponse({"success": True, "new_status": "SO Created"})
    return JsonResponse({"success": False}, status=400)

from datetime import date

# def view_sales_order_details(request, order_id):
#     sales_order = get_object_or_404(SalesOrder, id=order_id)

#     if request.method == 'POST':
#         sales_order.order_taken = True
#         sales_order.save()
#         print("Marked as taken!")  # Debug
#         return redirect('view_sales_order_details', order_id=order_id)

#     grand_total = sales_order.total_amount + sales_order.tax
#     order_items = sales_order.items.all()

#     has_undercost_items = False
#     has_over_limit = False

#     for item in order_items:
#         item.line_total = item.quantity * item.price
#         item.is_undercost = item.price < item.item.item_price
#         if item.is_undercost:
#             has_undercost_items = True

#     # 🔶 Set default values in case API fails
#     payment_terms = None
#     monthly_pending_data = None
#     credit_limit = 0
#     pending_total = 0
#     current_month_receivables = 0
#     pdf_message = format_whatsapp_order(sales_order, order_items, request)  # Default message

#     # 🔶 Try fetching API data
#     try:
#         api_url = f"http://192.168.2.44:8000/api/monthly-pending/{sales_order.customer.customer_code}/?year={date.today().year}"
#         response = requests.get(api_url)
#         response.raise_for_status()

#         data = response.json()
#         payment_terms = data.get("payment_term")
#         monthly_pending_data = data.get("monthly_pending", [])
#         credit_limit = float(data.get("credit_limit", 0))
#         pending_total = float(data.get("pending_total", 0))
#         current_month_receivables = float(data.get("current_month_receivables", 0))

#         if (pending_total + grand_total) > credit_limit:
#             has_over_limit = True
#             pdf_message = format_whatsapp_order(sales_order, order_items, request)  # Only override if over limit

#     except Exception as e:
#         print("❌ API error or unreachable:", e)
#         # All values stay as safe defaults

#     return render(request, 'so/view_sales_order_details.html', {
#         'sales_order': sales_order,
#         'order_items': order_items,
#         'grand_total': grand_total,
#         'has_undercost_items': has_undercost_items,
#         'pdf_message': pdf_message,
#         'monthly_pending_data': monthly_pending_data,
#         'credit_limit': credit_limit,
#         'payment_terms': payment_terms,
#         'has_over_limit': has_over_limit,
#         'current_month_receivables': current_month_receivables,
#         'pending_total': pending_total + grand_total,
#     })

from datetime import date, timedelta
import calendar

def get_last_six_months():
    """Returns list of last 6 months (latest first), e.g. ['Jul', 'Jun', ..., 'Feb']"""
    today = date.today()
    months = []
    for i in range(6):
        month = (today.month - i - 1) % 12 + 1
        year = today.year if today.month - i > 0 else today.year - 1
        months.append((calendar.month_abbr[month], f'month_pending_{6 - i}'))  # ('Jul', 'month_pending_6')
    return list(reversed(months))  # So oldest first like Feb → Jul


def view_sales_order_details(request, order_id):
    sales_order = get_object_or_404(SalesOrder, id=order_id)
    customer = sales_order.customer

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            sales_order.order_status = 'Approved'
            sales_order.order_taken = True
            sales_order.approved_by = request.user   # 🔹 Store approver
            sales_order.save()

            msg = (
            f"🆕 Sales Order Approved\n"
            f"Customer: {sales_order.customer.customer_name}\n"
            f"Salesman: {sales_order.salesman.salesman_name if sales_order.salesman else 'N/A'}"
            f"Location: {sales_order.location}\n"
            f"Amount: {sales_order.total_amount}\n\n"
            f"Approved by: {request.user.get_full_name() or request.user.username}\n"
            f"Please proceed."
            )
                # Notify Person B
            send_telegram_message(
                settings.TELEGRAM_APPROVE_CHAT_ID,msg
            )

            return redirect('view_sales_order_details', order_id=order_id)
        elif action == 'hold':
            sales_order.order_status = 'Hold by A/c'
            sales_order.order_taken = False
            messag = "Order put on hold by A/c."
            sales_order.save()
            return redirect('view_sales_order_details', order_id=order_id)
        elif action == 'update_remarks':
            # Handle remarks update
            sales_order.remarks = request.POST.get('remarks', '')
            sales_order.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # For AJAX requests
                return JsonResponse({
                    'success': True,
                    'remarks': sales_order.remarks
                })
            else:
                # For regular form submissions
                messages.success(request, 'Remarks updated successfully!')
                return redirect('view_sales_order_details', order_id=order_id)


    grand_total = sales_order.total_amount + sales_order.tax
    order_items = sales_order.items.all()

    has_undercost_items = False
    has_over_limit = False



    for item in order_items:
        item.line_total = item.quantity * item.price
        item.is_undercost = item.price < ( item.item.item_cost + (item.item.item_cost * 0.03) )  # Assuming 3% margin
        if item.is_undercost:
            has_undercost_items = True

    credit_limit = customer.credit_limit
    payment_terms = customer.credit_days
    current_month_receivables = customer.pdc_received
    pending_total = customer.total_outstanding_with_pdc
    total_with_new_order = pending_total + grand_total
    old_months = customer.old_months_pending

    if total_with_new_order > credit_limit:
        has_over_limit = True

    # if not has_over_limit and not has_undercost_items:
    #     sales_order.order_taken = True
    #     sales_order.order_status = 'Approved'
    #     messag="Order Approved Automatically as there are no undercost items and total is within credit limit."
    #     sales_order.save()


    # 🟩 Dynamically generate monthly pending data and labels
    last_six = get_last_six_months()
    monthly_pending_data = []
    for month_abbr, field in last_six:
        monthly_pending_data.append({
            'month_abbr': month_abbr,
            'pending_total': getattr(customer, field, 0.0)
        })

    # Add old months as 7th entry
    monthly_pending_data.append({
        'month_abbr': '6+ mo',
        'pending_total': customer.old_months_pending
    })

    # Fill to make 12 months if needed
    while len(monthly_pending_data) < 12:
        monthly_pending_data.append({'month_abbr': '', 'pending_total': 0.0})

    pdf_message = format_whatsapp_order(sales_order, order_items, request)

    return render(request, 'so/view_sales_order_details.html', {
        'sales_order': sales_order,
        'order_items': order_items,
        'grand_total': grand_total,
        'has_undercost_items': has_undercost_items,
        'pdf_message': pdf_message,
        'monthly_pending_data': monthly_pending_data,
        'credit_limit': credit_limit,
        'payment_terms': payment_terms,
        'has_over_limit': has_over_limit,
        'current_month_receivables': current_month_receivables,
        'pending_total': total_with_new_order,
        'old_months': old_months,
        'pending_total_without': pending_total,
        'messag': messag if 'messag' in locals() else None,
        # No need to explicitly pass remarks as it's part of sales_order
    })

def export_sales_order_to_excel(request, order_id):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    
    sales_order = get_object_or_404(SalesOrder, id=order_id)
    order_items = sales_order.items.all()

    # Create a workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Order_{sales_order.order_number}"

    # Define styles
    # Company header style
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sub-header style
    subheader_font = Font(name='Arial', size=12, bold=True)
    subheader_alignment = Alignment(horizontal="left", vertical="center")
    
    # Table header style
    table_header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    table_header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin')
    )
    
    # Number format for currency
    currency_format = '#,##0.00'
    
    # Start building the document
    row_num = 1
    
    # Company Header
    ws.merge_cells(f'A{row_num}:F{row_num}')
    company_cell = ws[f'A{row_num}']
    company_cell.value = "CUSTOMER ORDER FORM"
    company_cell.font = header_font
    company_cell.fill = header_fill
    company_cell.alignment = header_alignment
    ws.row_dimensions[row_num].height = 30
    row_num += 2
    
    # Order Information Section
    order_info = [
        ("Order Number:", f"{sales_order.order_number}"),
        ("Order Date:", sales_order.order_date.strftime("%d-%m-%Y")),
        ("Customer Code:", sales_order.customer.customer_code),
        ("Customer:", sales_order.customer.customer_name),
        ("Salesman:", sales_order.salesman.salesman_name if sales_order.salesman else "N/A"),
    ]
    
    # Add customer details if available
    if hasattr(sales_order.customer, 'address'):
        order_info.append(("Customer Address:", getattr(sales_order.customer, 'address', 'N/A')))
    if hasattr(sales_order.customer, 'phone'):
        order_info.append(("Phone:", getattr(sales_order.customer, 'phone', 'N/A')))
    
    # Write order information
    info_start_row = row_num
    for label, value in order_info:
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = Font(bold=True, size=10)
        ws[f'B{row_num}'] = value
        ws[f'B{row_num}'].font = Font(size=10)
        ws.merge_cells(f'B{row_num}:D{row_num}')
        row_num += 1
    
    row_num += 1  # Empty row
    
    # Items Table Header
    table_start_row = row_num
    headers = ['S.No', 'Item Code', 'Item Description', 'Quantity', 'unit','Unit Price', 'Total']
    col_widths = [8, 15, 35, 12, 15, 15]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_alignment
        cell.border = thin_border
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    row_num += 1
    
    # Add item rows
    subtotal = 0.00
    for idx, item in enumerate(order_items, 1):
        line_total = item.quantity * item.price
        subtotal += line_total
        
        row_data = [
            idx,
            item.item.item_code,
            item.item.item_description,
            item.quantity,
            item.unit,
            float(item.price),
            float(line_total)
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num in [1, 4] else "left", 
                                     vertical="center")
            
            # Format numbers
            if col_num in [5, 6]:  # Price columns
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 4:  # Quantity
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num += 1
    
    # Summary Section
    row_num += 1
    
    # Subtotal
    ws[f'E{row_num}'] = "Subtotal:"
    ws[f'E{row_num}'].font = Font(bold=True, size=10)
    ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
    
    subtotal_cell = ws[f'F{row_num}']
    subtotal_cell.value = float(subtotal)
    subtotal_cell.number_format = currency_format
    subtotal_cell.font = Font(bold=True, size=10)
    subtotal_cell.alignment = Alignment(horizontal="right")
    subtotal_cell.border = thick_border
    row_num += 1
    
    # Tax (if applicable)
    if hasattr(sales_order, 'tax_amount') and sales_order.tax_amount:
        ws[f'E{row_num}'] = "Tax:"
        ws[f'E{row_num}'].font = Font(size=10)
        ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
        
        tax_cell = ws[f'F{row_num}']
        tax_cell.value = float(sales_order.tax_amount)
        tax_cell.number_format = currency_format
        tax_cell.alignment = Alignment(horizontal="right")
        tax_cell.border = thin_border
        row_num += 1
    
    # Total
    ws[f'E{row_num}'] = "Total Amount:"
    ws[f'E{row_num}'].font = Font(bold=True, size=12)
    ws[f'E{row_num}'].alignment = Alignment(horizontal="right")
    
    total_cell = ws[f'F{row_num}']
    total_cell.value = float(sales_order.total_amount)
    total_cell.number_format = currency_format
    total_cell.font = Font(bold=True, size=12, color="2C3E50")
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_cell.border = thin_border
    
    # Add footer notes
    row_num += 3
    notes_row = row_num
    ws[f'A{notes_row}'] = "Notes:"
    ws[f'A{notes_row}'].font = Font(bold=True, size=10)
    row_num += 1
    
    # Add any order notes
    if hasattr(sales_order, 'notes') and sales_order.notes:
        ws[f'A{row_num}'] = sales_order.notes
        ws.merge_cells(f'A{row_num}:F{row_num + 2}')
        ws[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical="top")
        row_num += 3
    
    # # Terms and conditions
    # row_num += 1
    # ws[f'A{row_num}'] = "Terms & Conditions:"
    # ws[f'A{row_num}'].font = Font(bold=True, size=9)
    # row_num += 1
    
    # terms = [
    #     "1. Goods once sold will not be taken back.",
    #     "2. Subject to jurisdiction only.",
    #     "3. Our risk and responsibility ceases as soon as goods leave our premises."
    # ]
    
    # for term in terms:
    #     ws[f'A{row_num}'] = term
    #     ws[f'A{row_num}'].font = Font(size=8)
    #     ws.merge_cells(f'A{row_num}:F{row_num}')
    #     row_num += 1
    
    # # Signature section
    # row_num += 3
    # ws[f'A{row_num}'] = "Authorized Signature"
    # ws[f'A{row_num}'].font = Font(size=10)
    # ws[f'A{row_num}'].alignment = Alignment(horizontal="center")
    # ws.merge_cells(f'A{row_num}:B{row_num}')
    
    # ws[f'E{row_num}'] = "Customer Signature"
    # ws[f'E{row_num}'].font = Font(size=10)
    # ws[f'E{row_num}'].alignment = Alignment(horizontal="center")
    # ws.merge_cells(f'E{row_num}:F{row_num}')
    
    # # Add line above signatures
    # row_num -= 1
    # ws[f'A{row_num}'] = "_" * 25
    # ws[f'E{row_num}'] = "_" * 25
    # ws[f'A{row_num}'].alignment = Alignment(horizontal="center")
    # ws[f'E{row_num}'].alignment = Alignment(horizontal="center")
    
    # Print settings
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Set print area
    ws.print_area = f'A1:F{row_num + 1}'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"CustomerOrder_{sales_order.order_number}_{sales_order.order_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

def login_view(request):
    if request.user.is_authenticated:
        # Redirect based on user role
        if request.user.role.role == 'Salesman':
            return redirect('sales_home')
        else:
            return redirect('home')
    
    # Your normal login logic here
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # Redirect after login based on role
            if request.user.role.role == 'Salesman':
                return redirect('sales_home')
            else:
                return redirect('home')
        else:
            # Invalid login, show error
            return render(request, 'so/login.html', {'error': 'Invalid credentials'})
    
    return render(request, 'so/login.html')



def logout_view(request):
    logout(request)
    return redirect('login')


@role_required('Admin')
@login_required
def home(request):
    if request.user.username.lower() == 'alabamaadmin':
        total_orders = SalesOrder.objects.filter(division='ALABAMA').count()
    elif request.user.username.lower() not in ['so','manager']:
        total_orders = SalesOrder.objects.filter(division='JUNAID').count()
    else:
        total_orders = SalesOrder.objects.count()
    today = date.today()
    total_orders_today = SalesOrder.objects.filter(
        order_date__year=today.year,
        order_date__month=today.month,
        order_date__day=today.day
    ).count()
    total_customers = Customer.objects.count()
    yesterday = today - timedelta(days=1)
    orders_yesterday = SalesOrder.objects.filter(
        order_date__year=yesterday.year,
        order_date__month=yesterday.month,
        order_date__day=yesterday.day
    ).count()
    if orders_yesterday > 0:
        order_increase_pct = ((total_orders_today - orders_yesterday) / orders_yesterday) * 100
    else:
        order_increase_pct = 0
    # Redirect salesman to sales_home.html, others to home.html
    
    return render(request, 'so/home.html', {
        'total_orders': total_orders,
        'total_orders_today': total_orders_today,
        'total_customers': total_customers,
        'order_increase_pct': order_increase_pct
    })

def sales_home(request):
    if request.user.is_authenticated:
        return render(request, 'so/sales_home.html')
    else:
        return redirect('login')

def details(request):
    return render(request, 'so/SALESORDERWEB.html')
#import messages
from django.contrib import messages
def customer_list(request):
    customers = Customer.objects.all().order_by('customer_name')
    salesmen = Salesman.objects.all().order_by('salesman_name')
    return render(request, 'so/customers/customer_list.html', {
        'customers': customers,
        'form': CustomerForm(),
        'salesmen':salesmen  # Empty form for adding new customers
    })

@require_POST
def add_customer(request):
    form = CustomerForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Customer added successfully!')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")
    return redirect('customer_list')

@require_POST
def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    customer.delete()
    messages.success(request, 'Customer deleted successfully!')
    return redirect('customer_list')


#####################################  ITEM SECTION ################################################################
# List all items
def item_list(request):
    selected_firm = request.GET.get('firm')
    firms = Items.objects.values_list('item_firm', flat=True).distinct().order_by('item_firm')

    if selected_firm:
        items = Items.objects.filter(item_firm=selected_firm)[:200]
    else:
        items = Items.objects.all()[:200]

    return render(request, 'so/items/item_list.html', {
        'items': items,
        'firms': firms,
        'selected_firm': selected_firm
    })

# Add a new item
def item_create(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Item created successfully.")
            return redirect('item_list')
    else:
        form = ItemForm()
    return render(request, 'so/items/item_form.html', {'form': form, 'title': 'Add Item'})

# Edit an existing item
def item_edit(request, pk):
    item = get_object_or_404(Items, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Item updated successfully.")
            return redirect('item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'so/items/item_form.html', {'form': form, 'title': 'Edit Item'})

@csrf_exempt
def item_delete(request, pk):
    item = get_object_or_404(Items, pk=pk)
    if request.method == 'POST':
        item.delete()
        messages.success(request, "Item deleted successfully.")
        return redirect('item_list')


from django.http import JsonResponse
from django.template.loader import render_to_string

def item_list_ajax(request):
    firm = request.GET.get('firm', '')
    query = request.GET.get('q', '')

    items = Items.objects.all()
    if firm:
        items = items.filter(item_firm=firm)
    if query:
        items = items.filter(
            models.Q(item_code__icontains=query) |
            models.Q(item_description__icontains=query) |
            models.Q(item_firm__icontains=query) |
            models.Q(item_upvc__icontains=query)
        )
    items = items[:200]
    html = render_to_string('so/items/partials/item_table.html', {'items': items})
    return JsonResponse({'html': html})
#####################################################################################################################


############################################   REST API VIEWS   ####################################################
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from .serializers import *
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.permissions import AllowAny
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token


@method_decorator(csrf_exempt, name='dispatch')
class SalesOrderCreateView(APIView):
    permission_classes = [AllowAny]  # or adjust as needed

    def post(self, request):
        serializer = SalesOrderSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class CustomerBySalesmanView(APIView):
    def get(self, request):
        salesman_id = request.GET.get('salesman_id')
        customers = Customer.objects.filter(salesman_id=salesman_id)
        serializer = CustomerSerializer(customers, many=True)
        return Response(serializer.data)

# class ItemsByFirmView(APIView):
#     def get(self, request):
#         firm = request.GET.get('firm')
#         items = Items.objects.filter(item_firm=firm)
#         serializer = ItemsSerializer(items, many=True)
#         return Response(serializer.data)



class ItemPriceView(APIView):
    def get(self, request):
        item_id = request.GET.get('item_id')
        customer_id = request.GET.get('customer_id')
        item = Items.objects.get(id=item_id)
        custom_price = CustomerPrice.objects.filter(
            customer_id=customer_id,
            item_id=item_id
        ).first()
        return Response({
            "default_price": float(item.item_price),
            "custom_price": str(custom_price.custom_price) if custom_price else None
        })

class SalesmanList(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        qs = Salesman.objects.values('id', 'salesman_name')
        return Response(list(qs))

class CustomersBySalesman(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        salesman_id = request.GET.get('salesman_id')
        qs = Customer.objects.filter(salesman_id=salesman_id).values('id', 'customer_name')
        return Response(list(qs))

class UniqueFirms(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        firms = Items.objects.values_list('item_firm', flat=True).distinct()
        return Response(list(firms))


# class ItemsByFirm(APIView):
#     permission_classes = [AllowAny]
#     def get(self, request):
#         firm = request.GET.get('firm')
#         qs = Items.objects.filter(item_firm=firm).values('id', 'item_description')
#         return Response({'items': list(qs)})
@method_decorator(csrf_exempt, name='dispatch')
class ItemsByFirm(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        firm = request.GET.get('firm')
        
        if firm == 'All':
            # Return all items with firm information
            qs = Items.objects.all().values('id', 'item_description', 'item_firm')
            items = [
                {
                    'id': item['id'],
                    'item_description': item['item_description'],
                    'firm': item['item_firm']  # Include firm info
                }
                for item in qs
            ]
        else:
            # Return items for specific firm
            qs = Items.objects.filter(item_firm=firm).values('id', 'item_description')
            items = [
                {
                    'id': item['id'],
                    'item_description': item['item_description'],
                    'firm': firm  # Include the firm
                }
                for item in qs
            ]
        
        return Response({'items': items})
class ItemPriceView(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        item_id = request.GET.get('item_id')
        customer_id = request.GET.get('customer_id')
        item = Items.objects.get(id=item_id)
        custom = CustomerPrice.objects.filter(customer_id=customer_id, item_id=item_id).first()
        return Response({
            'default_price': float(item.item_price),
            'custom_price': str(custom.custom_price) if custom else None
        })



from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated

class SalesOrderListAPI(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        filter_status = request.GET.get("status")
        salesman_name = request.user.first_name

        qs = SalesOrder.objects.filter(salesman__salesman_name=salesman_name)

        if filter_status == "taken":
            qs = qs.filter(order_taken=True)
        elif filter_status == "pending":
            qs = qs.filter(order_taken=False)

        qs = qs.order_by('-order_date')
        from .serializers import SalesOrderListSerializer
        serializer = SalesOrderListSerializer(qs, many=True)
        return Response(serializer.data)

class SalesOrderDetailAPI(APIView):
    permission_classes = [AllowAny]
    def get(self, request, pk):
        from .serializers import SalesOrderDetailSerializer
        so = SalesOrder.objects.get(pk=pk)
        serializer = SalesOrderDetailSerializer(so)
        return Response(serializer.data)
    
    def post(self, request, pk):
        # This endpoint POST is for marking order as taken, just like your old logic
        so = SalesOrder.objects.get(pk=pk)
        mark = request.data.get("mark_taken")
        if mark:
            so.order_taken = True
            so.save()
        from .serializers import SalesOrderDetailSerializer

        serializer = SalesOrderDetailSerializer(so)
        return Response(serializer.data)


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get("username")
        password = request.data.get("password")

        user = authenticate(username=username, password=password)
        if user:
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                "token": token.key,
                "username": user.username,
                "first_name": user.first_name,
            })
        return Response({"error": "Invalid credentials"}, status=401)


###################### CUSTOMER CRED ##############
class CreateCustomerView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Create a new customer"""
        try:
            # Extract data from request
            customer_code = request.data.get('customer_code', '').strip()
            customer_name = request.data.get('customer_name', '').strip()
            salesman_id = request.data.get('salesman')
            
            # Validate required fields
            if not customer_code:
                return Response(
                    {'error': 'Customer code is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not customer_name:
                return Response(
                    {'error': 'Customer name is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if customer code already exists
            if Customer.objects.filter(customer_code=customer_code).exists():
                return Response(
                    {'error': 'Customer with this code already exists'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate salesman if provided
            salesman = None
            if salesman_id:
                try:
                    salesman = Salesman.objects.get(id=salesman_id)
                except Salesman.DoesNotExist:
                    return Response(
                        {'error': 'Invalid salesman selected'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Create customer
            customer = Customer.objects.create(
                customer_code=customer_code,
                customer_name=customer_name,
                salesman=salesman
            )
            
            # Serialize and return response
            serializer = CustomerSerializer(customer)
            return Response(
                {
                    'message': 'Customer created successfully',
                    'customer': serializer.data
                }, 
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            return Response(
                {'error': f'An error occurred: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CustomerListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get all customers with optional filtering"""
        try:
            # Get query parameters for filtering
            salesman_id = request.query_params.get('salesman_id')
            search = request.query_params.get('search', '').strip()
            
            # Base queryset
            queryset = Customer.objects.select_related('salesman').all()
            
            # Apply filters
            if salesman_id:
                queryset = queryset.filter(salesman_id=salesman_id)
            
            if search:
                queryset = queryset.filter(
                    models.Q(customer_name__icontains=search) |
                    models.Q(customer_code__icontains=search)
                )
            
            # Order by name
            queryset = queryset.order_by('customer_name')
            
            # Serialize
            serializer = CustomerSerializer(queryset, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            return Response(
                {'error': f'An error occurred: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CustomerDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        """Get customer details"""
        try:
            customer = Customer.objects.select_related('salesman').get(pk=pk)
            serializer = CustomerSerializer(customer)
            return Response(serializer.data)
        except Customer.DoesNotExist:
            return Response(
                {'error': 'Customer not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    def put(self, request, pk):
        """Update customer"""
        try:
            customer = Customer.objects.get(pk=pk)
            
            # Update fields if provided
            if 'customer_name' in request.data:
                customer.customer_name = request.data['customer_name'].strip()
            
            if 'customer_code' in request.data:
                new_code = request.data['customer_code'].strip()
                # Check if new code already exists (excluding current customer)
                if Customer.objects.exclude(pk=pk).filter(customer_code=new_code).exists():
                    return Response(
                        {'error': 'Customer with this code already exists'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer.customer_code = new_code
            
            if 'salesman' in request.data:
                salesman_id = request.data['salesman']
                if salesman_id:
                    try:
                        customer.salesman = Salesman.objects.get(id=salesman_id)
                    except Salesman.DoesNotExist:
                        return Response(
                            {'error': 'Invalid salesman selected'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    customer.salesman = None
            
            customer.save()
            
            serializer = CustomerSerializer(customer)
            return Response({
                'message': 'Customer updated successfully',
                'customer': serializer.data
            })
            
        except Customer.DoesNotExist:
            return Response(
                {'error': 'Customer not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'An error occurred: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def delete(self, request, pk):
        """Delete customer"""
        try:
            customer = Customer.objects.get(pk=pk)
            customer.delete()
            return Response(
                {'message': 'Customer deleted successfully'}, 
                status=status.HTTP_204_NO_CONTENT
            )
        except Customer.DoesNotExist:
            return Response(
                {'error': 'Customer not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )




##########################  PDF  #######################
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from io import BytesIO
import requests

from django.conf import settings
import os
import logging
from PyPDF2 import PdfReader, PdfWriter

logger = logging.getLogger(__name__)


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.grey)
        self.drawRightString(
            letter[0] - 0.5*inch,
            0.5*inch,
            f"Page {self._pageNumber} of {page_count}"
        )
        # Add generation timestamp
        self.drawString(
            0.5*inch,
            0.5*inch,
            f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        )
#import response
from django.http import HttpResponse



def export_sales_order_to_pdf(request, order_id):
    sales_order = get_object_or_404(SalesOrder, id=order_id)
    
    # 1. Setup the HTTP Response (Shared logic)
    response = HttpResponse(content_type='application/pdf')
    # Use the specific order number in the filename
    filename = f"Order_{sales_order.order_number}_{sales_order.order_date.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 2. Create Buffer
    buffer = BytesIO()

    # 3. DISPATCH: Choose the correct design function based on the order
    # (Assuming you added the 'division' field we discussed)
    if sales_order.division == 'ALABAMA':
        generate_alabama_pdf(buffer, sales_order)
    else:
        generate_junaid(buffer, sales_order)

    # 4. Finalize and Return
    pdf_value = buffer.getvalue()
    buffer.close()
    response.write(pdf_value)
    return response


from io import BytesIO
import os
import requests
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

def generate_junaid(buffer, sales_order):
    # 1. Get Items
    order_items = sales_order.items.all()
    
    # 2. Define Junaid Specific Settings (Logo & Header)
    current_logo_url = "https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp"
    local_logo_path = os.path.join(settings.BASE_DIR, 'static/images/footer-logo.png.webp')
    header_text = "CUSTOMER ORDER FORM"

    # 3. Create the PDF document using the passed buffer
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495E'),
        spaceAfter=12
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        fontName='Helvetica-Bold'
    )
    
    # Try to load logo
    try:
        response_img = requests.get(current_logo_url, timeout=5)
        if response_img.status_code == 200:
            logo = Image(BytesIO(response_img.content), width=150, height=50)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.3*inch))
    except Exception:
        # Fallback: load from local static file
        try:
            if os.path.exists(local_logo_path):
                logo = Image(local_logo_path, width=150, height=50)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 0.3 * inch))
        except Exception:
            pass 
    
    # Add title
    elements.append(Paragraph(header_text, title_style))
    
    # Order Information Section
    order_data = [
        ['ORDER INFORMATION', ''],
        ['Order Number:', f'{sales_order.order_number}'],
        ['Order Date:', sales_order.order_date.strftime('%d-%m-%Y')],
        ['Location:', getattr(sales_order, 'location', 'Not Specified')],
    ]
    
    order_table = Table(order_data, colWidths=[2.5*inch, 4*inch])
    order_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('SPAN', (0, 0), (-1, 0)),
        
        # Data rows
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#555555')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(order_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Customer Information Section
    customer_data = [
        ['CUSTOMER INFORMATION', ''],
        ['Customer Name:', sales_order.customer.customer_name],
        ['Customer Code:', sales_order.customer.customer_code]
    ]
    
    # Add salesman info if available
    if sales_order.salesman:
        customer_data.append(['Salesman:', sales_order.salesman.salesman_name])
    
    customer_table = Table(customer_data, colWidths=[2.5*inch, 4*inch])
    customer_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('SPAN', (0, 0), (-1, 0)),
        
        # Data rows
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#555555')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Items Table
    items_data = [
        ['S.No', 'Item Code', 'Description', 'Qty', 'Unit Price', 'Total']
    ]
    
    subtotal = 0.00
    for idx, item in enumerate(order_items, 1):
        line_total = item.quantity * item.price
        subtotal += line_total
        
        items_data.append([
            str(idx),
            item.item.item_code,
            Paragraph(item.item.item_description[:50] + '...' if len(item.item.item_description) > 50 else item.item.item_description, normal_style),
            f"{str(item.quantity)} {item.unit}",
            f"{item.price:,.2f} ",
            f"{line_total:,.2f} "
        ])
    
    # Create items table
    items_table = Table(
        items_data,
        colWidths=[0.5*inch, 1*inch, 2.5*inch, 0.7*inch, 1*inch, 1*inch]
    )
    
    items_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S.No
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Item Code
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Description
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Qty
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Unit Price
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),   # Total
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternate row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary Section
    summary_data = []
    
    # Subtotal
    summary_data.append(['', '', '', '', 'Subtotal:', f"{subtotal:,.2f} AED"])
    
    # Tax if applicable
    tax_amount = 0.00
    if hasattr(sales_order, 'tax') and sales_order.tax:
        tax_amount = sales_order.tax
        summary_data.append(['', '', '', '', f'VAT (5%):', f"{tax_amount:,.2f} AED"])
    
    # Discount if applicable
    if hasattr(sales_order, 'discount_amount') and sales_order.discount_amount:
        summary_data.append(['', '', '', '', 'Discount:', f"-{sales_order.discount_amount:,.2f} AED"])
    
    # Total
    total_amount = sales_order.total_amount
    grand_total = round(total_amount + tax_amount, 2)
    summary_data.append(['', '', '', '', 'Total Amount:      ', f"{grand_total:,.2f} AED"])
    
    summary_table = Table(
        summary_data,
        colWidths=[0.5*inch, 1*inch, 2.5*inch, 0.7*inch, 1*inch, 1*inch]
    )
    
    summary_table.setStyle(TableStyle([
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('FONTNAME', (4, 0), (5, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, 0), (5, -1), 10),
        
        # Total row styling
        ('FONTNAME', (4, -1), (5, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, -1), (5, -1), 12),
        ('TEXTCOLOR', (4, -1), (5, -1), colors.HexColor('#2C3E50')),
        ('LINEABOVE', (4, -1), (5, -1), 1.5, colors.HexColor('#2C3E50')),
        ('BACKGROUND', (4, -1), (5, -1), colors.HexColor('#E8F5E9')),
        ('TOPPADDING', (4, -1), (5, -1), 8),
        ('BOTTOMPADDING', (4, -1), (5, -1), 8),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Notes Section
    if hasattr(sales_order, 'notes') and sales_order.notes:
        elements.append(Paragraph("Notes:", heading_style))
        notes_style = ParagraphStyle(
            'NotesStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#555555'),
            leftIndent=20,
            rightIndent=20,
            borderColor=colors.HexColor('#CCCCCC'),
            borderWidth=1,
            borderPadding=10,
            backColor=colors.HexColor('#F8F9FA')
        )
        elements.append(Paragraph(sales_order.notes, notes_style))
        elements.append(Spacer(1, 0.3*inch))
    
    # Terms and Conditions
    terms_heading = Paragraph("System Generated - Terms & Conditions", heading_style)
    elements.append(terms_heading)
    
    terms_style = ParagraphStyle(
        'TermsStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        leftIndent=20
    )
    
    terms = [
        "1. This document is automatically generated as per prior sales agreement and does not require a physical signature.",
        "2. All disputes are subject to local jurisdiction only.",
        "3. Payment due as per customer account terms or prior agreement.",
    ]
    
    for term in terms:
        elements.append(Paragraph(term, terms_style))
    
    elements.append(Spacer(1, 0.5*inch))
    
    # Finalize PDF
    doc.build(elements)

def generate_alabama_pdf(buffer, sales_order):
    # 1. Get Items
    order_items = sales_order.items.all()
    
    # 2. Define Alabama Specific Settings (Red Theme)
    # Using the specific logo URL provided
    current_logo_url = "https://alabamauae.com/alabama4.png"
    local_logo_path = os.path.join(settings.BASE_DIR, 'static/images/alabama-logo.png')
    header_text = "CUSTOMER ORDER FORM"
    
    # Theme Colors
    THEME_RED = colors.HexColor("#211F1F")  # Dark Red for headers
    THEME_TEXT = colors.HexColor('#000000') # Black for standard text

    # 3. Create the PDF document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    # Custom styles - RED THEME
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=THEME_RED,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=THEME_RED,
        spaceAfter=12
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=THEME_TEXT,
        spaceAfter=6
    )
    
    label_style = ParagraphStyle(
        'LabelStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        fontName='Helvetica-Bold'
    )
    
    # Try to load logo
    try:
        response_img = requests.get(current_logo_url, timeout=5)
        if response_img.status_code == 200:
            logo = Image(BytesIO(response_img.content), width=150, height=50)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.3*inch))
    except Exception:
        # Fallback: load from local static file
        try:
            if os.path.exists(local_logo_path):
                logo = Image(local_logo_path, width=150, height=50)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 0.3 * inch))
        except Exception:
            pass 
    
    # Add title
    elements.append(Paragraph(header_text, title_style))
    
    # Order Information Section
    order_data = [
        ['ORDER INFORMATION', ''],
        ['Order Number:', f'{sales_order.order_number}'],
        ['Order Date:', sales_order.order_date.strftime('%d-%m-%Y')],
        ['Location:', getattr(sales_order, 'location', 'Not Specified')],
    ]
    
    order_table = Table(order_data, colWidths=[2.5*inch, 4*inch])
    order_table.setStyle(TableStyle([
        # Header row - RED Background
        ('BACKGROUND', (0, 0), (-1, 0), THEME_RED),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('SPAN', (0, 0), (-1, 0)),
        
        # Data rows
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#333333')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]), # Slight Red Tint
    ]))
    elements.append(order_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Customer Information Section
    customer_data = [
        ['CUSTOMER INFORMATION', ''],
        ['Customer Name:', sales_order.customer.customer_name],
    ]
    
    # Add salesman info if available
    if sales_order.salesman:
        customer_data.append(['Salesman:', sales_order.salesman.salesman_name])
    
    customer_table = Table(customer_data, colWidths=[2.5*inch, 4*inch])
    customer_table.setStyle(TableStyle([
        # Header row - RED Background
        ('BACKGROUND', (0, 0), (-1, 0), THEME_RED),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('SPAN', (0, 0), (-1, 0)),
        
        # Data rows
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#333333')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]), # Slight Red Tint
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Items Table
    items_data = [
        ['S.No', 'Item Code', 'Description', 'Qty', 'Unit Price', 'Total']
    ]
    
    subtotal = 0.00
    for idx, item in enumerate(order_items, 1):
        line_total = item.quantity * item.price
        subtotal += line_total
        
        items_data.append([
            str(idx),
            item.item.item_code,
            Paragraph(item.item.item_description[:50] + '...' if len(item.item.item_description) > 50 else item.item.item_description, normal_style),
            f"{str(item.quantity)} {item.unit}",
            f"{item.price:,.2f} ",
            f"{line_total:,.2f} "
        ])
    
    # Create items table
    items_table = Table(
        items_data,
        colWidths=[0.5*inch, 1*inch, 2.5*inch, 0.7*inch, 1*inch, 1*inch]
    )
    
    items_table.setStyle(TableStyle([
        # Header row - RED Background
        ('BACKGROUND', (0, 0), (-1, 0), THEME_RED),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S.No
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Item Code
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Description
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Qty
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Unit Price
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),   # Total
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternate row colors - White and Very Pale Red
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]),
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary Section
    summary_data = []
    
    # Subtotal
    summary_data.append(['', '', '', '', 'Subtotal:', f"{subtotal:,.2f} AED"])
    
    # Tax if applicable
    tax_amount = 0.00
    if hasattr(sales_order, 'tax') and sales_order.tax:
        tax_amount = sales_order.tax
        summary_data.append(['', '', '', '', f'VAT (5%):', f"{tax_amount:,.2f} AED"])
    
    # Discount if applicable
    if hasattr(sales_order, 'discount_amount') and sales_order.discount_amount:
        summary_data.append(['', '', '', '', 'Discount:', f"-{sales_order.discount_amount:,.2f} AED"])
    
    # Total
    total_amount = sales_order.total_amount
    grand_total = round(total_amount + tax_amount, 2)
    summary_data.append(['', '', '', '', 'Total Amount:      ', f"{grand_total:,.2f} AED"])
    
    summary_table = Table(
        summary_data,
        colWidths=[0.5*inch, 1*inch, 2.5*inch, 0.7*inch, 1*inch, 1*inch]
    )
    
    summary_table.setStyle(TableStyle([
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('FONTNAME', (4, 0), (5, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, 0), (5, -1), 10),
        
        # Total row styling - Red Line Above
        ('FONTNAME', (4, -1), (5, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, -1), (5, -1), 12),
        ('TEXTCOLOR', (4, -1), (5, -1), THEME_RED),
        ('LINEABOVE', (4, -1), (5, -1), 1.5, THEME_RED),
        ('BACKGROUND', (4, -1), (5, -1), colors.HexColor('#FFF0F0')),
        ('TOPPADDING', (4, -1), (5, -1), 8),
        ('BOTTOMPADDING', (4, -1), (5, -1), 8),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Notes Section
    if hasattr(sales_order, 'notes') and sales_order.notes:
        elements.append(Paragraph("Notes:", heading_style))
        notes_style = ParagraphStyle(
            'NotesStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#555555'),
            leftIndent=20,
            rightIndent=20,
            borderColor=colors.HexColor('#CCCCCC'),
            borderWidth=1,
            borderPadding=10,
            backColor=colors.HexColor('#FFF9F9')
        )
        elements.append(Paragraph(sales_order.notes, notes_style))
        elements.append(Spacer(1, 0.3*inch))
    
    # Terms and Conditions
    terms_heading = Paragraph("System Generated - Terms & Conditions", heading_style)
    elements.append(terms_heading)
    
    terms_style = ParagraphStyle(
        'TermsStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        leftIndent=20
    )
    
    terms = [
        "1. This document is automatically generated by Alabama Systems.",
        "2. All disputes are subject to local jurisdiction only.",
        "3. Payment due as per customer account terms or prior agreement.",
    ]
    
    for term in terms:
        elements.append(Paragraph(term, terms_style))
    
    elements.append(Spacer(1, 0.5*inch))
    
    # Finalize PDF
    doc.build(elements)



import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Customer
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def upload_customer_credit_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file , sheet_name = 'CREDIT')

            updated = 0
            for _, row in df.iterrows():
                code = str(row.get('Customer Code')).strip()
                if not code:
                    continue

                try:
                    customer = Customer.objects.get(customer_code=code)

                    # Clean and convert values
                    def safe_float(val):
                        try: return float(str(val).replace(",", "").strip())
                        except: return 0.0

                    customer.credit_days = row.get('CREDIT DAYS')
                    customer.credit_limit = safe_float(row.get('CREDIT LIMIT'))
                    customer.total_outstanding_with_pdc = safe_float(row.get('BALANCE'))
                    customer.pdc_received = safe_float(row.get('PDC RECEIVED'))
                    customer.total_outstanding = safe_float(row.get('TOTAL BALANCE'))
                    customer.month_pending_1 = safe_float(row.get('Month1'))
                    customer.month_pending_2 = safe_float(row.get('Month2'))
                    customer.month_pending_3 = safe_float(row.get('Month3'))
                    customer.month_pending_4 = safe_float(row.get('Month4'))
                    customer.month_pending_5 = safe_float(row.get('Month5'))
                    customer.month_pending_6 = safe_float(row.get('Month6'))
                    customer.old_months_pending = safe_float(row.get('6 + MONTHS '))

                    customer.save()
                    updated += 1
                except Customer.DoesNotExist:
                    print(f"Customer with code {code} not found.")
                    continue

            messages.success(request, f"Successfully updated {updated} customers.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        return redirect('upload_customer_credit_excel')

    return render(request, 'so/upload_credit_excel.html')


from django.db.models import Sum
from .models import Customer
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny

@api_view(['GET'])
@permission_classes([AllowAny])
def total_outstanding_sum(request):
    total = Customer.objects.aggregate(total_outstanding_sum=Sum('total_outstanding'))['total_outstanding_sum'] or 0.0
    return Response({'total_outstanding': round(total, 2)})


from django.http import JsonResponse
from django.db.models import Q
from so.models import Items


from django.http import JsonResponse

def get_last_location(request, customer_id):
    last_order = SalesOrder.objects.filter(customer_id=customer_id).order_by('-id').first()
    return JsonResponse({
        'location': last_order.location if last_order else ''
    })


from datetime import datetime
from datetime import timedelta
from decimal import Decimal

# =====================
# Quotation: Upload/List/Detail
# =====================
from django.db.models import Q
from django.http import Http404

# Map usernames -> the exact salesman_name values they are allowed to see.
# Use lowercase keys for usernames.
SALES_USER_MAP = {
    "muzain": ["B.MR.MUZAIN"],
    "dip": ["D.RETAIL CUST DIP"],
    "abubaqar": ["B. MR.RAFIQ ABU- PROJ","A.MR.RAFIQ ABU-TRD"],
    "rashid": ["A.MR.RASHID", "A.MR.RASHID CONT"],
    "parthiban": ["B.MR.PARTHIBAN"],
    "siyab": ["A.MR.SIYAB", "A.MR.SIYAB CONT"],
    "mr. nasheer": ["B.MR.NASHEER AHMAD"],
    "deira 2 store": ["R.DEIRA 2"],
    "rafiq": ["A.MR.RAFIQ"],
    "krishnan": ["I.KRISHNAN", "A.KRISHNAN"],  # combined both
    "alabama": ["D. ALABAMA"],     # both entries for Meraj
    "anish": ["ANISH DIP"],
    "musharaf": ["A.MUSHARAF"],
    "ibrahim": ["A.IBRAHIM"],
    "adil": ["A.DIP ADIL"],
    "kadar": ["A.DIP KADAR"],
    "stephy": ["A.DIP STEFFY"],
    "muzammil1": ["A.DIP MUZAMMIL" ,"A.DIP STEFFY","A.DIP KADAR","A.DIP ADIL","D.RETAIL CUST DIP" ],
    "retail": ["R.NAH","R.ABUDHABI" ,"R.AJMAN","R.QUSAIS","R.STORES","E.DEIRA 1","R.DEIRA 2"],
    "retailabudhabi": ["R.ABUDHABI"],
    "retailajman": ["R.AJMAN"],
    "retailqusais": ["R.QUSAIS"],
    "retailstores": ["R.STORES"],
    "exportdeira1": ["E.DEIRA 1"],
    "retaildeira2": ["R.DEIRA 2"],
    "retailnah": ["R.NAH"],
    
}

def salesman_scope_q(user: "User") -> Q:
    """Return a Q filter limiting SAPQuotation by salesman_name for non-staff users."""
    if user.is_superuser or user.role.role == "Admin":
        return Q()  # no restriction

    uname = (user.username or "").strip().lower()
    names = SALES_USER_MAP.get(uname)
    if names:
        q = Q(pk__isnull=False) & Q()  # start with something truthy
        q = Q()  # cleaner: start empty
        for n in names:
            q |= Q(salesman_name__iexact=n)
        return q

    # Sensible fallback if no explicit mapping:
    # match username token inside salesman_name (case-insensitive)
    token = uname.replace(".", " ").strip()
    if token:
        return Q(salesman_name__icontains=token)
    # If nothing to match, return an always-false Q to avoid leaking data
    return Q(pk__in=[])

@login_required
def upload_quotations(request):
    messages_list = []
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages_list.append('Please upload an Excel file.')
        else:
            try:
                df = pd.read_excel(excel_file)

                # Ensure expected columns exist
                required_cols = [
                    'Document Internal ID', 'Document Number', 'Posting Date',
                    'Customer/Supplier No.', 'Customer/Supplier Name', 'Sales Employee Name',
                    'Manufacturer Name', 'BP Reference No.', 'Item No.', 'Item/Service Description',
                    'Quantity', 'Price', 'Row Total', 'Document Total', 'Status','Bill To'
                ]
                missing = [c for c in required_cols if c not in df.columns]
                if missing:
                    messages_list.append(f"Missing columns: {', '.join(missing)}")
                else:
                    # Normalize numeric/text columns
                    def as_str(x):
                        try:
                            # preserve as string (e.g., to keep leading zeros)
                            return str(x).strip()
                        except Exception:
                            return ''

                    def to_decimal(x):
                        if pd.isna(x):
                            return None
                        try:
                            return Decimal(str(x).replace(',', '').strip())
                        except Exception:
                            return None

                    # Convert posting date
                    def parse_date(val):
                        if pd.isna(val):
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        s = str(val).strip()
                        for fmt in ["%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                            try:
                                return datetime.strptime(s, fmt).date()
                            except ValueError:
                                continue
                        return None

                    # Group by Document Number to create header + items
                    for doc_no, grp in df.groupby('Document Number'):
                        q_number = as_str(doc_no)
                        first = grp.iloc[0]

                        quotation, _ = SAPQuotation.objects.update_or_create(
                            q_number=q_number,
                            defaults={
                                'internal_number': as_str(first['Document Internal ID']),
                                'posting_date': parse_date(first['Posting Date']),
                                'customer_code': as_str(first['Customer/Supplier No.']),
                                'customer_name': as_str(first['Customer/Supplier Name']),
                                'salesman_name': as_str(first['Sales Employee Name']),
                                'brand': as_str(first['Manufacturer Name']),
                                'bp_reference_no': as_str(first['BP Reference No.']),
                                'document_total': to_decimal(first['Document Total']),
                                'status': as_str(first['Status']),
                                'bill_to': as_str(first['Bill To']),
                            }
                        )

                        # Refresh items: remove old, add new
                        quotation.items.all().delete()
                        items_to_create = []
                        for _, row in grp.iterrows():
                            items_to_create.append(SAPQuotationItem(
                                quotation=quotation,
                                item_no=as_str(row['Item No.']),
                                description=as_str(row['Item/Service Description']),
                                quantity=to_decimal(row['Quantity']) or Decimal('0'),
                                price=to_decimal(row['Price']) or Decimal('0'),
                                row_total=to_decimal(row['Row Total'])
                            ))
                        if items_to_create:
                            SAPQuotationItem.objects.bulk_create(items_to_create)

                    return redirect('quotation_list')

            except Exception as e:
                messages_list.append(f"Error processing Excel file: {str(e)}")

    return render(request, 'quotes/upload_quotations.html', {
        'messages': messages_list
    })


# =====================
# Quotation: List
# =====================
@login_required
def quotation_list(request):
    # Scope by logged-in user
    qs = SAPQuotation.objects.all().filter(salesman_scope_q(request.user))

    # Filters
    q = request.GET.get('q', '').strip()
    # salesman = request.GET.get('salesman', '').strip()
    # In quotation_list and quotation_search views:
    salesmen_filter = request.GET.getlist('salesman') # Gets ['Name1', 'Name2']

    # ✅ FIX: Apply List Filter
    if salesmen_filter:
        # Filter out empty strings
        clean_salesmen = [s for s in salesmen_filter if s.strip()]
        if clean_salesmen:
            qs = qs.filter(salesman_name__in=clean_salesmen)
    start = request.GET.get('start', '').strip()
    end = request.GET.get('end', '').strip()
    status = request.GET.get('status', '').strip()  # ✅ added
    total_range = request.GET.get('total', '').strip()     # ✅ NEW
    remarks_filter = request.GET.get('remarks', '').strip()  # ✅ NEW

    if total_range:
        if total_range == "0-5000":
            qs = qs.filter(document_total__gte=0, document_total__lte=5000)
        elif total_range == "5001-10000":
            qs = qs.filter(document_total__gte=5001, document_total__lte=10000)
        elif total_range == "10001-25000":
            qs = qs.filter(document_total__gte=10001, document_total__lte=25000)
        elif total_range == "25001-50000":
            qs = qs.filter(document_total__gte=25001, document_total__lte=50000)
        elif total_range == "50001-100000":
            qs = qs.filter(document_total__gte=50001, document_total__lte=100000)
        elif total_range == "100000+":
            qs = qs.filter(document_total__gt=100000)

    if remarks_filter == "YES":
        qs = qs.filter(remarks__isnull=False).exclude(remarks__exact="")
    elif remarks_filter == "NO":
        qs = qs.filter(Q(remarks__isnull=True) | Q(remarks__exact=""))



    if q:
        if q.isdigit():
            qs = qs.filter(q_number__istartswith=q)
        elif len(q) < 3:
            qs = qs.filter(
                Q(customer_name__istartswith=q) |
                Q(salesman_name__istartswith=q)
            )
        else:
            qs = qs.filter(
                Q(q_number__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(salesman_name__icontains=q)
            )



    # ✅ Status filter
    if status:
        qs = qs.filter(status__iexact=status)

    # Parse dates (YYYY-MM or YYYY-MM-DD)
    def parse_date(s):
        if not s:
            return None
        try:
            if len(s) == 7:  # YYYY-MM
                return datetime.strptime(s + '-01', '%Y-%m-%d').date()
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return None
    qs_for_years = qs.all() 
    start_date = parse_date(start)
    end_date = parse_date(end)
    if start_date:
        qs = qs.filter(posting_date__gte=start_date)
    if end_date:
        qs = qs.filter(posting_date__lte=end_date)


    # ---------------------------------------------------------
    # 6. CALCULATIONS
    # ---------------------------------------------------------
    
    # A. Calculate Grand Total from 'qs' (Respects Date + All Filters)
    grand_total_agg = qs.aggregate(
        total=Coalesce(Sum('document_total'), Value(0, output_field=DecimalField()))
    )
    total_value = grand_total_agg['total']

    # B. Calculate Years from 'qs_for_years' (Respects Salesman/Status, IGNORES Date)
    yearly_agg = qs_for_years.aggregate(
        total_2025=Coalesce(Sum('document_total', filter=Q(posting_date__year=2025)), Value(0, output_field=DecimalField())),
        total_2026=Coalesce(Sum('document_total', filter=Q(posting_date__year=2026)), Value(0, output_field=DecimalField())),
    )
    total_2025 = yearly_agg['total_2025']
    total_2026 = yearly_agg['total_2026']



    qs = qs.order_by('-posting_date', '-created_at')

    # Pagination
    try:
        page_size = int(request.GET.get('page_size', 100))
    except ValueError:
        page_size = 20
    page_size = max(5, min(page_size, 100))
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Distinct salesmen list (restricted to the same scope)
    salesmen = (
        SAPQuotation.objects.filter(salesman_scope_q(request.user))
        .exclude(salesman_name__isnull=True)
        .exclude(salesman_name='')
        .values_list('salesman_name', flat=True)
        .distinct()
        .order_by('salesman_name')
    )

    return render(request, 'quotes/quotation_list.html', {
        'page_obj': page_obj,
        'total_count': paginator.count,
        'total_2025': total_2025,      # ✅ send to template
        'total_2026': total_2026,      # ✅ send to template
        'salesmen': salesmen,
        'total_value': total_value,      # ✅ send to template
        'filters': {
            'q': q,
            'salesmen_filter': salesmen_filter,  # ✅ for multi-select
            'status': status,  # ✅ added back
            'start': start,
            'end': end,
            'page_size': page_size,
            'total': total_range,      # ✅ for keeping dropdown selected
            'remarks': remarks_filter, # ✅ for keeping dropdown selected
        }
    })

from django.db.models import Sum, Value, Q, DecimalField
from django.db.models.functions import Coalesce
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

@login_required
def export_quotation_list_pdf(request):
    """
    Exports the filtered list of quotations to a PDF report.
    Respects: q, salesman, start/end date, status, total range, remarks.
    """
    # 1. APPLY FILTERS (Exact copy from quotation_list)
    qs = SAPQuotation.objects.all().filter(salesman_scope_q(request.user))
    
    q = request.GET.get('q', '').strip()
    salesman = request.GET.get('salesman', '').strip()
    start = request.GET.get('start', '').strip()
    end = request.GET.get('end', '').strip()
    status = request.GET.get('status', '').strip()
    total_range = request.GET.get('total', '').strip()
    remarks_filter = request.GET.get('remarks', '').strip()

    # Apply Total Range Filter
    if total_range:
        if total_range == "0-5000": qs = qs.filter(document_total__gte=0, document_total__lte=5000)
        elif total_range == "5001-10000": qs = qs.filter(document_total__gte=5001, document_total__lte=10000)
        elif total_range == "10001-25000": qs = qs.filter(document_total__gte=10001, document_total__lte=25000)
        elif total_range == "25001-50000": qs = qs.filter(document_total__gte=25001, document_total__lte=50000)
        elif total_range == "50001-100000": qs = qs.filter(document_total__gte=50001, document_total__lte=100000)
        elif total_range == "100000+": qs = qs.filter(document_total__gt=100000)

    # Apply Remarks Filter
    if remarks_filter == "YES":
        qs = qs.filter(remarks__isnull=False).exclude(remarks__exact="")
    elif remarks_filter == "NO":
        qs = qs.filter(Q(remarks__isnull=True) | Q(remarks__exact=""))

    # Apply Search (q)
    if q:
        if q.isdigit():
            qs = qs.filter(q_number__istartswith=q)
        elif len(q) < 3:
            qs = qs.filter(Q(customer_name__istartswith=q) | Q(salesman_name__istartswith=q))
        else:
            qs = qs.filter(Q(q_number__icontains=q) | Q(customer_name__icontains=q) | Q(salesman_name__icontains=q))

    if salesman:
        qs = qs.filter(salesman_name__iexact=salesman)
    if status:
        qs = qs.filter(status__iexact=status)

    # Apply Dates
    def parse_date(s):
        if not s: return None
        try:
            if len(s) == 7: return datetime.strptime(s + '-01', '%Y-%m-%d').date()
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError: return None

    start_date = parse_date(start)
    end_date = parse_date(end)
    if start_date: qs = qs.filter(posting_date__gte=start_date)
    if end_date: qs = qs.filter(posting_date__lte=end_date)

    # Ordering
    qs = qs.order_by('-posting_date', '-created_at')

    # Calculate Total Value of Report
    total_value = qs.aggregate(
        total=Coalesce(Sum('document_total'), Value(0, output_field=DecimalField()))
    )['total']

    # --- 2. GENERATE PDF ---
    response = HttpResponse(content_type='application/pdf')
    filename = f"Quotation_Report_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Landscape A4 because lists are wide
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_text = "Quotation Sales Report"
    if start and end:
        title_text += f" ({start} to {end})"
    elements.append(Paragraph(title_text, styles['Title']))
    elements.append(Spacer(1, 20))

    # Table Header
    headers = ['Date', 'Quote #', 'Customer Name', 'Salesman', 'Status', 'Total (AED)']
    data = [headers]

    # Table Rows
    for item in qs:
        doc_total = item.document_total if item.document_total else 0
        date_str = item.posting_date.strftime('%Y-%m-%d') if item.posting_date else "-"
        
        row = [
            date_str,
            item.q_number,
            Paragraph(item.customer_name[:35] + '...' if len(item.customer_name or '') > 35 else (item.customer_name or ''), styles['Normal']),
            Paragraph(item.salesman_name or '-', styles['Normal']),
            item.status or '-',
            f"{doc_total:,.2f}"
        ]
        data.append(row)

    # Grand Total Row
    data.append(['', '', '', '', 'GRAND TOTAL:', f"{total_value:,.2f}"])

    # Table Styling
    # Calculate column widths (Landscape A4 width approx 840 points)
    col_widths = [70, 70, 280, 150, 80, 80]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5530')), # Header Color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        
        # Data Rows
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'), # Right align totals
        
        # Grand Total Row styling
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)

    # Footer/Summary
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total Records: {qs.count()}", styles['Normal']))

    doc.build(elements)
    return response

# =====================
# Quotation: Detail
# =====================
import requests
from django.core.cache import cache  # For caching API response
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.db.models import Q
# Import your models here

def get_stock_costs():
    """
    Fetches stock data from API and returns a dictionary: 
    {'item_code': cost_price_float}
    Cached for 1 hour to improve performance.
    """
    # Try to get data from cache first
    stock_data = cache.get('junaid_stock_data')
    
    if stock_data is None:
        try:
            url = "https://stock.junaidworld.com/api/stock"
            response = requests.get(url, timeout=5)
            response.raise_for_status()
            data = response.json()
            
            # Convert list to a dictionary for fast lookup: {'100003': 12.50}
            stock_data = {}
            for item in data:
                try:
                    # Key is item_code, Value is cost_price (converted to float)
                    code = str(item.get('item_code', '')).strip()
                    cost = float(item.get('cost_price', 0.0))
                    stock_data[code] = cost
                except (ValueError, TypeError):
                    continue
            
            # Save to cache for 60 minutes (3600 seconds)
            cache.set('junaid_stock_data', stock_data, 3600)
            
        except requests.RequestException:
            # If API fails, return empty dict so the page doesn't crash
            return {}
            
    return stock_data

@login_required
def quotation_detail(request, q_number):
    quotation = get_object_or_404(SAPQuotation, q_number=q_number)

    # Enforce scope for non-staff users
    if not (request.user.is_superuser or request.user.is_staff):
        allowed = SAPQuotation.objects.filter(
            Q(pk=quotation.pk) & salesman_scope_q(request.user)
        ).exists()
        if not allowed:
            raise Http404("Quotation not found")

    # Get items
    items = quotation.items.all().order_by('id')

    # --- NEW CALCULATION LOGIC ---
    stock_map = get_stock_costs()
    total_estimated_cost = 0.0

    # We iterate over items to calculate cost based on the API map
    for item in items:
        # Match item.item_no with the API's item_code
        item_code = str(item.item_no).strip()
        
        # Get unit cost from map, default to 0.0 if not found
        unit_cost = stock_map.get(item_code, 0.0)
        
        # Calculate row cost (Unit Cost * Quantity)
        # Convert Decimal quantity to float for calculation
        qty = float(item.quantity)
        total_estimated_cost += (unit_cost * qty)

    # Calculate Profit/Margin (Optional, but usually helpful)
    # Convert document_total to float for math
    doc_total = float(quotation.document_total or 0)
    total_profit = doc_total - total_estimated_cost

    # Check if user came from old PI list
    from_old_pi = request.GET.get('from') == 'old_pi'

    context = {
        'quotation': quotation,
        'items': items,
        'total_cost': total_estimated_cost, # Passed to template
        'total_profit': total_profit,       # Passed to template
        'from_old_pi': from_old_pi,         # Passed to template
    }

    return render(request, 'quotes/quotation_detail.html', context)

# =====================
# Quotation: AJAX Search (rows + pagination HTML)
# =====================
@login_required
def quotation_search(request):
    # Scope by logged-in user
    qs = SAPQuotation.objects.all().filter(salesman_scope_q(request.user))

    q = request.GET.get('q', '').strip()

    # In quotation_list and quotation_search views:
    salesmen_filter = request.GET.getlist('salesman') # Gets ['Name1', 'Name2']

    # 2. Logic
    if salesmen_filter:
         clean_salesmen = [s for s in salesmen_filter if s.strip()]
         if clean_salesmen:
             qs = qs.filter(salesman_name__in=clean_salesmen)

    start = request.GET.get('start', '').strip()
    end = request.GET.get('end', '').strip()
    status = request.GET.get('status', '').strip()
    total_range = request.GET.get('total', '').strip()   # ✅ NEW
    remarks_filter = request.GET.get('remarks', '').strip()

    # --- Existing filters ---
    if q:
        if q.isdigit():
            qs = qs.filter(q_number__istartswith=q)
        elif len(q) < 3:
            qs = qs.filter(
                Q(customer_name__istartswith=q) |
                Q(salesman_name__istartswith=q)
            )
        else:
            qs = qs.filter(
                Q(q_number__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(salesman_name__icontains=q)
            )



    if status:
        qs = qs.filter(status__iexact=status)

    # --- ✅ REMARKS FILTER ---
    if remarks_filter == "YES":
        qs = qs.filter(remarks__isnull=False).exclude(remarks__exact="")
    elif remarks_filter == "NO":
        qs = qs.filter(Q(remarks__isnull=True) | Q(remarks__exact=""))

    # --- DATE FILTER ---
    def parse_date(s):
        if not s:
            return None
        try:
            if len(s) == 7:
                return datetime.strptime(s + '-01', '%Y-%m-%d').date()
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return None

    start_date = parse_date(start)
    end_date = parse_date(end)

    if start_date:
        qs = qs.filter(posting_date__gte=start_date)
    if end_date:
        qs = qs.filter(posting_date__lte=end_date)

    # --- ✅ DOCUMENT TOTAL FILTER ---
    if total_range:
        if total_range == "0-5000":
            qs = qs.filter(document_total__gte=0, document_total__lte=5000)
        elif total_range == "5001-10000":
            qs = qs.filter(document_total__gte=5001, document_total__lte=10000)
        elif total_range == "10001-25000":
            qs = qs.filter(document_total__gte=10001, document_total__lte=25000)
        elif total_range == "25001-50000":
            qs = qs.filter(document_total__gte=25001, document_total__lte=50000)
        elif total_range == "50001-100000":
            qs = qs.filter(document_total__gte=50001, document_total__lte=100000)
        elif total_range == "100000+":
            qs = qs.filter(document_total__gt=100000)

    # --- ✅ TOTAL VALUE (sum of document_total on FILTERED qs) ---
    total_value = qs.aggregate(
        total=Coalesce(Sum('document_total'), Value(0, output_field=DecimalField()))
    )['total']

    # Order + Pagination (unchanged)
    qs = qs.order_by('-posting_date', '-created_at')

    try:
        page_size = int(request.GET.get('page_size', 20))
    except ValueError:
        page_size = 20
    page_size = max(5, min(page_size, 100))
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    rows_html = render_to_string('quotes/_quotation_rows.html', {
        'page_obj': page_obj
    }, request=request)

    pagination_html = render_to_string('quotes/_pagination.html', {
        'page_obj': page_obj
    }, request=request)

    return JsonResponse({
        'rows_html': rows_html,
        'pagination_html': pagination_html,
        'count': paginator.count,
        'total_value': float(total_value or 0),   # ✅ send to frontend
    })



# views.py
from io import BytesIO
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

# If you used the scoped access earlier, optionally re-check permission here too
from django.db.models import Q

# --- reportlab imports (same as in your other app) ---
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, white
from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, KeepTogether

# Reuse your template & styles (make sure these are importable)
# from .pdf import QuotationPDFTemplate, styles   # <-- adjust to your actual module
# If they're in the other app, import from there instead:
from .views_quotation import QuotationPDFTemplate, styles  # <- update path

from .models import SAPQuotation, SAPQuotationItem  # adjust if your items model path differs


@login_required
def export_sap_quotation_pdf(request, q_number):
    """
    Generate a PDF for SAPQuotation using the new Dynamic Template.
    """
    # Fetch quotation
    quotation = get_object_or_404(SAPQuotation, q_number=q_number)
    items_qs = quotation.items.all().order_by('id')

    # Prepare HTTP response
    response = HttpResponse(content_type='application/pdf')
    date_str = quotation.posting_date.strftime('%Y%m%d') if quotation.posting_date else 'NA'
    filename = f"SAP_Quotation_{quotation.q_number}_{date_str}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = BytesIO()

    # --- 1. DEFINE DEFAULT CONFIG (Junaid Settings) ---
    company_config = {
        'name': "Junaid Sanitary & Electrical Trading LLC",
        'address': "Dubai Investment Parks 2, Dubai, UAE",
        'contact': "Email: sales@junaid.ae | Phone: +97142367723",
        'logo_url': "https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp",
        'local_logo_path': os.path.join(settings.BASE_DIR, 'static', 'images', 'footer-logo.png.webp')
    }
    
    # Default Green Theme
    theme_config = {'primary': HexColor('#2C5530')}

    # --- 2. INITIALIZE TEMPLATE WITH CONFIG ---
    doc = QuotationPDFTemplate(
        buffer,
        company_config=company_config,  # <--- Passed here
        theme_config=theme_config,      # <--- Passed here
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=1.0*inch
    )

    elements = []

    # --- Title ---
    elements.append(Spacer(1, -1.3*inch))

    title_table = Table(
        [[Paragraph('QUOTATION', styles['MainTitle'])]],
        colWidths=[7.5*inch]
    )
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        # Set text color to Theme Primary
        ('TEXTCOLOR', (0, 0), (-1, -1), theme_config['primary']), 
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 0.1*inch))

    # --- Two-column info (Quotation / Customer) ---
    main_table_width = 7.2 * inch # Updated width to match new template logic

    quotation_data = [
        [Paragraph('Quotation Details', styles['SectionHeader'])],
        [Paragraph(f"<b>Number:</b> {quotation.q_number}", styles['Normal'])],
        [Paragraph(f"<b>Date:</b> {quotation.posting_date or '-'}", styles['Normal'])],
        [Paragraph(f"<b>BP Ref No:</b> {quotation.bp_reference_no or '—'}", styles['Normal'])],
    ]
    
    # Use theme color for background
    bg_color = theme_config['primary']
    
    quotation_info_table = Table(quotation_data, colWidths=[main_table_width / 2])
    quotation_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), bg_color), # Theme BG
        ('TEXTCOLOR', (0, 0), (0, 0), white),     # White text on header
    ]))

    customer_data = [
        [Paragraph('Customer Information', styles['SectionHeader'])],
        [Paragraph(f"<b>Name:</b> {quotation.customer_name or '—'}", styles['Normal'])],
        [Paragraph(f"<b>Code:</b> {quotation.customer_code or '—'}", styles['Normal'])],
        [Paragraph(f"<b>Salesman:</b> {quotation.salesman_name or '—'}", styles['Normal'])],
    ]
    
    customer_info_table = Table(customer_data, colWidths=[main_table_width / 2])
    customer_info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('BACKGROUND', (0, 0), (0, 0), bg_color), # Theme BG
        ('TEXTCOLOR', (0, 0), (0, 0), white),     # White text on header
    ]))

    info_table = Table([[quotation_info_table, customer_info_table]],
                       colWidths=[main_table_width / 2, main_table_width / 2])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.2 * inch))

    # --- Items table ---
    items_header = ['#', 'Item No.', 'Description', 'Qty', 'Unit Price', 'Total']
    items_data = [items_header]

    def _to_decimal(x):
        from decimal import Decimal
        if x is None: return Decimal('0')
        if isinstance(x, Decimal): return x
        try: return Decimal(str(x))
        except Exception: return Decimal('0')

    subtotal = Decimal('0')
    for idx, it in enumerate(items_qs, 1):
        qty = _to_decimal(it.quantity)
        price = _to_decimal(it.price)
        row_total = _to_decimal(it.row_total) if it.row_total is not None else (qty * price)
        subtotal += row_total

        # Use ItemDescription style
        desc_para = Paragraph(it.description or '—', styles['ItemDescription'])

        items_data.append([
            str(idx),
            it.item_no or '—',
            desc_para,
            f"{qty.normalize():f}".rstrip('0').rstrip('.') if qty else "0",
            f"AED {price:,.2f}",
            f"AED {row_total:,.2f}",
        ])

    items_table = Table(
        items_data,
        colWidths=[
            main_table_width * 0.05,   # #
            main_table_width * 0.15,   # Item No.
            main_table_width * 0.43,   # Description
            main_table_width * 0.07,   # Qty
            main_table_width * 0.15,   # Unit Price
            main_table_width * 0.15    # Total
        ],
        repeatRows=1
    )
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), bg_color), # Theme BG
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#F0F7F4'), white]),
        ('ALIGN', (0, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 0.1 * inch))

    # --- Summary (VAT 5%) ---
    tax_rate = Decimal('0.05')
    tax_amount = (subtotal * tax_rate).quantize(Decimal('0.01'))
    doc_total = _to_decimal(quotation.document_total)
    grand_total = doc_total if doc_total else (subtotal + tax_amount)

    summary_data = [
        ['Subtotal:', f"AED {subtotal:,.2f}"],
        [f'VAT ({(tax_rate*100):.0f}%):', f"AED {tax_amount:,.2f}"],
        ['Grand Total:', f"AED {grand_total:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[main_table_width * 0.5, main_table_width * 0.5])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#808080')),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('BACKGROUND', (0, 2), (-1, 2), bg_color), # Theme BG for Total
        ('TEXTCOLOR', (0, 2), (-1, 2), white),
    ]))
    
    summary_wrapper = Table([[summary_table]], colWidths=[main_table_width])
    summary_wrapper.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(KeepTogether(summary_wrapper))
    elements.append(Spacer(1, 0.3 * inch))

    # --- Optional: remarks / terms ---
    if getattr(quotation, 'remarks', None):
        elements.extend([
            Paragraph("Remarks:", styles['h3']),
            Paragraph(quotation.remarks, styles['Normal']),
            Spacer(1, 0.2 * inch)
        ])
        
    elements.extend([
        Paragraph("Terms & Conditions:", styles['h3']),
        Paragraph("1. This quotation is valid for 30 days from the date of issue.", styles['Normal']),
        Paragraph("2. Prices are subject to change after the validity period.", styles['Normal']),
        Paragraph("3. Delivery timelines to be confirmed upon order confirmation.", styles['Normal']),
        Paragraph("4. System-generated document.", styles['Normal']),
    ])

    # Build + return
    doc.multiBuild(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response



from django.views.decorators.http import require_POST
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.db.models import Q

@login_required
@require_POST
def quotation_update_remarks(request, q_number):
    quotation = get_object_or_404(SAPQuotation, q_number=q_number)

    # Enforce the same scope rules as detail view
    if not (request.user.is_superuser or request.user.is_staff):
        allowed = SAPQuotation.objects.filter(
            Q(pk=quotation.pk) & salesman_scope_q(request.user)
        ).exists()
        if not allowed:
            raise Http404("Quotation not found")

    # Update remarks
    new_remarks = (request.POST.get("remarks") or "").strip()
    quotation.remarks = new_remarks
    quotation.save(update_fields=["remarks"])

    messages.success(request, "Remarks updated.")
    return redirect("quotation_detail", q_number=quotation.q_number)




from django.http import JsonResponse
from django.db.models import Q

def items_search_api(request):
    """
    API endpoint for searching items with server-side processing for Select2.
    """
    search_term = request.GET.get('q', '')
    firm_filter = request.GET.get('firm', '') # Filter by brand/firm

    # Only start searching after a few characters are typed
    if len(search_term) < 2:
        return JsonResponse([], safe=False)

    # Base queryset
    items = Items.objects.all()

    # 1. Apply the firm filter if one is provided
    if firm_filter and firm_filter.lower() != 'all':
        items = items.filter(item_firm=firm_filter)

    # 2. Apply the search term across multiple relevant fields
    query = (
        Q(item_code__icontains=search_term) |
        Q(item_description__icontains=search_term)
    )
    items = items.filter(query)

    # 3. Limit the results to prevent sending too much data
    items = items[:50] # Return only the top 50 matches

    # 4. Format the data into the structure Select2 expects
    results = [
        {
            "id": item.id,
            "text": f"{item.item_description} ({item.item_code} - Stock: {item.item_stock})",
            # You can pass other data here if needed
            "stock": item.item_stock,
            "price": item.item_price 
        }
        for item in items
    ]

    return JsonResponse(results, safe=False)





# so/views.py

from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import Device
from .utils import get_client_ip

@login_required
@require_POST
def update_device_location(request):
    lat = request.POST.get("lat")
    lng = request.POST.get("lng")

    try:
        lat_val = float(lat)
        lng_val = float(lng)
    except (TypeError, ValueError):
        return JsonResponse({"status": "bad-coords"}, status=400)

    device = getattr(request, "device_obj", None)
    if not device:
        # fallback: try last device for this user
        device = Device.objects.filter(user=request.user).order_by('-last_seen').first()

    if not device:
        return JsonResponse({"status": "no-device"}, status=400)

    device.last_lat = lat_val
    device.last_lng = lng_val
    device.save(update_fields=["last_lat", "last_lng"])

    return JsonResponse({"status": "ok"})



import pandas as pd
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import OpenSalesOrder
from .forms import UploadFileForm

def upload_so_data(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            try:
                # 1. Read Excel using Pandas
                # dtype=str ensures Item Numbers like "00123" don't become 123
                df = pd.read_excel(excel_file, dtype={'Item No.': str, 'Document': str})
                
                # 2. Prepare objects list
                new_orders = []
                
                for _, row in df.iterrows():
                    # Date Parsing Logic (Handle DD.MM.YY)
                    p_date = None
                    raw_date = str(row.get('Posting Date', '')).strip()
                    if raw_date and raw_date != 'nan':
                        try:
                            # Try DD.MM.YY format first (as per your image)
                            p_date = datetime.strptime(raw_date, '%d.%m.%y').date()
                        except ValueError:
                            try:
                                # Fallback for standard Excel date format
                                p_date = pd.to_datetime(raw_date).date()
                            except:
                                p_date = None

                    # Create model instance
                    new_orders.append(OpenSalesOrder(
                        document_no=str(row.get('Document Number', '')),
                        posting_date=p_date,
                        bp_reference=str(row.get('BP Reference No.', '')),
                        customer_code=str(row.get('Customer/Supplier No.', '')),
                        customer_name=str(row.get('Customer/Supplier Name', '')),
                        item_no=str(row.get('Item No.', '')),
                        description=str(row.get('Item/Service Description', '')),
                        manufacturer=str(row.get('Manufacture', '')),
                        quantity=float(row.get('Quantity', 0) or 0),
                        row_total=float(row.get('Row Total', 0) or 0),
                        open_qty=float(row.get('Remaining Open Quantity', 0) or 0),
                        total_available=float(row.get('Total available Stock', 0) or 0),
                        salesman_name=str(row.get('Sales Employee', '')),
                        dip_stock=float(row.get('Dip warehouse stock', 0) or 0), 
                    ))

                # 3. Database Operation (Atomic Transaction)
                with transaction.atomic():
                    # Delete ALL existing records
                    OpenSalesOrder.objects.all().delete()
                    
                    # Insert NEW records
                    OpenSalesOrder.objects.bulk_create(new_orders)

                messages.success(request, f"Successfully imported {len(new_orders)} records.")
                return redirect('open_so_dashboard') # Redirect back to your dashboard

            except Exception as e:
                messages.error(request, f"Error importing file: {str(e)}")
                return redirect('upload_so_data')

    else:
        form = UploadFileForm()

    return render(request, 'so/upload_so.html', {'form': form})
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import OpenSalesOrder

# ReportLab Imports for PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO

# --- 2. HELPER FUNCTION (Filters Logic) ---
def _get_filtered_queryset(request):
    """
    Applies all filters (Permissions, HO/Others, Dropdowns, Months)
    and returns the QuerySet. Used by both Dashboard and PDF.
    """
    qs = OpenSalesOrder.objects.all().order_by('-posting_date')

    # A. PERMISSION CHECK
    try:
        user_role = request.user.role.role 
    except AttributeError:
        user_role = 'Salesman' 

    if user_role == 'Salesman':
        current_username = request.user.username.lower()
        if current_username in SALES_USER_MAP:
            allowed_names = SALES_USER_MAP[current_username]
            qs = qs.filter(salesman_name__in=allowed_names)
        else:
            qs = qs.none()

    # B. HO vs Others Logic
    show_ho_param = request.GET.get('show_ho')
    show_others_param = request.GET.get('show_others')
    
    show_ho = True if show_ho_param is None else (show_ho_param == 'true')
    show_others = True if show_others_param is None else (show_others_param == 'true')

    if show_ho and not show_others:
        qs = qs.filter(document_no__startswith='1')
    elif show_others and not show_ho:
        qs = qs.exclude(document_no__startswith='1')
    elif not show_ho and not show_others:
        qs = qs.none()

    # C. Multi-Select Filters
    salesmen = request.GET.getlist('salesman') 
    if salesmen:
        qs = qs.filter(salesman_name__in=salesmen)

    manufacturers = request.GET.getlist('manufacturer')
    if manufacturers:
        qs = qs.filter(manufacturer__in=manufacturers)

    items = request.GET.getlist('item')
    if items:
        qs = qs.filter(description__in=items)

    # --- ADD THIS BLOCK HERE ---
    customers = request.GET.getlist('customer')
    if customers:
        qs = qs.filter(customer_name__in=customers)
    # ---------------------------

    # D. Month Filter
    months = request.GET.getlist('month')
    if months:
        qs = qs.filter(posting_date__month__in=months)

    return qs

# --- 3. DASHBOARD VIEW ---
@login_required
def open_so_dashboard(request):
    # Call the helper
    qs = _get_filtered_queryset(request)

    # AJAX Response
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'so/partials/so_table_body.html', {'orders': qs})

    # Prepare Dropdowns (From the filtered QS as per your requirement)
    salesmen_list = qs.values_list('salesman_name', flat=True).distinct().order_by('salesman_name')
    manufacturers_list = qs.values_list('manufacturer', flat=True).distinct().order_by('manufacturer')
    items_list = qs.values_list('description', flat=True).distinct().order_by('description')
    customer_list = qs.values_list('customer_name', flat=True).distinct().order_by('customer_name')

    context = {
        'orders': qs,
        'salesmen': salesmen_list,
        'manufacturers': manufacturers_list,
        'items': items_list,
        'customers': customer_list,
    }
    return render(request, 'so/open_so_dashboard.html', context)

# --- 4. PDF EXPORT VIEW (ReportLab) ---
# --- 4. PDF EXPORT VIEW (ReportLab) ---
@login_required
def export_so_pdf(request):
    orders = _get_filtered_queryset(request)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()

    # --- LOGO & HEADER ---
    logo_url = "https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp"
    try:
        logo = Image(logo_url, width=150, height=50)
        logo.hAlign = 'LEFT'
        elements.append(logo)
        elements.append(Spacer(1, 10))
    except Exception:
        pass

    header_text = f"Open Sales Orders - {request.user.username}"
    elements.append(Paragraph(header_text, styles['Heading2']))
    elements.append(Spacer(1, 12))

    # --- TABLE HEADERS ---
    # 10 Columns total
    # 0:Date, 1:Doc, 2:LPO, 3:Cust, 4:Item, 5:Desc, 6:Qty, 7:Open, 8:Avail, 9:DIP
    data = [['Date', 'Doc No.', 'LPO', 'Customer', 'Item No', 'Description', 'Total SO', 'Open', 'Avail', 'DIP']]
    
    total_qty = 0
    total_open = 0
    
    normal_style = styles['Normal']
    normal_style.fontSize = 8 

    for obj in orders:
        p_date = obj.posting_date.strftime("%d/%m/%Y") if obj.posting_date else ""
        
        # --- FIX: WRAP TEXT FOR COLUMNS ---
        cust_cell = Paragraph(obj.customer_name[:35], normal_style)
        desc_cell = Paragraph(obj.description[:45], normal_style)
        
        # 1. Get string, 2. Wrap in Paragraph so it stays inside width 80
        lpo_text = obj.bp_reference if obj.bp_reference else "-"
        lpo_cell = Paragraph(lpo_text, normal_style) 
        
        qty = obj.quantity
        opn = obj.open_qty
        avl = obj.total_available
        dip = obj.dip_stock
        
        total_qty += qty
        total_open += opn
        
        data.append([
            p_date, 
            obj.document_no, 
            lpo_cell,       # <--- Insert the Paragraph object here
            cust_cell, 
            obj.item_no, 
            desc_cell, 
            f"{qty:,.0f}", 
            f"{opn:,.0f}",
            f"{avl:,.0f}",
            f"{dip:,.0f}"
        ])

    # --- FOOTER ROW ---
    # You added a column, so we need to add an extra empty string to keep alignment
    # Old logic: 'TOTALS:' was at index 4. Now it should be at index 5.
    data.append(['', '', '', '', '', 'TOTALS:', f"{total_qty:,.0f}", f"{total_open:,.0f}", '', ''])

    # --- COLUMN WIDTHS ---
    # Total Columns: 10
    col_widths = [60, 70, 80, 140, 65, 180, 50, 50, 50, 50] 
    
    t = Table(data, colWidths=col_widths, repeatRows=1)

    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),       
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),               
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),                   
        
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),      
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),                
        
        # --- ALIGNMENT UPDATES ---
        # Since we added LPO at index 2, the numbers shifted.
        # Numbers are now in columns 6 (Total SO), 7 (Open), 8 (Avail), 9 (DIP)
        ('ALIGN', (6, 1), (9, -1), 'RIGHT'),                
        
        # Footer Styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        # Align 'TOTALS:' text to the right
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'), 
    ]))

    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    
    unique_id = str(uuid.uuid4())[:4] 
    filename = f"sales_orders_{unique_id}.pdf"

    response = HttpResponse(buffer, content_type='application/pdf')
    # Using inline to open in new tab (as requested previously)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response







import uuid
from .models import TrustedDevice

# 1. View to show the warning page
# so/views.py
from django.shortcuts import render, redirect
from django.conf import settings
from .models import TrustedDevice
from .utils import send_telegram_message # Assuming you have this from your other code
import uuid

# 1. Register Page (Same as before)
@login_required
def register_device(request):
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')
    return render(request, 'so/register_device.html', {'user_agent': user_agent})

# 2. Logic to Request Approval (Modified)
@login_required
@require_POST
def approve_device(request):
    device_name = request.POST.get('device_name', 'Unknown Device')
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    ip_address = request.META.get('REMOTE_ADDR', '')
    new_token = str(uuid.uuid4())

    # Create the device, but is_approved=False by default (from model)
    TrustedDevice.objects.create(
        user=request.user,
        device_token=new_token,
        device_name=device_name,
        user_agent=user_agent,
        ip_address=ip_address,
        is_approved=False  # 🔒 Explicitly pending
    )

    # 🔔 Notify Admin via Telegram (Optional but recommended)
    try:
        msg = (
            f"⚠️ <b>New Device Request</b>\n"
            f"👤 User: {request.user.username}\n"
            f"💻 Device: {device_name}\n"
            f"🌐 IP: {ip_address}\n"
            f"Please check Django Admin to approve."
        )
        send_telegram_message(settings.TELEGRAM_CREATE_CHAT_ID, msg)
    except:
        pass

    # Set cookie so middleware knows this browser has *requested* access
    response = redirect('device_pending')
    response.set_cookie('trusted_device_token', new_token, max_age=315360000, httponly=True)
    
    return response

# 3. New "Waiting" Page
@login_required
def device_pending(request):
    # 1. Check the cookie
    token = request.COOKIES.get('trusted_device_token')
    
    if token:
        try:
            # 2. Check the Database Status
            device = TrustedDevice.objects.get(user=request.user, device_token=token)
            
            # 3. If Admin has approved it, redirect to Home immediately
            if device.is_approved:
                # Redirect based on your role logic
                if hasattr(request.user, 'role') and request.user.role.role == 'Salesman':
                    return redirect('sales_home')
                else:
                    return redirect('home')
                    
        except TrustedDevice.DoesNotExist:
            # If token exists in cookie but not in DB, force them to register again
            return redirect('register_device')

    # 4. If still not approved, show the pending page
    return render(request, 'so/device_pending.html')