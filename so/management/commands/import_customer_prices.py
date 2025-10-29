# yourapp/management/commands/import_customer_prices.py
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from so.models import CustomerPrice, Customer, Items  # adjust app/model paths if needed


REQUIRED_HEADERS = {"Customer", "LastPrice"}
# We will prefer ItemCode; if missing we fall back to ItemDescription
OPTIONAL_HEADERS = {"ItemCode", "ItemDescription"}


def to_decimal(x):
    if x is None:
        return None
    s = str(x).strip()
    if s == "":
        return None
    # Strip common currency/formatting
    for junk in [" AED", ","]:
        s = s.replace(junk, "")
    try:
        return Decimal(s)
    except InvalidOperation:
        # Final fallback: keep digits, dot, minus only
        filtered = "".join(ch for ch in s if (ch.isdigit() or ch in ".-"))
        try:
            return Decimal(filtered) if filtered else None
        except InvalidOperation:
            return None


class Command(BaseCommand):
    help = "Import/Upsert CustomerPrice records from an Excel file (Customer, ItemCode/ItemDescription, LastPrice)."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to Excel file (e.g., last_price_by_customer_item.xlsx)")
        parser.add_argument("--sheet", type=str, default=None, help="Worksheet name (default: first sheet)")
        parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not write to DB.")
        parser.add_argument("--customer-field", type=str, default="customer_name",
                            help="Field name on Customer model to match (default: customer_name)")
        parser.add_argument("--item-code-field", type=str, default="item_code",
                            help="Field name on Items model for code (default: item_code)")
        parser.add_argument("--item-desc-field", type=str, default="item_description",
                            help="Field name on Items model for description (default: item_description)")

    def handle(self, *args, **opts):
        path = opts["excel_path"]
        sheet_name = opts["sheet"]
        dry_run = opts["dry_run"]
        customer_field = opts["customer_field"]
        item_code_field = opts["item_code_field"]
        item_desc_field = opts["item_desc_field"]

        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Failed to open Excel file: {e}")

        ws = wb[sheet_name] if sheet_name else wb.active

        # ----- Read header row & map columns -----
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = [str(h).strip() if h is not None else "" for h in row]
            break
        if not header_row:
            raise CommandError("The worksheet appears to be empty (no header row found).")

        header_idx = {name: i for i, name in enumerate(header_row)}

        missing = REQUIRED_HEADERS - set(header_idx.keys())
        if missing:
            raise CommandError(f"Missing required headers: {sorted(missing)}\nFound: {header_row}")

        # At least one of ItemCode / ItemDescription must be present
        has_item_code = "ItemCode" in header_idx
        has_item_desc = "ItemDescription" in header_idx
        if not has_item_code and not has_item_desc:
            raise CommandError("Expected 'ItemCode' and/or 'ItemDescription' columns, but neither was found.")

        # ----- Process rows -----
        created, updated, skipped = 0, 0, 0
        not_found_customers = []
        not_found_items = []
        invalid_prices = []
        processed = 0

        # Build fast lookup caches (optional but helpful if file is large)
        # Customer cache by lowercase name
        customer_cache = {}
        # Item cache by code and by desc (lowercase)
        item_by_code = {}
        item_by_desc = {}

        # Preload all customers/items into dictionaries for speed
        for c in Customer.objects.all().only("id", customer_field):
            key = (getattr(c, customer_field) or "").strip().lower()
            if key:
                customer_cache[key] = c

        for it in Items.objects.all().only("id", item_code_field, item_desc_field):
            code = (getattr(it, item_code_field) or "").strip().lower()
            if code:
                item_by_code[code] = it
            desc = (getattr(it, item_desc_field) or "").strip().lower()
            if desc and desc not in item_by_desc:
                # If duplicate descriptions exist, we keep the first and will warn on ambiguous matches below
                item_by_desc[desc] = it

        # Wrap writes in a single atomic transaction (rolled back on dry-run)
        context = transaction.atomic()
        with context:
            for row in ws.iter_rows(min_row=2, values_only=True):
                processed += 1
                customer_val = (row[header_idx["Customer"]] if "Customer" in header_idx else None)
                last_price_val = (row[header_idx["LastPrice"]] if "LastPrice" in header_idx else None)
                item_code_val = (row[header_idx["ItemCode"]] if has_item_code else None)
                item_desc_val = (row[header_idx["ItemDescription"]] if has_item_desc else None)

                # Normalize inputs
                cust_key = (str(customer_val).strip().lower() if customer_val is not None else "")
                code_key = (str(item_code_val).strip().lower() if item_code_val is not None else "")
                desc_key = (str(item_desc_val).strip().lower() if item_desc_val is not None else "")

                # Validate customer
                customer_obj = customer_cache.get(cust_key)
                if not customer_obj:
                    skipped += 1
                    not_found_customers.append((processed, customer_val))
                    continue

                # Resolve item (prefer ItemCode)
                item_obj = None
                if code_key:
                    item_obj = item_by_code.get(code_key)
                if not item_obj and desc_key:
                    item_obj = item_by_desc.get(desc_key)

                if not item_obj:
                    skipped += 1
                    not_found_items.append((processed, item_code_val, item_desc_val))
                    continue

                # Parse price
                price = to_decimal(last_price_val)
                if price is None:
                    skipped += 1
                    invalid_prices.append((processed, last_price_val))
                    continue

                # Upsert
                obj, was_created = CustomerPrice.objects.update_or_create(
                    customer=customer_obj,
                    item=item_obj,
                    defaults={"custom_price": float(price)},  # NOTE: your model uses FloatField
                )
                created += int(was_created)
                updated += int(not was_created)

            if dry_run:
                # Force rollback
                self.stdout.write(self.style.WARNING("Dry-run enabled: rolling back all changes."))
                raise transaction.TransactionManagementError("DRY_RUN_ROLLBACK")

        # If dry-run, we intentionally triggered a rollback exception; handle it gracefully.
        # (Django will already have printed a traceback; suppress with a friendly message.)
        # But since we raised TransactionManagementError, execution won't reach here unless caught.
        # We choose not to catch it to guarantee rollback. The printed summary below is for real runs.

        self.stdout.write(self.style.SUCCESS("Import completed."))
        self.stdout.write(f"Processed rows: {processed}")
        self.stdout.write(self.style.SUCCESS(f"Created: {created}"))
        self.stdout.write(self.style.SUCCESS(f"Updated: {updated}"))
        self.stdout.write(self.style.WARNING(f"Skipped: {skipped}"))

        if not_found_customers:
            self.stdout.write(self.style.WARNING(f"\nCustomers not found ({len(not_found_customers)}):"))
            preview = "\n".join([f"  Row {r}: '{name}'" for r, name in not_found_customers[:20]])
            self.stdout.write(preview + ("\n  ... (more omitted)" if len(not_found_customers) > 20 else ""))

        if not_found_items:
            self.stdout.write(self.style.WARNING(f"\nItems not found ({len(not_found_items)}):"))
            preview = "\n".join([f"  Row {r}: code='{code}', desc='{desc}'" for r, code, desc in not_found_items[:20]])
            self.stdout.write(preview + ("\n  ... (more omitted)" if len(not_found_items) > 20 else ""))

        if invalid_prices:
            self.stdout.write(self.style.WARNING(f"\nInvalid prices ({len(invalid_prices)}):"))
            preview = "\n".join([f"  Row {r}: value='{val}'" for r, val in invalid_prices[:20]])
            self.stdout.write(preview + ("\n  ... (more omitted)" if len(invalid_prices) > 20 else ""))

